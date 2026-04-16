"""CSV/XLSX/JSON exchange helpers with deterministic schemas and reports."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import shutil
import sqlite3
import tempfile
from collections.abc import Iterator
from contextlib import contextmanager
from dataclasses import asdict
from datetime import datetime, timedelta
from datetime import time as dt_time
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from isrc_manager.code_registry import (
    BUILTIN_CATEGORY_CATALOG_NUMBER,
    BUILTIN_CATEGORY_CONTRACT_NUMBER,
    BUILTIN_CATEGORY_LICENSE_NUMBER,
    BUILTIN_CATEGORY_REGISTRY_SHA256_KEY,
    CLASSIFICATION_INTERNAL,
    CLASSIFICATION_MISMATCH,
    CodeRegistryService,
)
from isrc_manager.domain.codes import is_blank, to_compact_isrc, to_iso_isrc
from isrc_manager.domain.timecode import hms_to_seconds, parse_hms_text, seconds_to_hms
from isrc_manager.file_storage import coalesce_filename, infer_storage_mode
from isrc_manager.parties import PartyService
from isrc_manager.releases import ReleasePayload, ReleaseService, ReleaseTrackPlacement
from isrc_manager.services.custom_fields import CustomFieldDefinitionService
from isrc_manager.services.import_governance import GovernedImportCoordinator
from isrc_manager.services.import_repair_queue import TrackImportRepairQueueService
from isrc_manager.services.imports import XMLImportService
from isrc_manager.services.track_artist_sql import (
    track_additional_artists_expr,
    track_main_artist_join_sql,
)
from isrc_manager.services.tracks import TrackCreatePayload, TrackService, TrackUpdatePayload
from isrc_manager.works import WorkService

from .models import (
    ExchangeIdentifierClassificationOutcome,
    ExchangeIdentifierReviewRow,
    ExchangeImportOptions,
    ExchangeImportReport,
    ExchangeInspection,
)

JSON_SCHEMA_VERSION = 1
CSV_SNIFF_SAMPLE_SIZE = 4096
AUTO_CSV_DELIMITERS = ",;\t|"

_IDENTIFIER_FIELD_TO_SYSTEM_KEY = {
    "catalog_number": BUILTIN_CATEGORY_CATALOG_NUMBER,
    "release_catalog_number": BUILTIN_CATEGORY_CATALOG_NUMBER,
    "contract_number": BUILTIN_CATEGORY_CONTRACT_NUMBER,
    "license_number": BUILTIN_CATEGORY_LICENSE_NUMBER,
    "registry_sha256_key": BUILTIN_CATEGORY_REGISTRY_SHA256_KEY,
}
_UNBOUND_IDENTIFIER_IMPORT_FIELDS = frozenset(
    {
        "contract_number",
        "license_number",
        "registry_sha256_key",
    }
)
_SUPPORTED_IMPORT_ONLY_TARGETS = tuple(_UNBOUND_IDENTIFIER_IMPORT_FIELDS)


class ExchangeService:
    """Owns tabular and JSON import/export workflows."""

    _TITLE_NAME_IMPORT_TARGETS = frozenset(
        {
            "track_title",
            "album_title",
            "artist_name",
            "additional_artists",
            "release_title",
            "release_primary_artist",
            "release_album_artist",
        }
    )
    _TITLE_NAME_SMALL_WORDS = frozenset(
        {
            "a",
            "an",
            "and",
            "as",
            "at",
            "but",
            "by",
            "for",
            "in",
            "nor",
            "of",
            "on",
            "or",
            "the",
            "to",
            "via",
            "vs",
        }
    )
    _TITLE_NAME_TOKEN_RE = re.compile(r"[A-Za-z]+(?:'[A-Za-z]+)?")
    _TITLE_NAME_COMPACT_ACRONYM_RE = re.compile(r"(?<![A-Z])[A-Z]{1,2}(?:[&/][A-Z]{1,2})+(?![A-Z])")

    BASE_EXPORT_COLUMNS = (
        "track_id",
        "isrc",
        "track_title",
        "artist_name",
        "additional_artists",
        "album_title",
        "release_date",
        "track_length_sec",
        "track_length_hms",
        "iswc",
        "upc",
        "genre",
        "catalog_number",
        "buma_work_number",
        "composer",
        "publisher",
        "comments",
        "lyrics",
        "audio_file_path",
        "audio_file_storage_mode",
        "album_art_path",
        "album_art_storage_mode",
        "release_id",
        "release_title",
        "release_version_subtitle",
        "release_primary_artist",
        "release_album_artist",
        "release_type",
        "release_date_release",
        "release_original_release_date",
        "release_label",
        "release_sublabel",
        "release_catalog_number",
        "release_upc",
        "release_barcode_validation_status",
        "release_territory",
        "release_explicit_flag",
        "release_notes",
        "release_artwork_path",
        "release_artwork_storage_mode",
        "disc_number",
        "track_number",
        "sequence_number",
    )

    def __init__(
        self,
        conn: sqlite3.Connection,
        track_service: TrackService,
        release_service: ReleaseService,
        custom_fields: CustomFieldDefinitionService,
        data_root: str | Path | None = None,
        *,
        party_service: PartyService | None = None,
        work_service: WorkService | None = None,
        profile_name: str | None = None,
        repair_queue_service: TrackImportRepairQueueService | None = None,
    ):
        self.conn = conn
        self.track_service = track_service
        self.release_service = release_service
        self.custom_fields = custom_fields
        self.data_root = Path(data_root) if data_root is not None else None
        self.repair_queue_service = repair_queue_service or TrackImportRepairQueueService(conn)
        self.governed_imports = GovernedImportCoordinator(
            conn,
            track_service=track_service,
            party_service=party_service,
            work_service=work_service,
            profile_name=profile_name,
        )

    @staticmethod
    def _normalize_header_name(name: str) -> str:
        return str(name or "").strip().lower().replace(" ", "_")

    def _table_names(self) -> set[str]:
        return {
            str(row[0])
            for row in self.conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
            if row and row[0]
        }

    def _table_columns(self, table: str) -> set[str]:
        return {
            str(row[1])
            for row in self.conn.execute(f"PRAGMA table_info({table})").fetchall()
            if row and len(row) > 1 and row[1]
        }

    @staticmethod
    def _append_unique_name(target: dict[int, list[str]], key: int, value: object) -> None:
        clean = str(value or "").strip()
        if not clean:
            return
        bucket = target.setdefault(int(key), [])
        if clean not in bucket:
            bucket.append(clean)

    def _work_export_overrides(self, track_ids: list[int]) -> dict[int, dict[str, str]]:
        clean_track_ids = sorted({int(track_id) for track_id in track_ids if int(track_id) > 0})
        if not clean_track_ids:
            return {}
        if "Tracks" not in self._table_names():
            return {}
        if "work_id" not in self._table_columns("Tracks") or "Works" not in self._table_names():
            return {}

        placeholders = ",".join("?" for _ in clean_track_ids)
        track_rows = self.conn.execute(
            f"""
            SELECT
                t.id,
                t.work_id,
                COALESCE(w.iswc, '') AS work_iswc,
                COALESCE(w.registration_number, '') AS work_registration_number
            FROM Tracks t
            LEFT JOIN Works w ON w.id = t.work_id
            WHERE t.id IN ({placeholders})
            ORDER BY t.id
            """,
            clean_track_ids,
        ).fetchall()

        track_to_work: dict[int, int] = {}
        work_iswc_by_track: dict[int, str] = {}
        work_registration_by_track: dict[int, str] = {}
        work_ids: set[int] = set()
        for row in track_rows:
            track_id = int(row[0])
            work_id = int(row[1]) if row[1] is not None else 0
            if work_id > 0:
                track_to_work[track_id] = work_id
                work_ids.add(work_id)
                work_iswc_by_track[track_id] = str(row[2] or "").strip()
                work_registration_by_track[track_id] = str(row[3] or "").strip()

        if not work_ids:
            return {}

        work_placeholders = ",".join("?" for _ in work_ids)
        author_names: dict[int, list[str]] = {}
        publisher_names: dict[int, list[str]] = {}
        fallback_publisher_names: dict[int, list[str]] = {}

        if "WorkContributors" in self._table_names():
            for row in self.conn.execute(
                f"""
                SELECT
                    wc.work_id,
                    wc.role,
                    COALESCE(p.display_name, p.legal_name, wc.display_name, '') AS name
                FROM WorkContributors wc
                LEFT JOIN Parties p ON p.id = wc.party_id
                WHERE wc.work_id IN ({work_placeholders})
                ORDER BY wc.work_id, name, wc.id
                """,
                sorted(work_ids),
            ).fetchall():
                work_id = int(row[0])
                role = str(row[1] or "").strip().lower()
                name = row[2]
                if role in {"songwriter", "composer", "lyricist", "arranger", "adaptor"}:
                    self._append_unique_name(author_names, work_id, name)
                if role in {"publisher", "subpublisher"}:
                    self._append_unique_name(fallback_publisher_names, work_id, name)

        if "WorkOwnershipInterests" in self._table_names():
            for row in self.conn.execute(
                f"""
                SELECT
                    o.work_id,
                    COALESCE(p.display_name, p.legal_name, o.display_name, '') AS name
                FROM WorkOwnershipInterests o
                LEFT JOIN Parties p ON p.id = o.party_id
                WHERE o.work_id IN ({work_placeholders})
                  AND o.ownership_role IN ('publisher', 'subpublisher')
                ORDER BY o.work_id, name, o.id
                """,
                sorted(work_ids),
            ).fetchall():
                self._append_unique_name(publisher_names, int(row[0]), row[1])

        overrides: dict[int, dict[str, str]] = {}
        for track_id, work_id in track_to_work.items():
            resolved_publishers = (
                publisher_names.get(work_id) or fallback_publisher_names.get(work_id) or []
            )
            overrides[track_id] = {
                "iswc": work_iswc_by_track.get(track_id, ""),
                "buma_work_number": work_registration_by_track.get(track_id, ""),
                "composer": ", ".join(author_names.get(work_id, [])),
                "publisher": ", ".join(resolved_publishers),
            }
        return overrides

    def _apply_governed_work_import_override(
        self,
        payload_kwargs: dict[str, object],
        *,
        track_id: int,
        cache: dict[int, dict[str, str]],
    ) -> None:
        override = cache.get(int(track_id))
        if override is None:
            override = self._work_export_overrides([int(track_id)]).get(int(track_id), {})
            cache[int(track_id)] = override
        for field_name in ("iswc", "buma_work_number", "composer", "publisher"):
            clean_value = str(override.get(field_name) or "").strip()
            if clean_value:
                payload_kwargs[field_name] = clean_value

    def _release_columns_for_track_rows(self) -> dict[int, dict[str, object]]:
        release_rows = self.conn.execute(
            """
            SELECT
                rt.track_id,
                r.id,
                r.title,
                r.version_subtitle,
                r.primary_artist,
                r.album_artist,
                r.release_type,
                r.release_date,
                r.original_release_date,
                r.label,
                r.sublabel,
                r.catalog_number,
                r.upc,
                r.barcode_validation_status,
                r.territory,
                r.explicit_flag,
                r.release_notes,
                r.artwork_path,
                r.artwork_storage_mode,
                CASE WHEN r.artwork_blob IS NOT NULL THEN 1 ELSE 0 END,
                rt.disc_number,
                rt.track_number,
                rt.sequence_number
            FROM ReleaseTracks rt
            JOIN Releases r ON r.id = rt.release_id
            ORDER BY rt.track_id, rt.sequence_number, rt.release_id
            """
        ).fetchall()
        by_track: dict[int, list[dict[str, object]]] = {}
        for row in release_rows:
            by_track.setdefault(int(row[0]), []).append(
                {
                    "release_id": int(row[1]),
                    "release_title": row[2] or "",
                    "release_version_subtitle": row[3] or "",
                    "release_primary_artist": row[4] or "",
                    "release_album_artist": row[5] or "",
                    "release_type": row[6] or "",
                    "release_date_release": row[7] or "",
                    "release_original_release_date": row[8] or "",
                    "release_label": row[9] or "",
                    "release_sublabel": row[10] or "",
                    "release_catalog_number": row[11] or "",
                    "release_upc": row[12] or "",
                    "release_barcode_validation_status": row[13] or "",
                    "release_territory": row[14] or "",
                    "release_explicit_flag": int(row[15] or 0),
                    "release_notes": row[16] or "",
                    "release_artwork_path": row[17] or "",
                    "release_artwork_storage_mode": infer_storage_mode(
                        explicit_mode=row[18],
                        stored_path=row[17],
                        blob_value=b"\x00" if int(row[19] or 0) else None,
                    )
                    or "",
                    "disc_number": int(row[20] or 1),
                    "track_number": int(row[21] or 1),
                    "sequence_number": int(row[22] or 1),
                }
            )
        return by_track

    def _custom_field_maps(self) -> tuple[list[dict], dict[tuple[int, int], str]]:
        defs = self.custom_fields.list_active_fields()
        value_rows = self.conn.execute(
            """
            SELECT track_id, field_def_id, value
            FROM CustomFieldValues
            """
        ).fetchall()
        value_map = {
            (int(track_id), int(field_def_id)): ("" if value is None else str(value))
            for track_id, field_def_id, value in value_rows
        }
        return defs, value_map

    def _effective_track_media_paths(self, track_id: int) -> tuple[str, str]:
        audio_meta = self.track_service.get_media_meta(int(track_id), "audio_file")
        artwork_meta = self.track_service.get_media_meta(int(track_id), "album_art")
        return (
            str(audio_meta.get("path") or "").strip(),
            str(artwork_meta.get("path") or "").strip(),
        )

    @staticmethod
    def _synthetic_media_key(*parts: object, filename: str, default_stem: str) -> str:
        clean_parts = [str(part).strip().strip("/") for part in parts if str(part).strip()]
        clean_filename = coalesce_filename(filename, default_stem=default_stem)
        return "/".join([*clean_parts, clean_filename])

    def _resolve_packaged_media_source(self, stored_path: str) -> Path | None:
        clean_path = str(stored_path or "").strip()
        if not clean_path:
            return None
        path = Path(clean_path)
        if path.is_absolute():
            return path
        if self.data_root is None:
            return None
        return self.data_root / path

    @staticmethod
    def _package_media_arcname(stored_path: str) -> str:
        clean_path = str(stored_path or "").strip()
        path = Path(clean_path)
        if path.is_absolute():
            digest = hashlib.sha1(clean_path.encode("utf-8")).hexdigest()[:12]
            return f"media/external/{digest}_{path.name}"
        return f"media/{path.as_posix()}"

    @staticmethod
    def _append_unique_warning(
        warnings: list[str],
        seen_warnings: set[str],
        message: str,
    ) -> None:
        clean_message = str(message or "").strip()
        if not clean_message or clean_message in seen_warnings:
            return
        seen_warnings.add(clean_message)
        warnings.append(clean_message)

    def _package_track_media(
        self,
        archive: ZipFile,
        *,
        row: dict[str, object],
        field_name: str,
        track_id: int,
        media_key: str,
        written_media: set[str],
        packaged_media_index: dict[str, str],
        warnings: list[str],
        seen_warnings: set[str],
    ) -> None:
        meta = self.track_service.get_media_meta(int(track_id), media_key)
        if not bool(meta.get("has_media")):
            return
        storage_field = field_name.replace("_path", "_storage_mode")
        storage_mode = str(meta.get("storage_mode") or "").strip()
        package_key = str(meta.get("path") or "").strip()
        if not package_key:
            owner_scope = str(meta.get("owner_scope") or "track")
            owner_id = meta.get("owner_id") or track_id
            package_key = self._synthetic_media_key(
                "embedded",
                owner_scope,
                owner_id,
                media_key,
                filename=str(meta.get("filename") or ""),
                default_stem=media_key.replace("_", "-"),
            )
        arcname = self._package_media_arcname(package_key)
        if arcname not in written_media:
            stored_path = str(meta.get("path") or "").strip()
            try:
                if stored_path:
                    abs_path = self._resolve_packaged_media_source(stored_path)
                    if abs_path is not None and abs_path.exists():
                        archive.write(abs_path, arcname=arcname)
                    else:
                        data, _ = self.track_service.fetch_media_bytes(int(track_id), media_key)
                        archive.writestr(arcname, data)
                else:
                    data, _ = self.track_service.fetch_media_bytes(int(track_id), media_key)
                    archive.writestr(arcname, data)
            except FileNotFoundError:
                media_label = "audio file" if media_key == "audio_file" else "album art"
                missing_ref = stored_path or str(meta.get("filename") or "").strip() or media_label
                self._append_unique_warning(
                    warnings,
                    seen_warnings,
                    (
                        "Catalog exchange package omitted "
                        f"{media_label} for track {int(track_id)} because the stored media could "
                        f"not be found: {missing_ref}"
                    ),
                )
                row[field_name] = ""
                row[storage_field] = ""
                return
            written_media.add(arcname)
        packaged_media_index[package_key] = arcname
        row[field_name] = package_key
        row[storage_field] = storage_mode

    def _package_release_artwork(
        self,
        archive: ZipFile,
        *,
        row: dict[str, object],
        written_media: set[str],
        packaged_media_index: dict[str, str],
        warnings: list[str],
        seen_warnings: set[str],
    ) -> None:
        try:
            release_id = int(row.get("release_id") or 0)
        except Exception:
            release_id = 0
        if release_id <= 0:
            return
        release = self.release_service.fetch_release(release_id)
        if release is None:
            return
        storage_mode = str(release.artwork_storage_mode or "").strip()
        stored_path = str(release.artwork_path or "").strip()
        if not storage_mode and not stored_path:
            return
        package_key = stored_path
        if not package_key:
            package_key = self._synthetic_media_key(
                "embedded",
                "release",
                release_id,
                "artwork",
                filename=release.artwork_filename or "",
                default_stem="release-artwork",
            )
        arcname = self._package_media_arcname(package_key)
        if arcname not in written_media:
            try:
                if stored_path:
                    abs_path = self._resolve_packaged_media_source(stored_path)
                    if abs_path is not None and abs_path.exists():
                        archive.write(abs_path, arcname=arcname)
                    else:
                        data, _ = self.release_service.fetch_artwork_bytes(release_id)
                        archive.writestr(arcname, data)
                else:
                    data, _ = self.release_service.fetch_artwork_bytes(release_id)
                    archive.writestr(arcname, data)
            except FileNotFoundError:
                missing_ref = (
                    stored_path or str(release.artwork_filename or "").strip() or "artwork"
                )
                self._append_unique_warning(
                    warnings,
                    seen_warnings,
                    (
                        "Catalog exchange package omitted release artwork for "
                        f"release {release_id} because the stored media could not be found: "
                        f"{missing_ref}"
                    ),
                )
                row["release_artwork_path"] = ""
                row["release_artwork_storage_mode"] = ""
                return
            written_media.add(arcname)
        packaged_media_index[package_key] = arcname
        row["release_artwork_path"] = package_key
        row["release_artwork_storage_mode"] = storage_mode

    def export_rows(
        self, track_ids: list[int] | None = None, *, progress_callback=None
    ) -> tuple[list[str], list[dict[str, object]]]:
        self._report_progress(progress_callback, 5, "Collecting catalog export rows...")
        where_clause = ""
        params: list[object] = []
        if track_ids:
            placeholders = ",".join("?" * len(track_ids))
            where_clause = f"WHERE t.id IN ({placeholders})"
            params.extend(int(track_id) for track_id in track_ids)
        main_artist_join_sql, main_artist_name_expr = track_main_artist_join_sql(
            self.conn,
            track_alias="t",
            artist_alias="main_artist",
        )
        additional_artists_sql = track_additional_artists_expr(self.conn, track_id_expr="t.id")

        rows = self.conn.execute(
            f"""
            SELECT
                t.id,
                t.isrc,
                t.track_title,
                COALESCE({main_artist_name_expr}, '') AS artist_name,
                {additional_artists_sql} AS additional_artists,
                COALESCE(al.title, '') AS album_title,
                COALESCE(t.release_date, '') AS release_date,
                COALESCE(t.track_length_sec, 0) AS track_length_sec,
                COALESCE(t.iswc, '') AS iswc,
                COALESCE(t.upc, '') AS upc,
                COALESCE(t.genre, '') AS genre,
                COALESCE(t.catalog_number, '') AS catalog_number,
                COALESCE(t.buma_work_number, '') AS buma_work_number,
                COALESCE(t.composer, '') AS composer,
                COALESCE(t.publisher, '') AS publisher,
                COALESCE(t.comments, '') AS comments,
                COALESCE(t.lyrics, '') AS lyrics,
                COALESCE(t.audio_file_path, '') AS audio_file_path,
                COALESCE(t.album_art_path, '') AS album_art_path
            FROM Tracks t
            {main_artist_join_sql}
            LEFT JOIN Albums al ON al.id = t.album_id
            {where_clause}
            ORDER BY t.id
            """,
            params,
        ).fetchall()

        defs, custom_map = self._custom_field_maps()
        custom_headers = [
            f"custom::{field['name']}"
            for field in defs
            if field["field_type"] not in {"blob_image", "blob_audio"}
        ]
        headers = list(self.BASE_EXPORT_COLUMNS) + custom_headers
        release_map = self._release_columns_for_track_rows()
        work_overrides = self._work_export_overrides([int(row[0]) for row in rows])

        exported_rows: list[dict[str, object]] = []
        total_rows = max(len(rows), 1)
        for index, row in enumerate(rows, start=1):
            self._report_progress(
                progress_callback,
                10 + int(((index - 1) / total_rows) * 30),
                f"Preparing catalog export rows ({index} of {total_rows})...",
            )
            track_id = int(row[0])
            work_override = work_overrides.get(track_id, {})
            audio_meta = self.track_service.get_media_meta(track_id, "audio_file")
            artwork_meta = self.track_service.get_media_meta(track_id, "album_art")
            effective_audio_path = str(audio_meta.get("path") or "").strip()
            effective_album_art_path = str(artwork_meta.get("path") or "").strip()
            base = {
                "track_id": track_id,
                "isrc": row[1] or "",
                "track_title": row[2] or "",
                "artist_name": row[3] or "",
                "additional_artists": row[4] or "",
                "album_title": row[5] or "",
                "release_date": row[6] or "",
                "track_length_sec": int(row[7] or 0),
                "track_length_hms": seconds_to_hms(int(row[7] or 0)),
                "iswc": work_override.get("iswc") or row[8] or "",
                "upc": row[9] or "",
                "genre": row[10] or "",
                "catalog_number": row[11] or "",
                "buma_work_number": work_override.get("buma_work_number") or row[12] or "",
                "composer": work_override.get("composer") or row[13] or "",
                "publisher": work_override.get("publisher") or row[14] or "",
                "comments": row[15] or "",
                "lyrics": row[16] or "",
                "audio_file_path": effective_audio_path,
                "audio_file_storage_mode": str(audio_meta.get("storage_mode") or ""),
                "album_art_path": effective_album_art_path,
                "album_art_storage_mode": str(artwork_meta.get("storage_mode") or ""),
            }
            placements = release_map.get(track_id) or [
                {
                    "release_id": "",
                    "release_title": "",
                    "release_version_subtitle": "",
                    "release_primary_artist": "",
                    "release_album_artist": "",
                    "release_type": "",
                    "release_date_release": "",
                    "release_original_release_date": "",
                    "release_label": "",
                    "release_sublabel": "",
                    "release_catalog_number": "",
                    "release_upc": "",
                    "release_barcode_validation_status": "",
                    "release_territory": "",
                    "release_explicit_flag": 0,
                    "release_notes": "",
                    "release_artwork_path": "",
                    "release_artwork_storage_mode": "",
                    "disc_number": "",
                    "track_number": "",
                    "sequence_number": "",
                }
            ]
            for placement in placements:
                export_row = dict(base)
                export_row.update(placement)
                for field in defs:
                    if field["field_type"] in {"blob_image", "blob_audio"}:
                        continue
                    export_row[f"custom::{field['name']}"] = custom_map.get(
                        (track_id, int(field["id"])), ""
                    )
                exported_rows.append(export_row)
        self._report_progress(progress_callback, 40, "Catalog export rows prepared.")
        return headers, exported_rows

    def export_csv(
        self, path: str | Path, track_ids: list[int] | None = None, progress_callback=None
    ) -> int:
        headers, rows = self.export_rows(track_ids, progress_callback=progress_callback)
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._report_progress(progress_callback, 50, "Writing catalog CSV header...")
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            total_rows = max(len(rows), 1)
            for index, row in enumerate(rows, start=1):
                writer.writerow(row)
                self._report_progress(
                    progress_callback,
                    50 + int((index / total_rows) * 40),
                    f"Writing catalog CSV rows ({index} of {total_rows})...",
                )
        self._report_progress(progress_callback, 90, "Catalog CSV data written.")
        return len(rows)

    def export_xlsx(
        self, path: str | Path, track_ids: list[int] | None = None, progress_callback=None
    ) -> int:
        headers, rows = self.export_rows(track_ids, progress_callback=progress_callback)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CatalogExport"
        self._report_progress(progress_callback, 50, "Building catalog workbook header...")
        sheet.append(headers)
        total_rows = max(len(rows), 1)
        for index, row in enumerate(rows, start=1):
            sheet.append([row.get(header, "") for header in headers])
            self._report_progress(
                progress_callback,
                50 + int((index / total_rows) * 40),
                f"Writing catalog workbook rows ({index} of {total_rows})...",
            )
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        self._report_progress(progress_callback, 90, "Catalog workbook written.")
        return len(rows)

    def export_json(
        self, path: str | Path, track_ids: list[int] | None = None, progress_callback=None
    ) -> int:
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        headers, rows = self.export_rows(track_ids, progress_callback=progress_callback)
        self._report_progress(progress_callback, 50, "Serializing catalog JSON payload...")
        payload = {
            "schema_version": JSON_SCHEMA_VERSION,
            "exported_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "columns": headers,
            "rows": rows,
            "custom_field_defs": self.custom_fields.list_active_fields(),
        }
        self._report_progress(progress_callback, 80, "Writing catalog JSON file...")
        output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        self._report_progress(progress_callback, 90, "Catalog JSON data written.")
        return len(rows)

    def export_package(
        self, path: str | Path, track_ids: list[int] | None = None, progress_callback=None
    ) -> int:
        zip_path = Path(path)
        zip_path.parent.mkdir(parents=True, exist_ok=True)
        headers, rows = self.export_rows(track_ids, progress_callback=progress_callback)
        packaged_media_index: dict[str, str] = {}
        warnings: list[str] = []
        seen_warnings: set[str] = set()
        with ZipFile(zip_path, "w", compression=ZIP_DEFLATED) as archive:
            written_media: set[str] = set()
            total_rows = max(len(rows), 1)
            for index, row in enumerate(rows, start=1):
                self._report_progress(
                    progress_callback,
                    50 + int(((index - 1) / total_rows) * 35),
                    f"Packaging catalog media ({index} of {total_rows})...",
                )
                try:
                    track_id = int(row.get("track_id") or 0)
                except Exception:
                    track_id = 0
                if track_id > 0:
                    self._package_track_media(
                        archive,
                        row=row,
                        field_name="audio_file_path",
                        track_id=track_id,
                        media_key="audio_file",
                        written_media=written_media,
                        packaged_media_index=packaged_media_index,
                        warnings=warnings,
                        seen_warnings=seen_warnings,
                    )
                    self._package_track_media(
                        archive,
                        row=row,
                        field_name="album_art_path",
                        track_id=track_id,
                        media_key="album_art",
                        written_media=written_media,
                        packaged_media_index=packaged_media_index,
                        warnings=warnings,
                        seen_warnings=seen_warnings,
                    )
                self._package_release_artwork(
                    archive,
                    row=row,
                    written_media=written_media,
                    packaged_media_index=packaged_media_index,
                    warnings=warnings,
                    seen_warnings=seen_warnings,
                )
            payload = {
                "schema_version": JSON_SCHEMA_VERSION,
                "exported_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                "columns": headers,
                "rows": rows,
                "custom_field_defs": self.custom_fields.list_active_fields(),
                "packaged_media": True,
                "packaged_media_index": packaged_media_index,
                "warnings": warnings,
            }
            archive.writestr("manifest.json", json.dumps(payload, indent=2, ensure_ascii=False))
        self._report_progress(progress_callback, 90, "Catalog exchange package written.")
        return len(rows)

    def _load_json_payload(self, path: str | Path) -> dict[str, object]:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
        schema_version = int(payload.get("schema_version") or 0)
        if schema_version != JSON_SCHEMA_VERSION:
            raise ValueError(
                f"Unsupported JSON schema version {schema_version}. Expected {JSON_SCHEMA_VERSION}."
            )
        return payload

    def _load_package_payload(self, path: str | Path) -> dict[str, object]:
        with ZipFile(path, "r") as archive:
            try:
                manifest_bytes = archive.read("manifest.json")
            except KeyError as exc:
                raise ValueError("ZIP package does not contain manifest.json.") from exc
        payload = json.loads(manifest_bytes.decode("utf-8"))
        schema_version = int(payload.get("schema_version") or 0)
        if schema_version != JSON_SCHEMA_VERSION:
            raise ValueError(
                f"Unsupported package schema version {schema_version}. Expected {JSON_SCHEMA_VERSION}."
            )
        return payload

    @staticmethod
    def _safe_extract_zip(path: str | Path, target_dir: Path) -> None:
        target_root = target_dir.resolve()
        with ZipFile(path, "r") as archive:
            for member in archive.infolist():
                member_name = str(member.filename or "")
                destination = (target_root / member_name).resolve()
                try:
                    destination.relative_to(target_root)
                except ValueError as exc:
                    raise ValueError(f"ZIP package contains an unsafe path: {member_name}") from exc
                if member.is_dir():
                    destination.mkdir(parents=True, exist_ok=True)
                    continue
                destination.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(member, "r") as src, destination.open("wb") as dst:
                    shutil.copyfileobj(src, dst)

    @staticmethod
    def _normalize_package_media_key(value: object) -> str:
        return str(value or "").strip().replace("\\", "/")

    def _prepare_packaged_rows(
        self,
        payload: dict[str, object],
        *,
        extracted_root: Path,
    ) -> list[dict[str, object]]:
        packaged_media_index = payload.get("packaged_media_index")
        media_lookup: dict[str, str] = {}
        if isinstance(packaged_media_index, dict):
            for stored_path, arcname in packaged_media_index.items():
                key = self._normalize_package_media_key(stored_path)
                arc = str(arcname or "").strip()
                if key and arc:
                    media_lookup[key] = arc

        prepared_rows: list[dict[str, object]] = []
        for source_row in payload.get("rows") or []:
            row = dict(source_row)
            for field_name in ("audio_file_path", "album_art_path", "release_artwork_path"):
                raw_value = row.get(field_name)
                normalized_key = self._normalize_package_media_key(raw_value)
                if not normalized_key:
                    continue
                resolved_path = None
                if normalized_key in media_lookup:
                    candidate = extracted_root / media_lookup[normalized_key]
                    if candidate.exists():
                        resolved_path = candidate
                if resolved_path is None:
                    direct_candidate = extracted_root / normalized_key
                    if direct_candidate.exists():
                        resolved_path = direct_candidate
                if resolved_path is None:
                    media_candidate = extracted_root / "media" / normalized_key
                    if media_candidate.exists():
                        resolved_path = media_candidate
                if resolved_path is not None:
                    row[field_name] = str(resolved_path)
            prepared_rows.append(row)
        return prepared_rows

    @contextmanager
    def _open_csv_dict_reader(
        self, path: str | Path, *, delimiter: str | None = None
    ) -> Iterator[tuple[csv.DictReader, str]]:
        with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(CSV_SNIFF_SAMPLE_SIZE)
            handle.seek(0)
            dialect, resolved_delimiter = self._csv_dialect_for_sample(sample, delimiter=delimiter)
            if dialect is None:
                yield csv.DictReader(handle, delimiter=resolved_delimiter), resolved_delimiter
                return
            yield csv.DictReader(handle, dialect=dialect), resolved_delimiter

    @staticmethod
    def _validate_csv_delimiter(delimiter: str | None) -> str | None:
        if delimiter is None:
            return None
        clean = str(delimiter)
        if len(clean) != 1 or clean in {"\r", "\n"}:
            raise ValueError("CSV delimiter must be a single non-newline character.")
        return clean

    def _csv_dialect_for_sample(
        self, sample: str, *, delimiter: str | None = None
    ) -> tuple[type[csv.Dialect] | csv.Dialect | None, str]:
        explicit_delimiter = self._validate_csv_delimiter(delimiter)
        if explicit_delimiter is not None:
            return None, explicit_delimiter
        try:
            dialect = (
                csv.Sniffer().sniff(sample, delimiters=AUTO_CSV_DELIMITERS)
                if sample.strip()
                else csv.excel
            )
        except csv.Error:
            dialect = csv.excel
        return dialect, str(getattr(dialect, "delimiter", ",") or ",")

    def inspect_csv(
        self,
        path: str | Path,
        *,
        delimiter: str | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> ExchangeInspection:
        self._report_progress(progress_callback, 5, "Reading exchange CSV source...")
        if cancel_callback is not None:
            cancel_callback()
        with self._open_csv_dict_reader(path, delimiter=delimiter) as (reader, resolved_delimiter):
            headers = list(reader.fieldnames or [])
            preview_rows = []
            for _, row in zip(range(5), reader):
                preview_rows.append(dict(row))
        self._report_progress(progress_callback, 60, "Building exchange import preview...")
        return ExchangeInspection(
            file_path=str(path),
            format_name="csv",
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=self._suggest_mapping(headers),
            resolved_delimiter=resolved_delimiter,
            identifier_review_rows=self._build_identifier_review_rows(
                preview_rows,
                self._suggest_mapping(headers),
            ),
        )

    def inspect_xlsx(
        self, path: str | Path, *, progress_callback=None, cancel_callback=None
    ) -> ExchangeInspection:
        self._report_progress(progress_callback, 5, "Reading exchange workbook...")
        if cancel_callback is not None:
            cancel_callback()
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in (rows[0] if rows else ())]
        preview_rows = []
        for row in rows[1:6]:
            preview_rows.append({header: row[index] for index, header in enumerate(headers)})
        self._report_progress(progress_callback, 60, "Building exchange import preview...")
        return ExchangeInspection(
            file_path=str(path),
            format_name="xlsx",
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=self._suggest_mapping(headers),
            identifier_review_rows=self._build_identifier_review_rows(
                preview_rows,
                self._suggest_mapping(headers),
            ),
        )

    def inspect_json(
        self, path: str | Path, *, progress_callback=None, cancel_callback=None
    ) -> ExchangeInspection:
        self._report_progress(progress_callback, 5, "Reading exchange JSON source...")
        if cancel_callback is not None:
            cancel_callback()
        payload = self._load_json_payload(path)
        headers = list(payload.get("columns") or [])
        preview_rows = [dict(row) for row in list(payload.get("rows") or [])[:5]]
        warnings = []
        self._report_progress(progress_callback, 60, "Building exchange import preview...")
        return ExchangeInspection(
            file_path=str(path),
            format_name="json",
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=self._suggest_mapping(headers),
            warnings=warnings,
            identifier_review_rows=self._build_identifier_review_rows(
                preview_rows,
                self._suggest_mapping(headers),
            ),
        )

    def inspect_package(
        self, path: str | Path, *, progress_callback=None, cancel_callback=None
    ) -> ExchangeInspection:
        self._report_progress(progress_callback, 5, "Reading exchange package manifest...")
        if cancel_callback is not None:
            cancel_callback()
        payload = self._load_package_payload(path)
        headers = list(payload.get("columns") or [])
        preview_rows = [dict(row) for row in list(payload.get("rows") or [])[:5]]
        packaged_media_index = payload.get("packaged_media_index")
        media_count = len(packaged_media_index) if isinstance(packaged_media_index, dict) else 0
        warnings = []
        for warning in payload.get("warnings") or []:
            clean_warning = str(warning or "").strip()
            if clean_warning:
                warnings.append(clean_warning)
        if not bool(payload.get("packaged_media")):
            warnings.append("This ZIP does not advertise packaged media.")
        warnings.append(f"Packaged media entries detected: {media_count}")
        self._report_progress(progress_callback, 60, "Building exchange import preview...")
        return ExchangeInspection(
            file_path=str(path),
            format_name="package",
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=self._suggest_mapping(headers),
            warnings=warnings,
            identifier_review_rows=self._build_identifier_review_rows(
                preview_rows,
                self._suggest_mapping(headers),
            ),
        )

    def inspect_xml(self, path: str | Path) -> ExchangeInspection:
        xml_service = XMLImportService(
            self.conn,
            self.track_service,
            self.custom_fields,
            party_service=self.governed_imports.party_service,
            work_service=self.governed_imports.work_service,
            profile_name=self.governed_imports.profile_name,
            repair_queue_service=self.repair_queue_service,
        )
        _inspection, exchange_inspection = xml_service.build_exchange_inspection(str(path))
        return exchange_inspection

    def _suggest_mapping(self, headers: list[str]) -> dict[str, str]:
        supported = {
            self._normalize_header_name(name): name
            for name in self._supported_import_targets()
        }
        mapping: dict[str, str] = {}
        for header in headers:
            normalized = self._normalize_header_name(header)
            if normalized in supported:
                mapping[header] = supported[normalized]
            elif normalized.startswith("custom::") or normalized.startswith("custom__"):
                mapping[header] = header
        return mapping

    def _supported_import_targets(self) -> tuple[str, ...]:
        return (*self.BASE_EXPORT_COLUMNS, *_SUPPORTED_IMPORT_ONLY_TARGETS)

    @staticmethod
    def _identifier_system_key_for_field(field_name: str | None) -> str | None:
        return _IDENTIFIER_FIELD_TO_SYSTEM_KEY.get(str(field_name or "").strip())

    @staticmethod
    def _identifier_review_key(
        *,
        row_index: int,
        source_header: str,
        target_field_name: str,
        value: str,
    ) -> str:
        return "|".join(
            (
                str(int(row_index)),
                str(source_header or "").strip(),
                str(target_field_name or "").strip(),
                str(value or "").strip(),
            )
        )

    def _build_identifier_review_rows(
        self,
        preview_rows: list[dict[str, object]],
        mapping: dict[str, str],
    ) -> list[ExchangeIdentifierReviewRow]:
        review_rows: list[ExchangeIdentifierReviewRow] = []
        for row_index, row in enumerate(preview_rows, start=1):
            for source_header, target_name in mapping.items():
                clean_target = str(target_name or "").strip()
                if clean_target not in _UNBOUND_IDENTIFIER_IMPORT_FIELDS:
                    continue
                value = str(row.get(source_header) or "").strip()
                if not value:
                    continue
                category_system_key = self._identifier_system_key_for_field(clean_target)
                if not category_system_key:
                    continue
                review_rows.append(
                    ExchangeIdentifierReviewRow(
                        review_key=self._identifier_review_key(
                            row_index=row_index,
                            source_header=source_header,
                            target_field_name=clean_target,
                            value=value,
                        ),
                        row_index=int(row_index),
                        source_header=str(source_header),
                        target_field_name=clean_target,
                        suggested_category_system_key=category_system_key,
                        value=value,
                        reason=(
                            "These imported identifier values stay staged until apply. "
                            "Choose the external type Codespace should store."
                        ),
                    )
                )
        return review_rows

    def supported_import_targets(self) -> list[str]:
        targets: list[str] = []
        seen: set[str] = set()
        for name in self._supported_import_targets():
            if name in seen:
                continue
            seen.add(name)
            targets.append(name)
        for field in self.custom_fields.list_active_fields():
            if field.get("field_type") in {"blob_audio", "blob_image"}:
                continue
            target_name = f"custom::{field['name']}"
            if target_name in seen:
                continue
            seen.add(target_name)
            targets.append(target_name)
        return targets

    @staticmethod
    def _report_progress(
        progress_callback,
        value: int,
        message: str,
        *,
        maximum: int = 100,
    ) -> None:
        if progress_callback is None:
            return
        progress_callback(int(value), int(maximum), str(message or ""))

    @staticmethod
    def _classify_failure_category(message: str) -> str:
        clean = str(message or "").strip().lower()
        if any(
            token in clean
            for token in (
                "work",
                "govern",
                "isrc validation",
                "required",
                "not found",
                "unsupported",
                "conflict",
            )
        ):
            return "governance" if "work" in clean or "govern" in clean else "validation"
        return "validation"

    def _queue_failed_import_row(
        self,
        *,
        source_format: str,
        source_path: str | None,
        row_index: int,
        import_mode: str,
        normalized_row: dict[str, object],
        mapping: dict[str, str] | None,
        options: ExchangeImportOptions,
        failure_message: str,
        repair_entry_id: int | None = None,
    ) -> int:
        failure_category = self._classify_failure_category(failure_message)
        if repair_entry_id is not None:
            self.repair_queue_service.update_entry(
                int(repair_entry_id),
                normalized_row=normalized_row,
                failure_category=failure_category,
                failure_message=failure_message,
                mapping=mapping,
                options=asdict(options),
            )
            return int(repair_entry_id)
        return int(
            self.repair_queue_service.queue_failed_row(
                source_format=source_format,
                source_path=source_path,
                row_index=row_index,
                import_mode=options.mode,
                normalized_row=normalized_row,
                mapping=mapping,
                options=asdict(options),
                failure_category=failure_category,
                failure_message=failure_message,
            )
        )

    @contextmanager
    def _resolve_import_source_dir(
        self, *, format_name: str, source_path: str | Path | None
    ) -> Iterator[Path]:
        if str(format_name or "").strip().lower() == "package":
            if not source_path:
                with tempfile.TemporaryDirectory(prefix="exchange-package-repair-") as temp_dir:
                    yield Path(temp_dir)
                return
            with tempfile.TemporaryDirectory(prefix="exchange-package-repair-") as temp_dir:
                extracted_root = Path(temp_dir)
                self._safe_extract_zip(source_path, extracted_root)
                yield extracted_root
            return
        if source_path:
            yield Path(source_path).parent
            return
        yield Path.cwd()

    def import_csv(
        self,
        path: str | Path,
        *,
        mapping: dict[str, str] | None = None,
        options: ExchangeImportOptions | None = None,
        delimiter: str | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> ExchangeImportReport:
        path_obj = Path(path)
        self._report_progress(progress_callback, 5, "Reading source file...")
        if cancel_callback is not None:
            cancel_callback()
        with self._open_csv_dict_reader(path_obj, delimiter=delimiter) as (reader, _resolved):
            rows = [dict(row) for row in reader]
        self._report_progress(progress_callback, 20, "Parsing CSV rows...")
        return self.import_prepared_rows(
            rows,
            mapping=mapping,
            options=options,
            format_name="csv",
            source_path=path_obj,
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_xlsx(
        self,
        path: str | Path,
        *,
        mapping: dict[str, str] | None = None,
        options: ExchangeImportOptions | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> ExchangeImportReport:
        self._report_progress(progress_callback, 5, "Reading workbook...")
        if cancel_callback is not None:
            cancel_callback()
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in (values[0] if values else ())]
        rows = [{header: row[index] for index, header in enumerate(headers)} for row in values[1:]]
        self._report_progress(progress_callback, 20, "Parsing workbook rows...")
        return self.import_prepared_rows(
            rows,
            mapping=mapping,
            options=options,
            format_name="xlsx",
            source_path=path,
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_json(
        self,
        path: str | Path,
        *,
        mapping: dict[str, str] | None = None,
        options: ExchangeImportOptions | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> ExchangeImportReport:
        self._report_progress(progress_callback, 5, "Reading JSON source...")
        if cancel_callback is not None:
            cancel_callback()
        payload = self._load_json_payload(path)
        rows = [dict(row) for row in payload.get("rows") or []]
        self._report_progress(progress_callback, 20, "Parsing JSON rows...")
        return self.import_prepared_rows(
            rows,
            mapping=mapping,
            options=options,
            format_name="json",
            source_path=path,
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_package(
        self,
        path: str | Path,
        *,
        mapping: dict[str, str] | None = None,
        options: ExchangeImportOptions | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> ExchangeImportReport:
        self._report_progress(progress_callback, 5, "Reading package manifest...")
        if cancel_callback is not None:
            cancel_callback()
        payload = self._load_package_payload(path)
        self._report_progress(progress_callback, 15, "Extracting packaged media...")
        with tempfile.TemporaryDirectory(prefix="exchange-package-") as temp_dir:
            extracted_root = Path(temp_dir)
            self._safe_extract_zip(path, extracted_root)
            if cancel_callback is not None:
                cancel_callback()
            self._report_progress(progress_callback, 25, "Preparing package rows...")
            rows = self._prepare_packaged_rows(payload, extracted_root=extracted_root)
            return self._import_rows(
                rows,
                mapping=mapping,
                options=options,
                format_name="package",
                source_dir=extracted_root,
                source_path=str(path),
                progress_callback=progress_callback,
                cancel_callback=cancel_callback,
            )

    def import_xml(
        self,
        path: str | Path,
        *,
        mapping: dict[str, str] | None = None,
        options: ExchangeImportOptions | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> ExchangeImportReport:
        self._report_progress(progress_callback, 5, "Reading XML source...")
        if cancel_callback is not None:
            cancel_callback()
        xml_service = XMLImportService(
            self.conn,
            self.track_service,
            self.custom_fields,
            party_service=self.governed_imports.party_service,
            work_service=self.governed_imports.work_service,
            profile_name=self.governed_imports.profile_name,
            repair_queue_service=self.repair_queue_service,
        )
        inspection = xml_service.inspect_file(str(path))
        self._report_progress(progress_callback, 15, "Validating XML schema and fields...")
        rows = xml_service.exchange_rows_from_inspection(inspection)
        if inspection.conflicting_custom_fields:
            raise ValueError(
                "Custom field type conflicts were detected in the XML source: "
                + ", ".join(
                    f"{name} (xml={import_type}, profile={existing_type})"
                    for name, import_type, existing_type in inspection.conflicting_custom_fields
                )
            )
        opts = options or ExchangeImportOptions()
        if inspection.missing_custom_fields:
            if opts.create_missing_custom_fields:
                xml_service.ensure_missing_custom_fields(inspection)
            else:
                active_mapping = mapping or {
                    str(header): str(header)
                    for row in rows
                    for header in row
                    if str(header).startswith("custom::")
                }
                unresolved_targets = [
                    f"custom::{field_name}"
                    for field_name, _field_type in inspection.missing_custom_fields
                    if str(active_mapping.get(f"custom::{field_name}") or "").strip()
                    == f"custom::{field_name}"
                ]
                if unresolved_targets:
                    raise ValueError(
                        "This XML source references custom fields that are not in the current profile. "
                        "Enable 'Create missing custom fields' or skip those targets in the mapping first: "
                        + ", ".join(unresolved_targets)
                    )
        return self._import_rows(
            rows,
            mapping=mapping,
            options=opts,
            format_name="xml",
            source_dir=Path(path).parent,
            source_path=str(path),
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_prepared_rows(
        self,
        rows: list[dict[str, object]],
        *,
        mapping: dict[str, str] | None = None,
        options: ExchangeImportOptions | None = None,
        format_name: str,
        source_path: str | Path | None = None,
        progress_callback=None,
        cancel_callback=None,
        repair_entry_id: int | None = None,
        repair_override: dict[str, object] | None = None,
    ) -> ExchangeImportReport:
        with self._resolve_import_source_dir(
            format_name=format_name, source_path=source_path
        ) as source_dir:
            return self._import_rows(
                rows,
                mapping=mapping,
                options=options,
                format_name=format_name,
                source_dir=source_dir,
                source_path=str(source_path) if source_path is not None else None,
                progress_callback=progress_callback,
                cancel_callback=cancel_callback,
                repair_entry_id=repair_entry_id,
                repair_override=repair_override,
            )

    def _apply_mapping(
        self, row: dict[str, object], mapping: dict[str, str] | None
    ) -> dict[str, object]:
        if not mapping:
            return dict(row)
        normalized: dict[str, object] = {}
        for source_name, value in row.items():
            target_name = mapping.get(source_name)
            if not target_name:
                continue
            normalized[target_name] = value
        return normalized

    def _ensure_custom_headers(
        self, rows: list[dict[str, object]], *, create_missing: bool
    ) -> list[str]:
        custom_specs: list[dict[str, object]] = []
        seen_names: set[str] = set()
        existing_fields = {
            str(field["name"]): field for field in self.custom_fields.list_active_fields()
        }
        for row in rows:
            for key in row:
                if not str(key).startswith("custom::"):
                    continue
                field_name = str(key).split("::", 1)[1]
                if not field_name or field_name in seen_names:
                    continue
                seen_names.add(field_name)
                existing = existing_fields.get(field_name)
                if existing is not None:
                    continue
                custom_specs.append({"name": field_name, "field_type": "text"})
        if custom_specs and create_missing:
            self.custom_fields.ensure_fields(custom_specs)
        return [spec["name"] for spec in custom_specs]

    def _find_existing_track_id(
        self,
        row: dict[str, object],
        *,
        options: ExchangeImportOptions,
    ) -> int | None:
        if options.match_by_internal_id:
            try:
                track_id = int(row.get("track_id") or 0)
            except Exception:
                track_id = 0
            if track_id > 0:
                found = self.conn.execute(
                    "SELECT id FROM Tracks WHERE id=?", (track_id,)
                ).fetchone()
                if found:
                    return int(found[0])
        if options.match_by_isrc:
            compact = to_compact_isrc(str(row.get("isrc") or ""))
            if compact:
                found = self.conn.execute(
                    "SELECT id FROM Tracks WHERE isrc_compact=? ORDER BY id LIMIT 1",
                    (compact,),
                ).fetchone()
                if found:
                    return int(found[0])
        title = str(row.get("track_title") or "").strip()
        artist = str(row.get("artist_name") or "").strip()
        upc = str(row.get("upc") or row.get("release_upc") or "").strip()
        if options.match_by_upc_title and title and upc:
            found = self.conn.execute(
                """
                SELECT t.id
                FROM Tracks t
                WHERE lower(t.track_title)=lower(?) AND COALESCE(t.upc, '')=?
                ORDER BY t.id
                LIMIT 1
                """,
                (title, upc),
            ).fetchone()
            if found:
                return int(found[0])
        normalized_matches = (
            self._find_case_normalized_title_artist_matches(title, artist)
            if title and artist
            else []
        )
        if options.heuristic_match and normalized_matches:
            return normalized_matches[0]
        if options.mode == "merge" and len(normalized_matches) == 1:
            return normalized_matches[0]
        return None

    def _find_case_normalized_title_artist_matches(self, title: str, artist: str) -> list[int]:
        clean_title = str(title or "").strip()
        clean_artist = str(artist or "").strip()
        if not clean_title or not clean_artist:
            return []
        main_artist_join_sql, main_artist_name_expr = track_main_artist_join_sql(
            self.conn,
            track_alias="t",
            artist_alias="main_artist",
        )
        rows = self.conn.execute(
            f"""
            SELECT t.id
            FROM Tracks t
            {main_artist_join_sql}
            WHERE lower(t.track_title)=lower(?) AND lower({main_artist_name_expr})=lower(?)
            ORDER BY t.id
            """,
            (clean_title, clean_artist),
        ).fetchall()
        return [int(row[0]) for row in rows]

    @classmethod
    def _normalize_row_text_targets(cls, row: dict[str, object]) -> dict[str, object]:
        normalized: dict[str, object] = {}
        for key, value in row.items():
            normalized[key] = cls._normalize_text_target(str(key), value)
        return normalized

    @classmethod
    def _normalize_text_target(cls, target_name: str, raw_value: object) -> object:
        if target_name not in cls._TITLE_NAME_IMPORT_TARGETS or not isinstance(raw_value, str):
            return raw_value
        clean = raw_value.strip()
        if not clean:
            return clean
        if target_name == "additional_artists":
            return cls._normalize_additional_artists_target(clean)
        if not cls._should_normalize_title_name_text(clean):
            return clean
        return cls._to_display_title_name_case(clean)

    @staticmethod
    def _should_normalize_title_name_text(text: str) -> bool:
        has_alpha = False
        has_upper = False
        has_lower = False
        for char in str(text or ""):
            if not char.isalpha():
                continue
            has_alpha = True
            has_upper = has_upper or char.isupper()
            has_lower = has_lower or char.islower()
        return has_alpha and has_upper and not has_lower

    @classmethod
    def _normalize_additional_artists_target(cls, text: str) -> str:
        parts = [part.strip() for part in str(text or "").split(",")]
        normalized_parts: list[str] = []
        for part in parts:
            if not part:
                continue
            if cls._should_normalize_title_name_text(part):
                normalized_parts.append(cls._to_display_title_name_case(part))
            else:
                normalized_parts.append(part)
        return ", ".join(normalized_parts)

    @classmethod
    def _to_display_title_name_case(cls, text: str) -> str:
        clean = str(text or "").strip()
        if not clean:
            return clean
        lowered = clean.lower()
        titled = cls._TITLE_NAME_TOKEN_RE.sub(
            lambda match: cls._capitalize_title_name_token(match.group(0)),
            lowered,
        )
        normalized = cls._lowercase_middle_small_words(titled)
        return cls._restore_compact_acronym_spans(clean, normalized)

    @classmethod
    def _restore_compact_acronym_spans(cls, original: str, normalized: str) -> str:
        if not original or not normalized or len(original) != len(normalized):
            return normalized
        parts: list[str] = []
        last_end = 0
        has_match = False
        for match in cls._TITLE_NAME_COMPACT_ACRONYM_RE.finditer(original):
            has_match = True
            parts.append(normalized[last_end : match.start()])
            parts.append(match.group(0))
            last_end = match.end()
        if not has_match:
            return normalized
        parts.append(normalized[last_end:])
        return "".join(parts)

    @staticmethod
    def _capitalize_title_name_token(token: str) -> str:
        parts = str(token or "").split("'")
        if not parts:
            return ""
        normalized_parts: list[str] = []
        previous_length = 0
        for index, part in enumerate(parts):
            if not part:
                normalized_parts.append(part)
            elif index == 0 or previous_length == 1:
                normalized_parts.append(part[0].upper() + part[1:])
            else:
                normalized_parts.append(part)
            previous_length = len(part)
        return "'".join(normalized_parts)

    @classmethod
    def _lowercase_middle_small_words(cls, text: str) -> str:
        matches = list(cls._TITLE_NAME_TOKEN_RE.finditer(text))
        if len(matches) < 3:
            return text
        parts: list[str] = []
        last_end = 0
        last_index = len(matches) - 1
        for index, match in enumerate(matches):
            parts.append(text[last_end : match.start()])
            token = match.group(0)
            if 0 < index < last_index and token.lower() in cls._TITLE_NAME_SMALL_WORDS:
                parts.append(token.lower())
            else:
                parts.append(token)
            last_end = match.end()
        parts.append(text[last_end:])
        return "".join(parts)

    @staticmethod
    def _normalize_track_length_target(raw_value: object) -> object:
        if isinstance(raw_value, timedelta):
            return max(0, int(raw_value.total_seconds()))
        if isinstance(raw_value, dt_time):
            return hms_to_seconds(raw_value.hour, raw_value.minute, raw_value.second)
        if isinstance(raw_value, float) and math.isfinite(raw_value):
            return int(raw_value)
        if isinstance(raw_value, str):
            text = raw_value.strip()
            if not text:
                return raw_value
            parts = text.split(":")
            if len(parts) != 3:
                return raw_value
            if not all(part.isdigit() for part in parts):
                return raw_value
            if len(parts[1]) != 2 or len(parts[2]) != 2:
                return raw_value
            minutes = int(parts[1])
            seconds = int(parts[2])
            if minutes > 59 or seconds > 59:
                return raw_value
            return parse_hms_text(text)
        return raw_value

    def _resolve_media_path(self, source_dir: Path, raw_value: object) -> str | None:
        text = str(raw_value or "").strip()
        if not text:
            return None
        path = Path(text)
        if not path.is_absolute():
            path = source_dir / path
        if path.exists():
            return str(path)
        return None

    def _code_registry_service(self) -> CodeRegistryService | None:
        tables = {
            str(row[0])
            for row in self.conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
            if row and row[0]
        }
        if "CodeRegistryCategories" not in tables:
            return None
        return CodeRegistryService(self.conn)

    @staticmethod
    def _empty_identifier_totals() -> dict[str, dict[str, int]]:
        return {}

    @staticmethod
    def _increment_identifier_total(
        totals: dict[str, dict[str, int]],
        *,
        category_system_key: str,
        bucket: str,
    ) -> None:
        category_key = str(category_system_key or "").strip()
        if not category_key:
            return
        category_totals = totals.setdefault(category_key, {})
        category_totals[str(bucket or "").strip()] = int(category_totals.get(bucket) or 0) + 1

    def _record_identifier_classification(
        self,
        *,
        row_index: int,
        field_name: str,
        category_system_key: str,
        value: str | None,
        outcomes: list[ExchangeIdentifierClassificationOutcome],
        totals: dict[str, dict[str, int]],
        outcome_override: str | None = None,
    ) -> ExchangeIdentifierClassificationOutcome | None:
        clean_value = str(value or "").strip()
        if not clean_value:
            return None
        service = self._code_registry_service()
        classification = (
            service.classify_identifier_value(
                system_key=category_system_key,
                value=clean_value,
                allow_existing_internal_match=False,
            )
            if service is not None
            else None
        )
        classification_name = (
            str(classification.classification or "").strip()
            if classification is not None
            else "external"
        )
        category_label = (
            str(classification.category_display_name or "").strip()
            if classification is not None and classification.category_display_name
            else None
        )
        reason = str(classification.reason or "").strip() if classification is not None else None
        outcome = outcome_override or (
            "accepted_internal"
            if classification_name == CLASSIFICATION_INTERNAL
            else (
                "flagged_mismatch"
                if classification_name == CLASSIFICATION_MISMATCH
                else "stored_external"
            )
        )
        if outcome == "accepted_internal":
            self._increment_identifier_total(
                totals,
                category_system_key=category_system_key,
                bucket="internal",
            )
        elif outcome == "stored_external":
            self._increment_identifier_total(
                totals,
                category_system_key=category_system_key,
                bucket="external",
            )
        elif outcome == "flagged_mismatch":
            self._increment_identifier_total(
                totals,
                category_system_key=category_system_key,
                bucket="mismatch",
            )
        elif outcome.startswith("skipped"):
            self._increment_identifier_total(
                totals,
                category_system_key=category_system_key,
                bucket="skipped",
            )
        elif outcome.startswith("merged"):
            self._increment_identifier_total(
                totals,
                category_system_key=category_system_key,
                bucket="merged",
            )
        elif outcome.startswith("conflicted"):
            self._increment_identifier_total(
                totals,
                category_system_key=category_system_key,
                bucket="conflicted",
            )
        outcome_row = ExchangeIdentifierClassificationOutcome(
            row_index=int(row_index),
            field_name=str(field_name or "").strip(),
            category_system_key=str(category_system_key or "").strip(),
            value=clean_value,
            classification=classification_name or "external",
            outcome=outcome,
            category_label=category_label,
            reason=reason,
        )
        outcomes.append(outcome_row)
        return outcome_row

    def _record_catalog_classification(
        self,
        *,
        row_index: int,
        field_name: str,
        value: str | None,
        outcomes: list[ExchangeIdentifierClassificationOutcome],
        counters: dict[str, dict[str, int]],
        outcome_override: str | None = None,
    ) -> ExchangeIdentifierClassificationOutcome | None:
        return self._record_identifier_classification(
            row_index=row_index,
            field_name=field_name,
            category_system_key=BUILTIN_CATEGORY_CATALOG_NUMBER,
            value=value,
            outcomes=outcomes,
            totals=counters,
            outcome_override=outcome_override,
        )

    def _store_unbound_external_identifier(
        self,
        *,
        row_index: int,
        source_header: str,
        field_name: str,
        value: str | None,
        options: ExchangeImportOptions,
        outcomes: list[ExchangeIdentifierClassificationOutcome],
        totals: dict[str, dict[str, int]],
        cursor: sqlite3.Cursor | None,
    ) -> None:
        clean_value = str(value or "").strip()
        if not clean_value:
            return
        suggested_system_key = self._identifier_system_key_for_field(field_name)
        if suggested_system_key is None:
            return
        review_key = self._identifier_review_key(
            row_index=row_index,
            source_header=source_header,
            target_field_name=field_name,
            value=clean_value,
        )
        system_key = str(
            options.identifier_overrides.get(review_key) or suggested_system_key
        ).strip()
        service = self._code_registry_service()
        classification = (
            service.classify_identifier_value(
                system_key=system_key,
                value=clean_value,
                allow_existing_internal_match=False,
            )
            if service is not None
            else None
        )
        category_label = (
            str(classification.category_display_name or "").strip()
            if classification is not None and classification.category_display_name
            else None
        )
        reason = (
            (
                service._classification_reason_for_external_storage(
                    system_key=system_key,
                    classification=classification,
                )
                if service is not None and classification is not None
                else "Stored in External Identifiers from exchange import."
            )
            or "Stored in External Identifiers from exchange import."
        )
        outcome = ExchangeIdentifierClassificationOutcome(
            row_index=int(row_index),
            field_name=str(field_name or "").strip(),
            category_system_key=system_key,
            value=clean_value,
            classification=(
                str(classification.classification or "").strip()
                if classification is not None
                else "external"
            ),
            outcome="stored_external",
            category_label=category_label,
            reason=reason,
        )
        outcomes.append(outcome)
        self._increment_identifier_total(
            totals,
            category_system_key=system_key,
            bucket="external",
        )
        if options.mode == "dry_run" or service is None:
            return
        service.store_external_code_identifier(
            system_key=system_key,
            value=clean_value,
            provenance_kind="import",
            source_label="exchange.import",
            owner_kind="import",
            owner_id=0,
            cursor=cursor,
        )

    def _upsert_release_from_row(
        self,
        row: dict[str, object],
        track_id: int,
        *,
        source_dir: Path,
        source_release_map: dict[str, int] | None = None,
        preserve_source_release_identity: bool = False,
        cursor: sqlite3.Cursor | None = None,
        catalog_outcomes: list[ExchangeIdentifierClassificationOutcome] | None = None,
        catalog_counters: dict[str, dict[str, int]] | None = None,
        row_index: int | None = None,
    ) -> None:
        release_title = str(row.get("release_title") or "").strip()
        if not release_title:
            return
        source_release_key = str(row.get("release_id") or "").strip()
        existing_id = None
        if source_release_map is not None and source_release_key:
            existing_id = source_release_map.get(source_release_key)
        elif source_release_key:
            release_id = int(source_release_key)
            release = self.release_service.fetch_release(release_id)
            existing_id = release.id if release is not None else None
        if existing_id is None and not (preserve_source_release_identity and source_release_key):
            release_upc = str(row.get("release_upc") or "").strip()
            release_catalog = str(row.get("release_catalog_number") or "").strip()
            if release_upc:
                found = self.conn.execute(
                    "SELECT id FROM Releases WHERE upc=? ORDER BY id LIMIT 1",
                    (release_upc,),
                ).fetchone()
                if found:
                    existing_id = int(found[0])
            if existing_id is None and release_catalog and release_title:
                found = self.conn.execute(
                    "SELECT id FROM Releases WHERE title=? AND catalog_number=? ORDER BY id LIMIT 1",
                    (release_title, release_catalog),
                ).fetchone()
                if found:
                    existing_id = int(found[0])

        placement = ReleaseTrackPlacement(
            track_id=track_id,
            disc_number=max(1, int(row.get("disc_number") or 1)),
            track_number=max(1, int(row.get("track_number") or 1)),
            sequence_number=max(1, int(row.get("sequence_number") or row.get("track_number") or 1)),
        )
        payload = ReleasePayload(
            title=release_title,
            version_subtitle=str(row.get("release_version_subtitle") or "").strip() or None,
            primary_artist=str(
                row.get("release_primary_artist") or row.get("artist_name") or ""
            ).strip()
            or None,
            album_artist=str(
                row.get("release_album_artist")
                or row.get("release_primary_artist")
                or row.get("artist_name")
                or ""
            ).strip()
            or None,
            release_type=str(row.get("release_type") or "album").strip() or "album",
            release_date=str(
                row.get("release_date_release") or row.get("release_date") or ""
            ).strip()
            or None,
            original_release_date=str(row.get("release_original_release_date") or "").strip()
            or None,
            label=str(row.get("release_label") or "").strip() or None,
            sublabel=str(row.get("release_sublabel") or "").strip() or None,
            catalog_number=str(
                row.get("release_catalog_number") or row.get("catalog_number") or ""
            ).strip()
            or None,
            upc=str(row.get("release_upc") or row.get("upc") or "").strip() or None,
            territory=str(row.get("release_territory") or "").strip() or None,
            explicit_flag=str(row.get("release_explicit_flag") or "").strip().lower()
            in {"1", "true", "yes"},
            notes=str(row.get("release_notes") or "").strip() or None,
            artwork_source_path=self._resolve_media_path(
                source_dir, row.get("release_artwork_path")
            ),
            artwork_storage_mode=str(row.get("release_artwork_storage_mode") or "").strip() or None,
            placements=[placement],
        )
        if catalog_outcomes is not None and catalog_counters is not None and row_index is not None:
            self._record_catalog_classification(
                row_index=int(row_index),
                field_name="release_catalog_number",
                value=payload.catalog_number,
                outcomes=catalog_outcomes,
                counters=catalog_counters,
            )
        if existing_id is None:
            created_release_id = self.release_service.create_release(payload, cursor=cursor)
            if source_release_map is not None and source_release_key:
                source_release_map[source_release_key] = created_release_id
        else:
            summary = self.release_service.fetch_release_summary(existing_id)
            placements = list(summary.tracks) if summary is not None else []
            if all(existing.track_id != track_id for existing in placements):
                placements.append(placement)
            payload.placements = placements
            self.release_service.update_release(existing_id, payload, cursor=cursor)
            if source_release_map is not None and source_release_key:
                source_release_map[source_release_key] = existing_id

    def _import_rows(
        self,
        rows: list[dict[str, object]],
        *,
        mapping: dict[str, str] | None,
        options: ExchangeImportOptions | None,
        format_name: str,
        source_dir: Path,
        source_path: str | None = None,
        progress_callback=None,
        cancel_callback=None,
        repair_entry_id: int | None = None,
        repair_override: dict[str, object] | None = None,
    ) -> ExchangeImportReport:
        opts = options or ExchangeImportOptions()
        effective_mode = str(opts.preview_apply_mode or opts.mode or "dry_run").strip().lower()
        repair_queue_entry_ids: list[int] = []
        self._report_progress(progress_callback, 30, "Mapping and normalizing import rows...")
        if cancel_callback is not None:
            cancel_callback()
        normalized_rows = [
            self._normalize_row_text_targets(self._apply_mapping(row, mapping)) for row in rows
        ]
        source_headers_by_target: dict[str, list[str]] = {}
        for source_header, target_name in (mapping or {}).items():
            clean_target = str(target_name or "").strip()
            if not clean_target:
                continue
            source_headers_by_target.setdefault(clean_target, []).append(str(source_header))
        unknown_fields = sorted(
            {
                key
                for row in normalized_rows
                for key in row
                if key not in self._supported_import_targets()
                and not str(key).startswith("custom::")
            }
        )
        missing_custom_fields = self._ensure_custom_headers(
            normalized_rows,
            create_missing=(opts.create_missing_custom_fields and opts.mode != "dry_run"),
        )
        self._report_progress(
            progress_callback, 36, "Preparing import validation and custom fields..."
        )
        warnings: list[str] = []
        duplicates: list[str] = []
        passed = 0
        failed = 0
        skipped = 0
        would_create_tracks = 0
        would_update_tracks = 0
        created_tracks: list[int] = []
        updated_tracks: list[int] = []
        catalog_outcomes: list[ExchangeIdentifierClassificationOutcome] = []
        catalog_counters = self._empty_identifier_totals()
        total_rows = max(len(normalized_rows), 1)
        package_create_mode = format_name == "package" and (
            effective_mode == "create" or opts.preserve_source_package_identity
        )
        source_track_map: dict[str, int] = {}
        source_release_map: dict[str, int] = {}
        work_batch_cache: dict[str, int] = {}
        work_override_cache: dict[int, dict[str, str]] = {}
        override_mode = str((repair_override or {}).get("governance_mode") or "").strip().lower()
        override_work_id = repair_override.get("work_id") if repair_override is not None else None
        try:
            override_work_id = int(override_work_id) if override_work_id not in (None, "") else None
        except Exception:
            override_work_id = None

        custom_defs = {
            field["name"]: field["id"]
            for field in self.custom_fields.list_active_fields()
            if str(field.get("field_type") or "text") not in {"blob_audio", "blob_image"}
        }
        if missing_custom_fields and opts.mode == "dry_run" and opts.create_missing_custom_fields:
            warnings.append(
                "Dry run would create missing custom fields: "
                + ", ".join(sorted(missing_custom_fields))
            )

        def _apply_custom_fields(
            track_id: int,
            row: dict[str, object],
            *,
            cursor: sqlite3.Cursor,
        ) -> None:
            for key, value in row.items():
                if not str(key).startswith("custom::"):
                    continue
                field_name = str(key).split("::", 1)[1]
                field_id = custom_defs.get(field_name)
                if field_id is None:
                    continue
                cursor.execute(
                    """
                    INSERT INTO CustomFieldValues(
                        track_id,
                        field_def_id,
                        value,
                        blob_value,
                        managed_file_path,
                        storage_mode,
                        filename,
                        mime_type,
                        size_bytes
                    )
                    VALUES (?, ?, ?, NULL, '', '', '', '', 0)
                    ON CONFLICT(track_id, field_def_id) DO UPDATE SET
                        value=excluded.value,
                        blob_value=NULL,
                        managed_file_path=excluded.managed_file_path,
                        storage_mode=excluded.storage_mode,
                        filename=excluded.filename,
                        mime_type=excluded.mime_type,
                        size_bytes=excluded.size_bytes
                    """,
                    (track_id, field_id, str(value or "")),
                )

        for index, row in enumerate(normalized_rows, start=1):
            if cancel_callback is not None:
                cancel_callback()
            row_progress = 40 + int(((index - 1) / total_rows) * 50)
            self._report_progress(
                progress_callback,
                row_progress,
                f"Creating and updating catalog entities for row {index} of {total_rows}...",
            )
            track_title = str(row.get("track_title") or "").strip()
            artist_name = str(row.get("artist_name") or "").strip()
            if not track_title or not artist_name:
                failed += 1
                message = "Track Title and Artist are required."
                warnings.append(f"Row {index}: {message}")
                if opts.mode != "dry_run":
                    queued_id = self._queue_failed_import_row(
                        source_format=format_name,
                        source_path=source_path,
                        row_index=index,
                        import_mode=opts.mode,
                        normalized_row=row,
                        mapping=mapping,
                        options=opts,
                        failure_message=message,
                        repair_entry_id=repair_entry_id,
                    )
                    repair_queue_entry_ids.append(int(queued_id))
                continue
            source_track_key = str(row.get("track_id") or "").strip() if package_create_mode else ""
            reused_package_track = bool(source_track_key) and source_track_key in source_track_map
            if package_create_mode and reused_package_track:
                existing_track_id = source_track_map[source_track_key]
            else:
                existing_track_id = (
                    None
                    if effective_mode == "create"
                    else self._find_existing_track_id(row, options=opts)
                )
            if effective_mode == "update" and existing_track_id is None:
                self._record_catalog_classification(
                    row_index=index,
                    field_name="catalog_number",
                    value=str(row.get("catalog_number") or "").strip() or None,
                    outcomes=catalog_outcomes,
                    counters=catalog_counters,
                    outcome_override="skipped_no_match",
                )
                self._record_catalog_classification(
                    row_index=index,
                    field_name="release_catalog_number",
                    value=str(
                        row.get("release_catalog_number") or row.get("catalog_number") or ""
                    ).strip()
                    or None,
                    outcomes=catalog_outcomes,
                    counters=catalog_counters,
                    outcome_override="skipped_no_match",
                )
                skipped += 1
                warnings.append(f"Row {index}: no existing match was found for update mode.")
                continue
            if effective_mode == "insert_new" and existing_track_id is not None:
                self._record_catalog_classification(
                    row_index=index,
                    field_name="catalog_number",
                    value=str(row.get("catalog_number") or "").strip() or None,
                    outcomes=catalog_outcomes,
                    counters=catalog_counters,
                    outcome_override="skipped_duplicate_match",
                )
                self._record_catalog_classification(
                    row_index=index,
                    field_name="release_catalog_number",
                    value=str(
                        row.get("release_catalog_number") or row.get("catalog_number") or ""
                    ).strip()
                    or None,
                    outcomes=catalog_outcomes,
                    counters=catalog_counters,
                    outcome_override="skipped_duplicate_match",
                )
                skipped += 1
                duplicates.append(f"Row {index}: matched existing track {existing_track_id}")
                continue
            if opts.mode == "dry_run":
                self._record_catalog_classification(
                    row_index=index,
                    field_name="catalog_number",
                    value=str(row.get("catalog_number") or "").strip() or None,
                    outcomes=catalog_outcomes,
                    counters=catalog_counters,
                )
                self._record_catalog_classification(
                    row_index=index,
                    field_name="release_catalog_number",
                    value=str(
                        row.get("release_catalog_number") or row.get("catalog_number") or ""
                    ).strip()
                    or None,
                    outcomes=catalog_outcomes,
                    counters=catalog_counters,
                )
                for field_name in _UNBOUND_IDENTIFIER_IMPORT_FIELDS:
                    self._store_unbound_external_identifier(
                        row_index=index,
                        source_header=(
                            source_headers_by_target.get(field_name, [field_name])[0]
                        ),
                        field_name=field_name,
                        value=row.get(field_name),
                        options=opts,
                        outcomes=catalog_outcomes,
                        totals=catalog_counters,
                        cursor=None,
                    )
                if existing_track_id is None:
                    would_create_tracks += 1
                else:
                    would_update_tracks += 1
                passed += 1
                continue

            try:
                with self.conn:
                    cur = self.conn.cursor()
                    release_date = str(row.get("release_date") or "").strip() or None
                    track_length_value = self._normalize_track_length_target(
                        row.get("track_length_sec")
                    )
                    if is_blank(str(track_length_value or "")):
                        track_length_hms = row.get("track_length_hms")
                        if isinstance(track_length_hms, (dt_time, timedelta)):
                            track_length_value = self._normalize_track_length_target(
                                track_length_hms
                            )
                        else:
                            track_length_value = parse_hms_text(str(track_length_hms or ""))
                    track_length_sec = int(track_length_value or 0)
                    payload_kwargs = dict(
                        isrc=to_iso_isrc(str(row.get("isrc") or "")) or "",
                        track_title=track_title,
                        artist_name=artist_name,
                        additional_artists=[
                            part.strip()
                            for part in str(row.get("additional_artists") or "").split(",")
                            if part.strip()
                        ],
                        album_title=str(row.get("album_title") or "").strip() or None,
                        release_date=release_date,
                        track_length_sec=track_length_sec,
                        iswc=str(row.get("iswc") or "").strip() or None,
                        upc=str(row.get("upc") or "").strip() or None,
                        genre=str(row.get("genre") or "").strip() or None,
                        catalog_number=str(row.get("catalog_number") or "").strip() or None,
                        buma_work_number=str(row.get("buma_work_number") or "").strip() or None,
                        composer=str(row.get("composer") or "").strip() or None,
                        publisher=str(row.get("publisher") or "").strip() or None,
                        comments=str(row.get("comments") or "").strip() or None,
                        lyrics=str(row.get("lyrics") or "").strip() or None,
                        audio_file_source_path=self._resolve_media_path(
                            source_dir, row.get("audio_file_path")
                        ),
                        audio_file_storage_mode=str(
                            row.get("audio_file_storage_mode") or ""
                        ).strip()
                        or None,
                        album_art_source_path=self._resolve_media_path(
                            source_dir, row.get("album_art_path")
                        ),
                        album_art_storage_mode=str(row.get("album_art_storage_mode") or "").strip()
                        or None,
                    )
                    if (
                        row.get("audio_file_path")
                        and payload_kwargs["audio_file_source_path"] is None
                    ):
                        warnings.append(
                            f"Row {index}: Audio reference not found: {row.get('audio_file_path')}"
                        )
                    if (
                        row.get("album_art_path")
                        and payload_kwargs["album_art_source_path"] is None
                    ):
                        warnings.append(
                            f"Row {index}: Artwork reference not found: {row.get('album_art_path')}"
                        )

                    payload_kwargs["artist_name"] = (
                        self.governed_imports.resolve_party_backed_artist_name(
                            payload_kwargs["artist_name"],
                            cursor=cur,
                        )
                    )
                    payload_kwargs["additional_artists"] = (
                        self.governed_imports.resolve_party_backed_additional_artist_names(
                            list(payload_kwargs["additional_artists"]),
                            cursor=cur,
                        )
                    )

                    if existing_track_id is None:
                        if override_mode == "link_existing_work" and override_work_id is not None:
                            payload_kwargs["work_id"] = int(override_work_id)
                        create_result = self.governed_imports.create_governed_track(
                            TrackCreatePayload(
                                **payload_kwargs,
                            ),
                            cursor=cur,
                            batch_cache=work_batch_cache,
                            governance_mode=(
                                "link_existing_work"
                                if override_mode == "link_existing_work" and override_work_id
                                else (
                                    "create_new_work"
                                    if override_mode == "create_new_work"
                                    else "match_or_create_work"
                                )
                            ),
                        )
                        track_id = int(create_result.track_id)
                        if source_track_key:
                            source_track_map[source_track_key] = track_id
                        created_tracks.append(track_id)
                        passed += 1
                    elif package_create_mode and reused_package_track:
                        track_id = existing_track_id
                        if source_track_key:
                            source_track_map[source_track_key] = track_id
                        passed += 1
                    else:
                        snapshot = self.track_service.fetch_track_snapshot(existing_track_id)
                        if snapshot is None:
                            raise ValueError(f"Track {existing_track_id} not found")
                        track_catalog_outcome_override = None
                        if opts.mode == "merge":
                            incoming_catalog_number = (
                                str(row.get("catalog_number") or "").strip() or None
                            )
                            if (
                                incoming_catalog_number
                                and snapshot.catalog_number
                                and str(snapshot.catalog_number).strip() != incoming_catalog_number
                            ):
                                track_catalog_outcome_override = "merged_retained_existing"
                            payload_kwargs["track_title"] = (
                                snapshot.track_title or payload_kwargs["track_title"]
                            )
                            payload_kwargs["artist_name"] = (
                                snapshot.artist_name or payload_kwargs["artist_name"]
                            )
                            payload_kwargs["album_title"] = (
                                snapshot.album_title or payload_kwargs["album_title"]
                            )
                            payload_kwargs["release_date"] = (
                                snapshot.release_date or payload_kwargs["release_date"]
                            )
                            payload_kwargs["track_length_sec"] = (
                                snapshot.track_length_sec or payload_kwargs["track_length_sec"]
                            )
                            payload_kwargs["iswc"] = snapshot.iswc or payload_kwargs["iswc"]
                            payload_kwargs["upc"] = snapshot.upc or payload_kwargs["upc"]
                            payload_kwargs["genre"] = snapshot.genre or payload_kwargs["genre"]
                            payload_kwargs["catalog_number"] = (
                                snapshot.catalog_number or payload_kwargs["catalog_number"]
                            )
                            payload_kwargs["buma_work_number"] = (
                                snapshot.buma_work_number or payload_kwargs["buma_work_number"]
                            )
                            payload_kwargs["composer"] = (
                                snapshot.composer or payload_kwargs["composer"]
                            )
                            payload_kwargs["publisher"] = (
                                snapshot.publisher or payload_kwargs["publisher"]
                            )
                            payload_kwargs["comments"] = (
                                snapshot.comments or payload_kwargs["comments"]
                            )
                            payload_kwargs["lyrics"] = snapshot.lyrics or payload_kwargs["lyrics"]
                            payload_kwargs["audio_file_source_path"] = (
                                payload_kwargs["audio_file_source_path"] or None
                            )
                            payload_kwargs["album_art_source_path"] = (
                                payload_kwargs["album_art_source_path"] or None
                            )

                        if override_mode == "link_existing_work" and override_work_id is not None:
                            payload_kwargs["work_id"] = int(override_work_id)
                        else:
                            payload_kwargs["work_id"] = (
                                self.governed_imports.ensure_governed_work_id(
                                    track_title=payload_kwargs["track_title"],
                                    iswc=payload_kwargs["iswc"],
                                    registration_number=payload_kwargs["buma_work_number"],
                                    composer=payload_kwargs["composer"],
                                    publisher=payload_kwargs["publisher"],
                                    cursor=cur,
                                    batch_cache=work_batch_cache,
                                    existing_work_id=snapshot.work_id,
                                )
                            )

                        self._apply_governed_work_import_override(
                            payload_kwargs,
                            track_id=existing_track_id,
                            cache=work_override_cache,
                        )

                        self.track_service.update_track(
                            TrackUpdatePayload(
                                track_id=existing_track_id,
                                clear_audio_file=False,
                                clear_album_art=False,
                                **payload_kwargs,
                            ),
                            cursor=cur,
                        )
                        updated_tracks.append(existing_track_id)
                        passed += 1
                        track_id = existing_track_id
                        if source_track_key:
                            source_track_map[source_track_key] = int(track_id)

                    self._record_catalog_classification(
                        row_index=index,
                        field_name="catalog_number",
                        value=str(row.get("catalog_number") or "").strip() or None,
                        outcomes=catalog_outcomes,
                        counters=catalog_counters,
                        outcome_override=(
                            track_catalog_outcome_override
                            if "track_catalog_outcome_override" in locals()
                            else None
                        ),
                    )
                    _apply_custom_fields(track_id, row, cursor=cur)
                    self._upsert_release_from_row(
                        row,
                        track_id,
                        source_dir=source_dir,
                        source_release_map=(source_release_map if package_create_mode else None),
                        preserve_source_release_identity=package_create_mode,
                        cursor=cur,
                        catalog_outcomes=catalog_outcomes,
                        catalog_counters=catalog_counters,
                        row_index=index,
                    )
                    for field_name in _UNBOUND_IDENTIFIER_IMPORT_FIELDS:
                        self._store_unbound_external_identifier(
                            row_index=index,
                            source_header=(
                                source_headers_by_target.get(field_name, [field_name])[0]
                            ),
                            field_name=field_name,
                            value=row.get(field_name),
                            options=opts,
                            outcomes=catalog_outcomes,
                            totals=catalog_counters,
                            cursor=cur,
                        )
                    if repair_entry_id is not None:
                        self.repair_queue_service.mark_resolved(
                            int(repair_entry_id),
                            track_id=int(track_id),
                            work_id=int(
                                payload_kwargs.get("work_id")
                                or (
                                    create_result.work_id
                                    if "create_result" in locals()
                                    else snapshot.work_id
                                )
                            ),
                        )
            except Exception as exc:
                failed += 1
                message = str(exc)
                warnings.append(f"Row {index}: {message}")
                if opts.mode != "dry_run":
                    queued_id = self._queue_failed_import_row(
                        source_format=format_name,
                        source_path=source_path,
                        row_index=index,
                        import_mode=opts.mode,
                        normalized_row=row,
                        mapping=mapping,
                        options=opts,
                        failure_message=message,
                        repair_entry_id=repair_entry_id,
                    )
                    repair_queue_entry_ids.append(int(queued_id))

        self._report_progress(progress_callback, 96, "Finalizing import results...")
        self._report_progress(progress_callback, 100, "Import complete.")

        return ExchangeImportReport(
            format_name=format_name,
            mode=opts.mode,
            passed=passed,
            failed=failed,
            skipped=skipped,
            warnings=warnings,
            duplicates=duplicates,
            unknown_fields=unknown_fields,
            evaluated_mode=effective_mode,
            would_create_tracks=would_create_tracks,
            would_update_tracks=would_update_tracks,
            created_tracks=created_tracks,
            updated_tracks=updated_tracks,
            repair_queue_entry_ids=repair_queue_entry_ids,
            source_track_id_map={
                int(key): int(value)
                for key, value in source_track_map.items()
                if str(key).strip().isdigit()
            },
            source_release_id_map={
                int(key): int(value)
                for key, value in source_release_map.items()
                if str(key).strip().isdigit()
            },
            identifier_totals=catalog_counters,
            identifier_classifications=catalog_outcomes,
        )
