"""Exchange helpers for works, parties, contracts, rights, and assets."""

from __future__ import annotations

import csv
import json
import sqlite3
import tempfile
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from isrc_manager.assets import AssetService, AssetVersionPayload
from isrc_manager.contracts import (
    ContractDocumentPayload,
    ContractObligationPayload,
    ContractPartyPayload,
    ContractPayload,
    ContractService,
)
from isrc_manager.file_storage import coalesce_filename
from isrc_manager.parties import PartyPayload, PartyService
from isrc_manager.rights import RightPayload, RightsService
from isrc_manager.works import WorkContributorPayload, WorkPayload, WorkService

REPERTOIRE_JSON_SCHEMA_VERSION = 1


def _stage_progress(start: int, end: int, index: int, total: int) -> int:
    if total <= 0:
        return int(end)
    clamped = max(0, min(index, total))
    return int(start + ((end - start) * clamped / total))


class RepertoireExchangeService:
    """Imports and exports the expanded repertoire knowledge model."""

    ENTITY_FILENAMES = {
        "parties": "parties.csv",
        "works": "works.csv",
        "contracts": "contracts.csv",
        "rights": "rights.csv",
        "assets": "assets.csv",
    }

    def __init__(
        self,
        conn: sqlite3.Connection,
        *,
        party_service: PartyService,
        work_service: WorkService,
        contract_service: ContractService,
        rights_service: RightsService,
        asset_service: AssetService,
        data_root: str | Path | None = None,
    ):
        self.conn = conn
        self.party_service = party_service
        self.work_service = work_service
        self.contract_service = contract_service
        self.rights_service = rights_service
        self.asset_service = asset_service
        self.data_root = Path(data_root) if data_root is not None else None

    @staticmethod
    def _report_progress(
        progress_callback,
        value: int,
        message: str,
        *,
        maximum: int = 100,
    ) -> None:
        if progress_callback is None:
            return
        progress_callback(int(value), int(maximum), str(message or ""))

    def export_payload(self) -> dict[str, object]:
        return {
            "schema_version": REPERTOIRE_JSON_SCHEMA_VERSION,
            "parties": self.party_service.export_rows(),
            "works": self.work_service.export_rows(),
            "contracts": self.contract_service.export_rows(),
            "rights": self.rights_service.export_rows(),
            "assets": self.asset_service.export_rows(),
        }

    def export_json(self, path: str | Path) -> None:
        output = Path(path)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(
            json.dumps(self.export_payload(), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )

    def export_xlsx(self, path: str | Path) -> None:
        workbook = Workbook()
        payload = self.export_payload()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        for sheet_name in ("parties", "works", "contracts", "rights", "assets"):
            rows = [dict(row) for row in payload.get(sheet_name, [])]
            sheet = workbook.create_sheet(sheet_name.title())
            headers = sorted({key for row in rows for key in row}) if rows else ["id"]
            sheet.append(headers)
            for row in rows:
                values = []
                for header in headers:
                    value = row.get(header)
                    if isinstance(value, (dict, list, tuple)):
                        values.append(json.dumps(value, ensure_ascii=True))
                    else:
                        values.append(value)
                sheet.append(values)
        output = Path(path)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)

    def export_csv_bundle(self, directory: str | Path) -> None:
        output_dir = Path(directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        payload = self.export_payload()
        for entity_name, filename in self.ENTITY_FILENAMES.items():
            rows = [dict(row) for row in payload.get(entity_name, [])]
            headers = sorted({key for row in rows for key in row}) if rows else ["id"]
            with (output_dir / filename).open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=headers)
                writer.writeheader()
                for row in rows:
                    writer.writerow(
                        {
                            key: (
                                json.dumps(value, ensure_ascii=True)
                                if isinstance(value, (dict, list, tuple))
                                else value
                            )
                            for key, value in row.items()
                        }
                    )

    def export_package(self, path: str | Path) -> None:
        payload = self.export_payload()
        packaged_files: dict[str, str] = {}
        package_path = Path(path)
        package_path.parent.mkdir(parents=True, exist_ok=True)
        with ZipFile(package_path, "w", compression=ZIP_DEFLATED) as archive:
            for contract in payload["contracts"]:
                contract_id = int(contract.get("id") or 0)
                for document in contract.get("documents", []):
                    stored_path = str(document.get("file_path") or "").strip()
                    document_id = int(document.get("id") or 0)
                    abs_path = self.contract_service.resolve_document_path(stored_path)
                    if stored_path and abs_path is not None and abs_path.exists():
                        package_key = stored_path
                        arcname = f"files/contracts/{Path(stored_path).name}"
                        if arcname not in packaged_files.values():
                            archive.write(abs_path, arcname=arcname)
                        packaged_files[package_key] = arcname
                        continue
                    if document_id <= 0:
                        continue
                    try:
                        data, _ = self.contract_service.fetch_document_bytes(document_id)
                    except Exception:
                        continue
                    filename = coalesce_filename(
                        document.get("filename"),
                        default_stem=f"contract-document-{document_id or contract_id or 'file'}",
                    )
                    package_key = (
                        f"embedded/contracts/{contract_id or 'contract'}/{document_id}/{filename}"
                    )
                    arcname = f"files/contracts/{document_id}_{filename}"
                    if arcname not in packaged_files.values():
                        archive.writestr(arcname, data)
                    document["file_path"] = package_key
                    packaged_files[package_key] = arcname
            for asset in payload["assets"]:
                stored_path = str(asset.get("stored_path") or "").strip()
                asset_id = int(asset.get("id") or 0)
                abs_path = self.asset_service.resolve_asset_path(stored_path)
                if stored_path and abs_path is not None and abs_path.exists():
                    package_key = stored_path
                    arcname = f"files/assets/{Path(stored_path).name}"
                    if arcname not in packaged_files.values():
                        archive.write(abs_path, arcname=arcname)
                    packaged_files[package_key] = arcname
                    continue
                if asset_id <= 0:
                    continue
                try:
                    data, _ = self.asset_service.fetch_asset_bytes(asset_id)
                except Exception:
                    continue
                filename = coalesce_filename(
                    asset.get("filename"),
                    default_stem=f"asset-{asset_id}",
                )
                package_key = f"embedded/assets/{asset_id}/{filename}"
                arcname = f"files/assets/{asset_id}_{filename}"
                if arcname not in packaged_files.values():
                    archive.writestr(arcname, data)
                asset["stored_path"] = package_key
                packaged_files[package_key] = arcname
            payload["packaged_files"] = packaged_files
            archive.writestr("manifest.json", json.dumps(payload, indent=2, ensure_ascii=False))

    def _load_json_payload(self, path: str | Path) -> dict[str, object]:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
        version = int(payload.get("schema_version") or 0)
        if version != REPERTOIRE_JSON_SCHEMA_VERSION:
            raise ValueError(
                f"Unsupported repertoire JSON schema version {version}. Expected {REPERTOIRE_JSON_SCHEMA_VERSION}."
            )
        return payload

    def _rows_from_workbook(self, path: str | Path) -> dict[str, list[dict[str, object]]]:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        rows_by_sheet: dict[str, list[dict[str, object]]] = {}
        for sheet in workbook.worksheets:
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                rows_by_sheet[sheet.title.casefold()] = []
                continue
            headers = [str(value or "").strip() for value in values[0]]
            rows_by_sheet[sheet.title.casefold()] = [
                {headers[index]: row[index] for index in range(len(headers))} for row in values[1:]
            ]
        return rows_by_sheet

    def _rows_from_csv_bundle(self, directory: str | Path) -> dict[str, list[dict[str, object]]]:
        source_dir = Path(directory)
        rows_by_entity: dict[str, list[dict[str, object]]] = {}
        for entity_name, filename in self.ENTITY_FILENAMES.items():
            path = source_dir / filename
            if not path.exists():
                rows_by_entity[entity_name] = []
                continue
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows_by_entity[entity_name] = [dict(row) for row in csv.DictReader(handle)]
        return rows_by_entity

    @staticmethod
    def _decode_value(value: object) -> object:
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return ""
            if text.startswith("[") or text.startswith("{"):
                try:
                    return json.loads(text)
                except Exception:
                    return value
        return value

    def import_json(
        self,
        path: str | Path,
        *,
        progress_callback=None,
        cancel_callback=None,
    ) -> None:
        self._report_progress(progress_callback, 5, "Reading Contracts and Rights JSON...")
        if cancel_callback is not None:
            cancel_callback()
        self._import_payload(
            self._load_json_payload(path),
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_xlsx(
        self,
        path: str | Path,
        *,
        progress_callback=None,
        cancel_callback=None,
    ) -> None:
        self._report_progress(progress_callback, 5, "Reading Contracts and Rights workbook...")
        if cancel_callback is not None:
            cancel_callback()
        rows = self._rows_from_workbook(path)
        self._report_progress(progress_callback, 20, "Parsing repertoire workbook sheets...")
        self._import_payload(
            {
                "schema_version": REPERTOIRE_JSON_SCHEMA_VERSION,
                "parties": rows.get("parties", []),
                "works": rows.get("works", []),
                "contracts": rows.get("contracts", []),
                "rights": rows.get("rights", []),
                "assets": rows.get("assets", []),
            },
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_csv_bundle(
        self,
        directory: str | Path,
        *,
        progress_callback=None,
        cancel_callback=None,
    ) -> None:
        self._report_progress(progress_callback, 5, "Reading Contracts and Rights CSV bundle...")
        if cancel_callback is not None:
            cancel_callback()
        rows = self._rows_from_csv_bundle(directory)
        self._report_progress(progress_callback, 20, "Parsing repertoire CSV bundle...")
        self._import_payload(
            {
                "schema_version": REPERTOIRE_JSON_SCHEMA_VERSION,
                "parties": rows.get("parties", []),
                "works": rows.get("works", []),
                "contracts": rows.get("contracts", []),
                "rights": rows.get("rights", []),
                "assets": rows.get("assets", []),
            },
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_package(
        self,
        path: str | Path,
        *,
        progress_callback=None,
        cancel_callback=None,
    ) -> None:
        self._report_progress(progress_callback, 5, "Extracting Contracts and Rights package...")
        with tempfile.TemporaryDirectory(prefix="repertoire-package-") as tmpdir:
            target_dir = Path(tmpdir)
            with ZipFile(path, "r") as archive:
                archive.extractall(target_dir)
            if cancel_callback is not None:
                cancel_callback()
            payload = json.loads((target_dir / "manifest.json").read_text(encoding="utf-8"))
            packaged_files = payload.get("packaged_files", {})
            if isinstance(packaged_files, dict):
                for contract in payload.get("contracts", []):
                    for document in contract.get("documents", []):
                        stored_path = str(document.get("file_path") or "").strip()
                        arcname = packaged_files.get(stored_path)
                        if arcname:
                            document["source_path"] = str(target_dir / arcname)
                for asset in payload.get("assets", []):
                    stored_path = str(asset.get("stored_path") or "").strip()
                    arcname = packaged_files.get(stored_path)
                    if arcname:
                        asset["source_path"] = str(target_dir / arcname)
            self._report_progress(progress_callback, 20, "Parsing repertoire package manifest...")
            self._import_payload(
                payload,
                progress_callback=progress_callback,
                cancel_callback=cancel_callback,
            )

    def _import_payload(
        self,
        payload: dict[str, object],
        *,
        progress_callback=None,
        cancel_callback=None,
    ) -> None:
        version = int(payload.get("schema_version") or 0)
        if version != REPERTOIRE_JSON_SCHEMA_VERSION:
            raise ValueError(
                f"Unsupported repertoire schema version {version}. Expected {REPERTOIRE_JSON_SCHEMA_VERSION}."
            )
        party_id_map: dict[int, int] = {}
        work_id_map: dict[int, int] = {}
        contract_id_map: dict[int, int] = {}
        document_id_map: dict[int, int] = {}

        def _check_cancel() -> None:
            if cancel_callback is not None:
                cancel_callback()

        parties = list(payload.get("parties", []) or [])
        works = list(payload.get("works", []) or [])
        contracts = list(payload.get("contracts", []) or [])
        rights = list(payload.get("rights", []) or [])
        assets = list(payload.get("assets", []) or [])

        with self.conn:
            self._report_progress(progress_callback, 30, "Importing Party records...")
            for index, row in enumerate(parties, start=1):
                _check_cancel()
                self._report_progress(
                    progress_callback,
                    _stage_progress(30, 45, index - 1, len(parties)),
                    f"Importing Party records ({index} of {len(parties)})...",
                )
                source = {key: self._decode_value(value) for key, value in dict(row).items()}
                old_id = int(source.get("id") or 0)
                legal_name = str(source.get("legal_name") or "").strip()
                existing = self.conn.execute(
                    "SELECT id FROM Parties WHERE legal_name=? ORDER BY id LIMIT 1",
                    (legal_name,),
                ).fetchone()
                if existing:
                    party_id_map[old_id] = int(existing[0])
                    continue
                new_id = self.party_service.create_party(
                    PartyPayload(
                        legal_name=legal_name,
                        display_name=source.get("display_name"),
                        artist_name=source.get("artist_name"),
                        company_name=source.get("company_name"),
                        first_name=source.get("first_name"),
                        middle_name=source.get("middle_name"),
                        last_name=source.get("last_name"),
                        party_type=source.get("party_type") or "organization",
                        contact_person=source.get("contact_person"),
                        email=source.get("email"),
                        alternative_email=source.get("alternative_email"),
                        phone=source.get("phone"),
                        website=source.get("website"),
                        street_name=source.get("street_name"),
                        street_number=source.get("street_number"),
                        address_line1=source.get("address_line1"),
                        address_line2=source.get("address_line2"),
                        city=source.get("city"),
                        region=source.get("region"),
                        postal_code=source.get("postal_code"),
                        country=source.get("country"),
                        bank_account_number=source.get("bank_account_number"),
                        chamber_of_commerce_number=source.get("chamber_of_commerce_number"),
                        tax_id=source.get("tax_id"),
                        vat_number=source.get("vat_number"),
                        pro_affiliation=source.get("pro_affiliation"),
                        pro_number=source.get("pro_number"),
                        ipi_cae=source.get("ipi_cae"),
                        notes=source.get("notes"),
                        profile_name=source.get("profile_name"),
                        artist_aliases=[
                            str(item).strip()
                            for item in list(source.get("artist_aliases", []) or [])
                            if str(item).strip()
                        ],
                    ),
                    cursor=self.conn.cursor(),
                )
                party_id_map[old_id] = new_id

            self._report_progress(progress_callback, 45, "Importing Work records...")
            for index, row in enumerate(works, start=1):
                _check_cancel()
                self._report_progress(
                    progress_callback,
                    _stage_progress(45, 60, index - 1, len(works)),
                    f"Importing Work records ({index} of {len(works)})...",
                )
                source = {key: self._decode_value(value) for key, value in dict(row).items()}
                old_id = int(source.get("id") or 0)
                contributors = []
                for contributor in source.get("contributors", []) or []:
                    contributors.append(
                        WorkContributorPayload(
                            role=contributor.get("role") or "songwriter",
                            name=contributor.get("display_name") or "",
                            share_percent=contributor.get("share_percent"),
                            role_share_percent=contributor.get("role_share_percent"),
                            party_id=party_id_map.get(int(contributor.get("party_id") or 0)),
                            notes=contributor.get("notes"),
                        )
                    )
                track_ids = [
                    int(track_id)
                    for track_id in source.get("track_ids", []) or []
                    if self.conn.execute(
                        "SELECT 1 FROM Tracks WHERE id=?", (int(track_id),)
                    ).fetchone()
                ]
                new_id = self.work_service.create_work(
                    WorkPayload(
                        title=source.get("title") or "",
                        alternate_titles=list(source.get("alternate_titles", []) or []),
                        version_subtitle=source.get("version_subtitle"),
                        language=source.get("language"),
                        lyrics_flag=bool(source.get("lyrics_flag")),
                        instrumental_flag=bool(source.get("instrumental_flag")),
                        genre_notes=source.get("genre_notes"),
                        iswc=source.get("iswc"),
                        registration_number=source.get("registration_number"),
                        work_status=source.get("work_status"),
                        metadata_complete=bool(source.get("metadata_complete")),
                        contract_signed=bool(source.get("contract_signed")),
                        rights_verified=bool(source.get("rights_verified")),
                        notes=source.get("notes"),
                        profile_name=source.get("profile_name"),
                        contributors=contributors,
                        track_ids=track_ids,
                    ),
                    cursor=self.conn.cursor(),
                )
                work_id_map[old_id] = new_id

            contract_documents_buffer: list[tuple[int, list[dict[str, object]]]] = []
            self._report_progress(progress_callback, 60, "Importing Contract records...")
            for index, row in enumerate(contracts, start=1):
                _check_cancel()
                self._report_progress(
                    progress_callback,
                    _stage_progress(60, 75, index - 1, len(contracts)),
                    f"Importing Contract records ({index} of {len(contracts)})...",
                )
                source = {key: self._decode_value(value) for key, value in dict(row).items()}
                old_id = int(source.get("id") or 0)
                party_payloads = []
                for party in source.get("parties", []) or []:
                    mapped_party_id = party_id_map.get(int(party.get("party_id") or 0))
                    if not mapped_party_id:
                        continue
                    party_payloads.append(
                        ContractPartyPayload(
                            party_id=mapped_party_id,
                            role_label=party.get("role_label") or "counterparty",
                            is_primary=bool(party.get("is_primary")),
                            notes=party.get("notes"),
                        )
                    )
                obligation_payloads = [
                    ContractObligationPayload(
                        obligation_type=item.get("obligation_type") or "other",
                        title=item.get("title") or "",
                        due_date=item.get("due_date"),
                        follow_up_date=item.get("follow_up_date"),
                        reminder_date=item.get("reminder_date"),
                        completed=bool(item.get("completed")),
                        completed_at=item.get("completed_at"),
                        notes=item.get("notes"),
                    )
                    for item in source.get("obligations", []) or []
                ]
                document_payloads = []
                for item in source.get("documents", []) or []:
                    document_payloads.append(
                        ContractDocumentPayload(
                            title=item.get("title") or "",
                            document_type=item.get("document_type") or "other",
                            version_label=item.get("version_label"),
                            created_date=item.get("created_date"),
                            received_date=item.get("received_date"),
                            signed_status=item.get("signed_status"),
                            signed_by_all_parties=bool(item.get("signed_by_all_parties")),
                            active_flag=bool(item.get("active_flag")),
                            source_path=item.get("source_path"),
                            stored_path=item.get("file_path"),
                            storage_mode=item.get("storage_mode"),
                            filename=item.get("filename"),
                            checksum_sha256=item.get("checksum_sha256"),
                            notes=item.get("notes"),
                        )
                    )
                work_ids = [
                    work_id_map.get(int(item), 0) for item in source.get("work_ids", []) or []
                ]
                track_ids = [
                    int(item)
                    for item in source.get("track_ids", []) or []
                    if self.conn.execute("SELECT 1 FROM Tracks WHERE id=?", (int(item),)).fetchone()
                ]
                release_ids = [
                    int(item)
                    for item in source.get("release_ids", []) or []
                    if self.conn.execute(
                        "SELECT 1 FROM Releases WHERE id=?", (int(item),)
                    ).fetchone()
                ]
                new_id = self.contract_service.create_contract(
                    ContractPayload(
                        title=source.get("title") or "",
                        contract_type=source.get("contract_type"),
                        draft_date=source.get("draft_date"),
                        signature_date=source.get("signature_date"),
                        effective_date=source.get("effective_date"),
                        start_date=source.get("start_date"),
                        end_date=source.get("end_date"),
                        renewal_date=source.get("renewal_date"),
                        notice_deadline=source.get("notice_deadline"),
                        option_periods=source.get("option_periods"),
                        reversion_date=source.get("reversion_date"),
                        termination_date=source.get("termination_date"),
                        status=source.get("status") or "draft",
                        summary=source.get("summary"),
                        notes=source.get("notes"),
                        profile_name=source.get("profile_name"),
                        parties=party_payloads,
                        obligations=obligation_payloads,
                        documents=document_payloads,
                        work_ids=[item for item in work_ids if item],
                        track_ids=track_ids,
                        release_ids=release_ids,
                    ),
                    cursor=self.conn.cursor(),
                )
                contract_id_map[old_id] = new_id
                contract_documents_buffer.append((new_id, list(source.get("documents", []) or [])))

            self._report_progress(progress_callback, 75, "Relinking contract document chains...")
            for new_contract_id, source_documents in contract_documents_buffer:
                _check_cancel()
                detail = self.contract_service.fetch_contract_detail(new_contract_id)
                if detail is None:
                    continue
                local_docs = detail.documents
                for source_doc, local_doc in zip(source_documents, local_docs):
                    old_doc_id = int(source_doc.get("id") or 0)
                    if old_doc_id:
                        document_id_map[old_doc_id] = local_doc.id
                for source_doc in source_documents:
                    local_doc_id = document_id_map.get(int(source_doc.get("id") or 0))
                    if not local_doc_id:
                        continue
                    supersedes_old = int(source_doc.get("supersedes_document_id") or 0)
                    superseded_by_old = int(source_doc.get("superseded_by_document_id") or 0)
                    self.conn.execute(
                        """
                        UPDATE ContractDocuments
                        SET supersedes_document_id=?,
                            superseded_by_document_id=?
                        WHERE id=?
                        """,
                        (
                            document_id_map.get(supersedes_old) if supersedes_old else None,
                            document_id_map.get(superseded_by_old) if superseded_by_old else None,
                            int(local_doc_id),
                        ),
                    )

            self._report_progress(progress_callback, 78, "Importing Right records...")
            for index, row in enumerate(rights, start=1):
                _check_cancel()
                self._report_progress(
                    progress_callback,
                    _stage_progress(78, 88, index - 1, len(rights)),
                    f"Importing Right records ({index} of {len(rights)})...",
                )
                source = {key: self._decode_value(value) for key, value in dict(row).items()}
                self.rights_service.create_right(
                    RightPayload(
                        title=source.get("title"),
                        right_type=source.get("right_type") or "other",
                        exclusive_flag=bool(source.get("exclusive_flag")),
                        territory=source.get("territory"),
                        media_use_type=source.get("media_use_type"),
                        start_date=source.get("start_date"),
                        end_date=source.get("end_date"),
                        perpetual_flag=bool(source.get("perpetual_flag")),
                        granted_by_party_id=party_id_map.get(
                            int(source.get("granted_by_party_id") or 0)
                        ),
                        granted_to_party_id=party_id_map.get(
                            int(source.get("granted_to_party_id") or 0)
                        ),
                        retained_by_party_id=party_id_map.get(
                            int(source.get("retained_by_party_id") or 0)
                        ),
                        source_contract_id=contract_id_map.get(
                            int(source.get("source_contract_id") or 0)
                        ),
                        work_id=work_id_map.get(int(source.get("work_id") or 0)),
                        track_id=(
                            int(source.get("track_id"))
                            if source.get("track_id")
                            and self.conn.execute(
                                "SELECT 1 FROM Tracks WHERE id=?",
                                (int(source.get("track_id")),),
                            ).fetchone()
                            else None
                        ),
                        release_id=(
                            int(source.get("release_id"))
                            if source.get("release_id")
                            and self.conn.execute(
                                "SELECT 1 FROM Releases WHERE id=?",
                                (int(source.get("release_id")),),
                            ).fetchone()
                            else None
                        ),
                        notes=source.get("notes"),
                        profile_name=source.get("profile_name"),
                    )
                )

            self._report_progress(progress_callback, 88, "Importing Asset records...")
            for index, row in enumerate(assets, start=1):
                _check_cancel()
                self._report_progress(
                    progress_callback,
                    _stage_progress(88, 96, index - 1, len(assets)),
                    f"Importing Asset records ({index} of {len(assets)})...",
                )
                source = {key: self._decode_value(value) for key, value in dict(row).items()}
                self.asset_service.create_asset(
                    AssetVersionPayload(
                        asset_type=source.get("asset_type") or "other",
                        filename=source.get("filename"),
                        source_path=source.get("source_path"),
                        stored_path=source.get("stored_path"),
                        storage_mode=source.get("storage_mode"),
                        checksum_sha256=source.get("checksum_sha256"),
                        duration_sec=(
                            int(source["duration_sec"])
                            if source.get("duration_sec") not in (None, "")
                            else None
                        ),
                        sample_rate=(
                            int(source["sample_rate"])
                            if source.get("sample_rate") not in (None, "")
                            else None
                        ),
                        bit_depth=(
                            int(source["bit_depth"])
                            if source.get("bit_depth") not in (None, "")
                            else None
                        ),
                        format=source.get("format"),
                        derived_from_asset_id=None,
                        approved_for_use=bool(source.get("approved_for_use")),
                        primary_flag=bool(source.get("primary_flag")),
                        version_status=source.get("version_status"),
                        notes=source.get("notes"),
                        track_id=(
                            int(source.get("track_id"))
                            if source.get("track_id")
                            and self.conn.execute(
                                "SELECT 1 FROM Tracks WHERE id=?",
                                (int(source.get("track_id")),),
                            ).fetchone()
                            else None
                        ),
                        release_id=(
                            int(source.get("release_id"))
                            if source.get("release_id")
                            and self.conn.execute(
                                "SELECT 1 FROM Releases WHERE id=?",
                                (int(source.get("release_id")),),
                            ).fetchone()
                            else None
                        ),
                    )
                )
        self._report_progress(progress_callback, 98, "Finalizing Contracts and Rights import...")
        self._report_progress(progress_callback, 100, "Contracts and Rights import complete.")
