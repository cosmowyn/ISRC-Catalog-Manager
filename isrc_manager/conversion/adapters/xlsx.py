"""XLSX template and source adapters for conversion."""

from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from ..mapping import normalize_field_name
from ..models import (
    REQUIRED_STATUS_OPTIONAL,
    REQUIRED_STATUS_REQUIRED,
    SOURCE_MODE_FILE,
    ConversionExportResult,
    ConversionSourceProfile,
    ConversionTargetField,
    ConversionTemplateProfile,
)
from .base import SourceAdapter, TemplateAdapter


def _iter_non_empty_cells(row) -> bool:
    return any(str(cell.value or "").strip() for cell in row)


def _column_letter(index: int) -> str:
    letters = ""
    value = int(index)
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters or "A"


def _open_workbook(
    source: Path | bytes,
    *,
    read_only: bool,
    data_only: bool,
    keep_vba: bool = False,
):
    if isinstance(source, Path):
        return load_workbook(
            filename=str(source),
            read_only=read_only,
            data_only=data_only,
            keep_vba=keep_vba,
        )
    return load_workbook(
        filename=BytesIO(bytes(source)),
        read_only=read_only,
        data_only=data_only,
        keep_vba=keep_vba,
    )


def _sheet_profiles(source: Path | bytes) -> dict[str, dict[str, object]]:
    workbook = _open_workbook(source, read_only=False, data_only=False)
    try:
        profiles: dict[str, dict[str, object]] = {}
        for worksheet in workbook.worksheets:
            header_row_index = None
            for row_index in range(1, worksheet.max_row + 1):
                row = [
                    worksheet.cell(row=row_index, column=column)
                    for column in range(1, worksheet.max_column + 1)
                ]
                if _iter_non_empty_cells(row):
                    header_row_index = row_index
                    break
            if header_row_index is None:
                continue
            header_row = [
                worksheet.cell(row=header_row_index, column=column).value
                for column in range(1, worksheet.max_column + 1)
            ]
            target_fields: list[ConversionTargetField] = []
            for column_index, header in enumerate(header_row, start=1):
                clean_header = str(header or "").strip()
                if not clean_header:
                    continue
                required_status = (
                    REQUIRED_STATUS_REQUIRED
                    if clean_header.endswith("*")
                    else REQUIRED_STATUS_OPTIONAL
                )
                target_fields.append(
                    ConversionTargetField(
                        field_key=normalize_field_name(clean_header),
                        display_name=clean_header,
                        location=f"{worksheet.title}!{_column_letter(column_index)}{header_row_index}",
                        required_status=required_status,
                        kind="cell",
                        metadata={"column_index": column_index},
                    )
                )
            sample_start = header_row_index + 1
            sample_end = sample_start - 1
            for row_index in range(sample_start, worksheet.max_row + 1):
                row_values = [
                    worksheet.cell(row=row_index, column=column).value
                    for column in range(1, worksheet.max_column + 1)
                ]
                if not any(str(value or "").strip() for value in row_values):
                    break
                sample_end = row_index
            profiles[worksheet.title] = {
                "sheet_name": worksheet.title,
                "header_row_index": header_row_index,
                "target_fields": tuple(target_fields),
                "sample_start": sample_start,
                "sample_end": sample_end,
                "style_row": sample_start if sample_end >= sample_start else None,
            }
        return profiles
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


class XlsxTemplateAdapter(TemplateAdapter):
    format_name = "xlsx"

    def inspect_template(self, path: str | Path) -> ConversionTemplateProfile:
        template_path = Path(path)
        sheet_profiles = _sheet_profiles(template_path)
        if not sheet_profiles:
            raise ValueError("The workbook does not contain a usable data sheet.")
        available_scopes = tuple((name, name) for name in sheet_profiles)
        chosen_scope = available_scopes[0][0]
        selected = sheet_profiles[chosen_scope]
        target_fields = tuple(selected["target_fields"])
        signature = self._template_signature(chosen_scope, target_fields)
        warnings: list[str] = []
        if len(available_scopes) > 1:
            warnings.append("Multiple workbook sheets look writable. Review the selected sheet.")
        return ConversionTemplateProfile(
            template_path=template_path,
            format_name=self.format_name,
            output_suffix=template_path.suffix.lower() or ".xlsx",
            structure_label=f"Workbook template ({template_path.name})",
            target_fields=target_fields,
            template_signature=signature,
            template_bytes=None,
            available_scopes=available_scopes,
            chosen_scope=chosen_scope,
            warnings=tuple(warnings),
            adapter_state={"sheet_profiles": sheet_profiles},
        )

    def select_scope(
        self,
        profile: ConversionTemplateProfile,
        scope_key: str,
    ) -> ConversionTemplateProfile:
        sheet_profiles = profile.adapter_state.get("sheet_profiles") or {}
        selected = sheet_profiles.get(scope_key)
        if not selected:
            return profile
        return ConversionTemplateProfile(
            template_path=profile.template_path,
            format_name=profile.format_name,
            output_suffix=profile.output_suffix,
            structure_label=profile.structure_label,
            target_fields=tuple(selected["target_fields"]),
            template_signature=self._template_signature(
                scope_key, tuple(selected["target_fields"])
            ),
            template_bytes=profile.template_bytes,
            available_scopes=profile.available_scopes,
            chosen_scope=str(scope_key),
            warnings=profile.warnings,
            adapter_state=profile.adapter_state,
        )

    def build_preview(
        self,
        profile: ConversionTemplateProfile,
        rendered_field_rows: list[dict[str, object]],
    ) -> tuple[
        tuple[str, ...],
        tuple[tuple[str, ...], ...],
        str,
        tuple[str, ...],
        dict[str, object],
    ]:
        headers = tuple(field.display_name for field in profile.target_fields)
        rows = [
            tuple(str(rendered.get(field.field_key, "")) for field in profile.target_fields)
            for rendered in rendered_field_rows
        ]
        return headers, tuple(rows), "", (), {}

    def export_preview(
        self,
        preview,
        output_path: str | Path,
        *,
        progress_callback=None,
    ) -> ConversionExportResult:
        template_path = preview.template_profile.template_path
        workbook = _open_workbook(
            (
                preview.template_profile.template_bytes
                if preview.template_profile.template_bytes is not None
                else template_path
            ),
            read_only=False,
            data_only=False,
            keep_vba=template_path.suffix.lower() == ".xlsm",
        )
        try:
            if callable(progress_callback):
                progress_callback(15, 100, "Opening workbook template...")
            sheet_name = preview.template_profile.chosen_scope
            worksheet = workbook[sheet_name]
            sheet_profiles = preview.template_profile.adapter_state.get("sheet_profiles") or {}
            selected = sheet_profiles.get(sheet_name) or {}
            start_row = int(
                selected.get("sample_start") or (int(selected.get("header_row_index") or 1) + 1)
            )
            sample_end = int(selected.get("sample_end") or (start_row - 1))
            style_row = selected.get("style_row")
            target_fields = tuple(preview.template_profile.target_fields)
            for index, rendered in enumerate(preview.rendered_rows, start=0):
                row_number = start_row + index
                if style_row and row_number > worksheet.max_row:
                    self._clone_row_style(worksheet, int(style_row), row_number)
                for field, value in zip(target_fields, rendered):
                    column_index = int(field.metadata.get("column_index") or 1)
                    worksheet.cell(row=row_number, column=column_index, value=str(value))
            for row_number in range(start_row + len(preview.rendered_rows), sample_end + 1):
                for field in target_fields:
                    column_index = int(field.metadata.get("column_index") or 1)
                    worksheet.cell(row=row_number, column=column_index, value=None)
            if callable(progress_callback):
                progress_callback(80, 100, "Writing converted workbook...")
            target = Path(output_path)
            target.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(str(target))
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()
        if callable(progress_callback):
            progress_callback(90, 100, "Workbook conversion export written.")
        return ConversionExportResult(
            output_path=Path(output_path),
            target_format=self.format_name,
            exported_row_count=len(preview.rendered_rows),
            summary_lines=(
                f"Template: {template_path.name}",
                f"Sheet: {preview.template_profile.chosen_scope}",
                f"Rows written: {len(preview.rendered_rows)}",
            ),
        )

    @staticmethod
    def _clone_row_style(worksheet, source_row: int, target_row: int) -> None:
        if source_row <= 0 or source_row > worksheet.max_row or target_row <= worksheet.max_row:
            return
        for column_index in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(row=source_row, column=column_index)
            target_cell = worksheet.cell(row=target_row, column=column_index)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
        source_dimension = worksheet.row_dimensions.get(source_row)
        if source_dimension is not None:
            worksheet.row_dimensions[target_row].height = source_dimension.height
            worksheet.row_dimensions[target_row].hidden = source_dimension.hidden

    @staticmethod
    def _template_signature(
        sheet_name: str, target_fields: tuple[ConversionTargetField, ...]
    ) -> str:
        field_ids = ",".join(field.field_key for field in target_fields)
        return f"xlsx|sheet:{sheet_name}|{field_ids}"


class XlsxSourceAdapter(SourceAdapter):
    format_name = "xlsx"

    def inspect_source(
        self,
        source,
        *,
        preferred_csv_delimiter: str | None = None,
    ) -> ConversionSourceProfile:
        del preferred_csv_delimiter
        source_path = Path(source)
        sheet_profiles = _sheet_profiles(source_path)
        if not sheet_profiles:
            raise ValueError("The workbook does not contain a usable source sheet.")
        available_scopes = tuple((name, name) for name in sheet_profiles)
        chosen_scope = available_scopes[0][0]
        return self.select_scope(
            ConversionSourceProfile(
                source_mode=SOURCE_MODE_FILE,
                format_name=self.format_name,
                source_label=str(source_path),
                source_path=str(source_path),
                headers=(),
                rows=(),
                preview_rows=(),
                available_scopes=available_scopes,
                chosen_scope=chosen_scope,
                warnings=tuple(
                    ["Multiple source sheets look usable. Review the selected sheet."]
                    if len(available_scopes) > 1
                    else []
                ),
                adapter_state={"sheet_profiles": sheet_profiles},
            ),
            chosen_scope,
        )

    def select_scope(
        self,
        profile: ConversionSourceProfile,
        scope_key: str,
    ) -> ConversionSourceProfile:
        sheet_profiles = profile.adapter_state.get("sheet_profiles") or {}
        selected = sheet_profiles.get(scope_key)
        if not selected:
            return profile
        source_path = Path(profile.source_path)
        workbook = load_workbook(filename=str(source_path), read_only=True, data_only=True)
        try:
            worksheet = workbook[scope_key]
            header_row_index = int(selected.get("header_row_index") or 1)
            headers = [
                str(worksheet.cell(row=header_row_index, column=column).value or "").strip()
                for column in range(1, worksheet.max_column + 1)
            ]
            rows: list[dict[str, object]] = []
            for row_index in range(header_row_index + 1, worksheet.max_row + 1):
                raw_values = [
                    worksheet.cell(row=row_index, column=column).value
                    for column in range(1, worksheet.max_column + 1)
                ]
                if not any(str(value or "").strip() for value in raw_values):
                    continue
                rows.append(
                    {
                        headers[index]: raw_values[index]
                        for index in range(len(headers))
                        if headers[index]
                    }
                )
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()
        return ConversionSourceProfile(
            source_mode=profile.source_mode,
            format_name=profile.format_name,
            source_label=profile.source_label,
            source_path=profile.source_path,
            headers=tuple(header for header in headers if header),
            rows=tuple(rows),
            preview_rows=tuple(rows[:10]),
            available_scopes=profile.available_scopes,
            chosen_scope=str(scope_key),
            warnings=profile.warnings,
            adapter_state=profile.adapter_state,
        )
