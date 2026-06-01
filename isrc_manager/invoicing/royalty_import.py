"""DSP statement import support for royalty source events."""

from __future__ import annotations

import csv
import re
import sqlite3
from calendar import monthrange
from collections.abc import Mapping
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

from isrc_manager.domain.repertoire import clean_text

from .models import DEFAULT_CURRENCY
from .money import normalize_currency, parse_money_minor
from .royalty_integration import RoyaltyIntegrationService, RoyaltySourceEventPayload

ROYALTY_IMPORT_SKIP_TARGET = "__skip_field__"

ROYALTY_SOURCE_IMPORT_TARGETS: tuple[tuple[str, str], ...] = (
    ("Description", "description"),
    ("Source type", "source_type"),
    ("Source ID", "source_id"),
    ("Contract ID", "contract_id"),
    ("Work ID", "work_id"),
    ("Work title", "work_title"),
    ("Track ID", "track_id"),
    ("Track ISRC", "track_isrc"),
    ("Track title", "track_title"),
    ("Release ID", "release_id"),
    ("Release title", "release_title"),
    ("Event date", "event_date"),
    ("Period start", "period_start"),
    ("Period end", "period_end"),
    ("Currency", "currency"),
    ("Gross amount", "gross_amount"),
    ("Net amount", "net_amount"),
    ("Gross amount minor", "gross_amount_minor"),
    ("Net amount minor", "net_amount_minor"),
)

_SUPPORTED_TARGET_VALUES = {target for _label, target in ROYALTY_SOURCE_IMPORT_TARGETS}
_AMOUNT_HEADER_CURRENCY_RE = re.compile(r"(?:^|_)(?P<currency>[a-zA-Z]{3})$")

_HEADER_ALIASES: dict[str, str] = {
    "description": "description",
    "title": "description",
    "track": "description",
    "tracktitle": "description",
    "track_title": "description",
    "work": "work_title",
    "worktitle": "work_title",
    "work_title": "work_title",
    "composition": "work_title",
    "compositiontitle": "work_title",
    "composition_title": "work_title",
    "isrc": "track_isrc",
    "trackisrc": "track_isrc",
    "track_isrc": "track_isrc",
    "recordingisrc": "track_isrc",
    "recording_isrc": "track_isrc",
    "recording": "track_title",
    "recordingtitle": "track_title",
    "recording_title": "track_title",
    "trackname": "track_title",
    "track_name": "track_title",
    "release": "release_title",
    "releasetitle": "release_title",
    "release_title": "release_title",
    "album": "release_title",
    "albumtitle": "release_title",
    "album_title": "release_title",
    "source": "source_type",
    "store": "source_type",
    "storename": "source_type",
    "store_name": "source_type",
    "provider": "source_type",
    "dsp": "source_type",
    "platform": "source_type",
    "servicename": "source_type",
    "service_name": "source_type",
    "sourcetype": "source_type",
    "source_type": "source_type",
    "id": "source_id",
    "sourceid": "source_id",
    "source_id": "source_id",
    "statementid": "source_id",
    "statement_id": "source_id",
    "transactionid": "source_id",
    "transaction_id": "source_id",
    "usageid": "source_id",
    "usage_id": "source_id",
    "lineid": "source_id",
    "line_id": "source_id",
    "contractid": "contract_id",
    "contract_id": "contract_id",
    "workid": "work_id",
    "work_id": "work_id",
    "trackid": "track_id",
    "track_id": "track_id",
    "recordingid": "track_id",
    "recording_id": "track_id",
    "releaseid": "release_id",
    "release_id": "release_id",
    "date": "event_date",
    "dateid": "period_start",
    "date_id": "period_start",
    "period": "period_start",
    "eventdate": "event_date",
    "event_date": "event_date",
    "transactiondate": "event_date",
    "transaction_date": "event_date",
    "usagedate": "event_date",
    "usage_date": "event_date",
    "periodstart": "period_start",
    "period_start": "period_start",
    "startdate": "period_start",
    "start_date": "period_start",
    "from": "period_start",
    "periodend": "period_end",
    "period_end": "period_end",
    "enddate": "period_end",
    "end_date": "period_end",
    "to": "period_end",
    "currency": "currency",
    "ccy": "currency",
    "gross": "gross_amount",
    "grossamount": "gross_amount",
    "gross_amount": "gross_amount",
    "grossrevenue": "gross_amount",
    "gross_revenue": "gross_amount",
    "grossroyalty": "gross_amount",
    "gross_royalty": "gross_amount",
    "net": "net_amount",
    "netamount": "net_amount",
    "net_amount": "net_amount",
    "netrevenue": "net_amount",
    "net_revenue": "net_amount",
    "payable": "net_amount",
    "royalty": "net_amount",
    "netroyalty": "net_amount",
    "net_royalty": "net_amount",
    "grossminor": "gross_amount_minor",
    "gross_minor": "gross_amount_minor",
    "grossamountminor": "gross_amount_minor",
    "gross_amount_minor": "gross_amount_minor",
    "netminor": "net_amount_minor",
    "net_minor": "net_amount_minor",
    "netamountminor": "net_amount_minor",
    "net_amount_minor": "net_amount_minor",
    "total": "net_amount",
    "totalgbp": "net_amount",
    "total_gbp": "net_amount",
    "totaleur": "net_amount",
    "total_eur": "net_amount",
    "totalusd": "net_amount",
    "total_usd": "net_amount",
}


@dataclass(frozen=True, slots=True)
class RoyaltySourceImportInspection:
    file_path: str
    format_name: str
    headers: list[str]
    preview_rows: list[dict[str, object]]
    suggested_mapping: dict[str, str]
    warnings: list[str] = field(default_factory=list)
    total_rows: int = 0


@dataclass(frozen=True, slots=True)
class RoyaltySourceImportPreviewRow:
    row_number: int
    status: str
    description: str
    contract_id: int | None
    work_id: int | None
    track_id: int | None
    release_id: int | None
    event_date: str | None
    period_start: str | None
    period_end: str | None
    source_type: str | None
    source_id: str | None
    currency: str
    gross_amount_minor: int
    net_amount_minor: int
    issues: tuple[str, ...] = ()

    def to_preview_dict(self) -> dict[str, object]:
        return {
            "Row": self.row_number,
            "Status": self.status,
            "Description": self.description,
            "Contract": self.contract_id or "",
            "Work": self.work_id or "",
            "Track": self.track_id or "",
            "Release": self.release_id or "",
            "Event date": self.event_date or "",
            "Period start": self.period_start or "",
            "Period end": self.period_end or "",
            "Source": self.source_type or "",
            "Source ID": self.source_id or "",
            "Currency": self.currency,
            "Gross minor": self.gross_amount_minor,
            "Net minor": self.net_amount_minor,
            "Issues": "; ".join(self.issues),
        }


@dataclass(frozen=True, slots=True)
class RoyaltySourceImportReport:
    format_name: str
    mode: str
    passed: int
    failed: int
    skipped: int
    warnings: list[str]
    preview_rows: tuple[RoyaltySourceImportPreviewRow, ...]
    created_event_ids: tuple[int, ...] = ()

    @property
    def summary_lines(self) -> list[str]:
        action = "would import" if self.mode == "preview" else "imported"
        return [
            f"Format: {self.format_name.upper()}",
            f"Rows {action}: {self.passed}",
            f"Rows failed: {self.failed}",
            f"Rows skipped: {self.skipped}",
        ]


class RoyaltySourceImportService:
    """Inspect and import DSP source rows into royalty source events."""

    def __init__(self, conn: sqlite3.Connection):
        self.conn = conn
        self.integration = RoyaltyIntegrationService(conn)

    def inspect_file(self, path: str | Path) -> RoyaltySourceImportInspection:
        source_path = Path(path)
        rows, format_name, warnings = self._read_rows(source_path)
        headers = self._headers_from_rows(rows)
        return RoyaltySourceImportInspection(
            file_path=str(source_path),
            format_name=format_name,
            headers=headers,
            preview_rows=rows[:20],
            suggested_mapping=self.suggest_mapping(headers),
            warnings=warnings,
            total_rows=len(rows),
        )

    def preview_import(
        self,
        path: str | Path,
        mapping: Mapping[str, str],
        *,
        default_contract_id: int | None = None,
    ) -> RoyaltySourceImportReport:
        rows, format_name, warnings = self._read_rows(Path(path))
        preview_rows = tuple(
            self._preview_row(
                row,
                row_number=index,
                mapping=mapping,
                default_contract_id=default_contract_id,
            )
            for index, row in enumerate(rows, start=1)
        )
        return self._report(format_name, "preview", warnings, preview_rows)

    def apply_import(
        self,
        path: str | Path,
        mapping: Mapping[str, str],
        *,
        default_contract_id: int | None = None,
    ) -> RoyaltySourceImportReport:
        preview = self.preview_import(
            path,
            mapping,
            default_contract_id=default_contract_id,
        )
        created_ids: list[int] = []
        with self.conn:
            for row in preview.preview_rows:
                if row.status != "ready":
                    continue
                event = self.integration.record_source_event(
                    RoyaltySourceEventPayload(
                        contract_id=row.contract_id,
                        work_id=row.work_id,
                        track_id=row.track_id,
                        release_id=row.release_id,
                        source_type=row.source_type,
                        source_id=row.source_id,
                        description=row.description,
                        event_date=row.event_date,
                        period_start=row.period_start,
                        period_end=row.period_end,
                        currency=row.currency,
                        gross_amount_minor=row.gross_amount_minor,
                        net_amount_minor=row.net_amount_minor,
                        metadata={
                            "import_format": preview.format_name,
                            "import_row_number": row.row_number,
                        },
                    )
                )
                created_ids.append(event.id)
        return RoyaltySourceImportReport(
            format_name=preview.format_name,
            mode="apply",
            passed=len(created_ids),
            failed=preview.failed,
            skipped=preview.skipped,
            warnings=preview.warnings,
            preview_rows=preview.preview_rows,
            created_event_ids=tuple(created_ids),
        )

    @staticmethod
    def suggest_mapping(headers: list[str]) -> dict[str, str]:
        suggestions: dict[str, str] = {}
        used_targets: set[str] = set()
        for header in headers:
            target = _HEADER_ALIASES.get(_normalize_header(header))
            if target and target not in used_targets:
                suggestions[header] = target
                used_targets.add(target)
        return suggestions

    def _preview_row(
        self,
        row: Mapping[str, object],
        *,
        row_number: int,
        mapping: Mapping[str, str],
        default_contract_id: int | None,
    ) -> RoyaltySourceImportPreviewRow:
        normalized = self._mapped_row(row, mapping)
        issues: list[str] = []
        description = (
            clean_text(normalized.get("description"))
            or clean_text(normalized.get("track_title"))
            or clean_text(normalized.get("work_title"))
            or clean_text(normalized.get("release_title"))
            or self._source_period_description(normalized)
            or ""
        )
        if not description:
            issues.append("description is required")
        contract_id = _optional_int(normalized.get("contract_id"))
        if contract_id is None and default_contract_id is not None:
            contract_id = int(default_contract_id)
        work_id = _optional_int(normalized.get("work_id"))
        if work_id is None:
            work_id = self._match_optional_id(
                "Works",
                "id",
                "title",
                normalized.get("work_title"),
                label="work title",
                issues=issues,
            )
        track_id = _optional_int(normalized.get("track_id"))
        if track_id is None:
            track_id = self._match_optional_id(
                "Tracks",
                "id",
                "isrc",
                normalized.get("track_isrc"),
                label="track ISRC",
                issues=issues,
            )
        if track_id is None:
            track_id = self._match_optional_id(
                "Tracks",
                "id",
                "track_title",
                normalized.get("track_title"),
                label="track title",
                issues=issues,
            )
        release_id = _optional_int(normalized.get("release_id"))
        if release_id is None:
            release_id = self._match_optional_id(
                "Releases",
                "id",
                "title",
                normalized.get("release_title"),
                label="release title",
                issues=issues,
            )
        try:
            currency = normalize_currency(str(normalized.get("currency") or DEFAULT_CURRENCY))
        except Exception as exc:
            currency = DEFAULT_CURRENCY
            issues.append(str(exc))
        gross_amount = _amount_minor(normalized, "gross_amount", "gross_amount_minor", issues)
        net_amount = _amount_minor(normalized, "net_amount", "net_amount_minor", issues)
        has_amount_value = _has_amount_value(
            normalized,
            "gross_amount",
            "gross_amount_minor",
        ) or _has_amount_value(normalized, "net_amount", "net_amount_minor")
        is_minor_rounding_skip = False
        if gross_amount <= 0 and net_amount <= 0:
            if has_amount_value:
                issues.append("amount rounds below one minor unit; skipped")
                is_minor_rounding_skip = True
            else:
                issues.append("gross or net amount is required")
        event_date = self._normalized_date(clean_text(normalized.get("event_date")))
        period_start, period_end = self._normalized_period(
            clean_text(normalized.get("period_start")),
            clean_text(normalized.get("period_end")),
        )
        status = (
            "ready"
            if not issues
            else "skipped" if is_minor_rounding_skip and len(issues) == 1 else "error"
        )
        return RoyaltySourceImportPreviewRow(
            row_number=row_number,
            status=status,
            description=description,
            contract_id=contract_id,
            work_id=work_id,
            track_id=track_id,
            release_id=release_id,
            event_date=event_date,
            period_start=period_start,
            period_end=period_end,
            source_type=clean_text(normalized.get("source_type")),
            source_id=clean_text(normalized.get("source_id")),
            currency=currency,
            gross_amount_minor=gross_amount,
            net_amount_minor=net_amount,
            issues=tuple(issues),
        )

    def _match_optional_id(
        self,
        table_name: str,
        id_column: str,
        match_column: str,
        value: object,
        *,
        label: str,
        issues: list[str],
    ) -> int | None:
        text = clean_text(value)
        if text is None:
            return None
        try:
            rows = self.conn.execute(
                f"""
                SELECT {id_column}
                FROM {table_name}
                WHERE LOWER(TRIM({match_column})) = LOWER(TRIM(?))
                ORDER BY {id_column}
                LIMIT 2
                """,
                (text,),
            ).fetchall()
        except sqlite3.Error:
            issues.append(f"{label} cannot be matched in this profile")
            return None
        if len(rows) == 1:
            return int(rows[0][0])
        if not rows:
            issues.append(f"{label} did not match an existing record: {text}")
        else:
            issues.append(f"{label} matched multiple records: {text}")
        return None

    @staticmethod
    def _mapped_row(row: Mapping[str, object], mapping: Mapping[str, str]) -> dict[str, object]:
        normalized: dict[str, object] = {}
        for source_header, target in mapping.items():
            clean_target = str(target or "").strip()
            if clean_target not in _SUPPORTED_TARGET_VALUES:
                continue
            normalized[clean_target] = row.get(str(source_header))
            if clean_target in {"gross_amount", "net_amount"} and not clean_text(
                normalized.get("currency")
            ):
                currency = _currency_from_amount_header(source_header)
                if currency:
                    normalized["currency"] = currency
        return normalized

    @staticmethod
    def _source_period_description(row: Mapping[str, object]) -> str | None:
        source_type = clean_text(row.get("source_type"))
        source_id = clean_text(row.get("source_id"))
        period = clean_text(row.get("period_start")) or clean_text(row.get("event_date"))
        if source_type and period:
            return f"{source_type} royalty revenue {period}"
        if source_type:
            return f"{source_type} royalty revenue"
        if source_id:
            return f"Royalty source {source_id}"
        return None

    @staticmethod
    def _normalized_date(value: str | None) -> str | None:
        if value is None:
            return None
        clean = value.strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", clean):
            return clean
        if re.fullmatch(r"\d{4}-\d{2}", clean):
            return f"{clean}-01"
        return clean

    @classmethod
    def _normalized_period(
        cls,
        start_value: str | None,
        end_value: str | None,
    ) -> tuple[str | None, str | None]:
        start = cls._normalized_date(start_value)
        end = cls._normalized_date(end_value)
        if end is None and start_value is not None:
            clean = start_value.strip()
            if re.fullmatch(r"\d{4}-\d{2}", clean):
                year_text, month_text = clean.split("-", 1)
                end = (
                    f"{year_text}-{month_text}-{monthrange(int(year_text), int(month_text))[1]:02d}"
                )
        return start, end

    @staticmethod
    def _report(
        format_name: str,
        mode: str,
        warnings: list[str],
        preview_rows: tuple[RoyaltySourceImportPreviewRow, ...],
    ) -> RoyaltySourceImportReport:
        return RoyaltySourceImportReport(
            format_name=format_name,
            mode=mode,
            passed=sum(1 for row in preview_rows if row.status == "ready"),
            failed=sum(1 for row in preview_rows if row.status == "error"),
            skipped=sum(1 for row in preview_rows if row.status == "skipped"),
            warnings=warnings,
            preview_rows=preview_rows,
        )

    @staticmethod
    def _headers_from_rows(rows: list[dict[str, object]]) -> list[str]:
        headers: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    headers.append(key)
                    seen.add(key)
        return headers

    @staticmethod
    def _read_rows(path: Path) -> tuple[list[dict[str, object]], str, list[str]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return _read_csv_rows(path), "csv", []
        if suffix in {".xlsx", ".xlsm"}:
            return _read_xlsx_rows(path), "xlsx", []
        if suffix == ".xml":
            rows = _read_xml_rows(path)
            warnings = [] if rows else ["No row-like XML elements were found."]
            return rows, "xml", warnings
        raise ValueError("Royalty imports support CSV, XML, and XLSX files.")


def _normalize_header(value: object) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum() or ch == "_")


def _currency_from_amount_header(value: object) -> str | None:
    match = _AMOUNT_HEADER_CURRENCY_RE.search(_normalize_header(value))
    if match is None:
        return None
    try:
        return normalize_currency(match.group("currency").upper())
    except ValueError:
        return None


def _optional_int(value: object) -> int | None:
    text = clean_text(value)
    if text is None:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _amount_minor(
    row: Mapping[str, object],
    amount_key: str,
    minor_key: str,
    issues: list[str],
) -> int:
    minor_text = clean_text(row.get(minor_key))
    if minor_text is not None:
        try:
            return int(minor_text)
        except ValueError:
            issues.append(f"{minor_key} must be an integer")
            return 0
    amount_text = clean_text(row.get(amount_key))
    if amount_text is None:
        return 0
    try:
        return parse_money_minor(amount_text)
    except Exception as exc:
        issues.append(f"{amount_key}: {exc}")
        return 0


def _has_amount_value(row: Mapping[str, object], amount_key: str, minor_key: str) -> bool:
    return clean_text(row.get(minor_key)) is not None or clean_text(row.get(amount_key)) is not None


def _read_csv_rows(path: Path) -> list[dict[str, object]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    return [{str(key): value for key, value in row.items() if key is not None} for row in reader]


def _read_xlsx_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header_index = 0
    while header_index < len(rows) and not any(_cell_text(value) for value in rows[header_index]):
        header_index += 1
    if header_index >= len(rows):
        return []
    headers = [
        _cell_text(value) or f"Column {index + 1}" for index, value in enumerate(rows[header_index])
    ]
    parsed: list[dict[str, object]] = []
    for values in rows[header_index + 1 :]:
        if not any(_cell_text(value) for value in values):
            continue
        parsed.append(
            {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        )
    return parsed


def _read_xml_rows(path: Path) -> list[dict[str, object]]:
    root = ElementTree.parse(path).getroot()
    children = [child for child in list(root) if isinstance(child.tag, str)]
    if not children:
        return [_flatten_xml_element(root)] if root.attrib else []
    repeated_tags = _repeated_child_tags(children)
    row_elements = [child for child in children if _local_name(child.tag) in repeated_tags]
    if not row_elements:
        row_elements = children
    return [_flatten_xml_element(element) for element in row_elements]


def _repeated_child_tags(children: list[ElementTree.Element]) -> set[str]:
    counts: dict[str, int] = {}
    for child in children:
        name = _local_name(child.tag)
        counts[name] = counts.get(name, 0) + 1
    return {name for name, count in counts.items() if count > 1}


def _flatten_xml_element(element: ElementTree.Element) -> dict[str, object]:
    row: dict[str, object] = {}
    for key, value in element.attrib.items():
        row[_local_name(key)] = value
    for child in element.iter():
        if child is element or not isinstance(child.tag, str):
            continue
        name = _local_name(child.tag)
        text = clean_text(child.text)
        if text is not None and name not in row:
            row[name] = text
        for key, value in child.attrib.items():
            row[f"{name}.{_local_name(key)}"] = value
    return row


def _local_name(name: str) -> str:
    return str(name).rsplit("}", 1)[-1]


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
