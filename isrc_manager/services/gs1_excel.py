"""Excel-first GS1 export implementation using verified official workbooks."""

from __future__ import annotations

from copy import copy
from pathlib import Path

from .gs1_mapping import localize_export_value
from .gs1_models import (
    GS1ExportPreview,
    GS1ExportResult,
    GS1PreparedRecord,
    GS1TemplateProfile,
    GS1TemplateSheetProfile,
)


def _load_openpyxl():
    try:
        from openpyxl import load_workbook
    except (
        ImportError
    ) as exc:  # pragma: no cover - exercised by UI on systems without the dependency
        from .gs1_models import GS1DependencyError

        raise GS1DependencyError(
            "GS1 workbook export requires the 'openpyxl' package to be installed."
        ) from exc
    return load_workbook


class GS1ExcelExportService:
    """Writes canonical GS1 data into the detected export sheet of an official workbook."""

    def build_preview(
        self,
        template_profile: GS1TemplateProfile,
        records: list[GS1PreparedRecord],
    ) -> GS1ExportPreview:
        headers: list[str] = []
        rows: list[tuple[str, ...]] = []
        row_sheet_names: list[str] = []
        ordered_columns = self._ordered_export_columns(
            template_profile.sheet_profile(template_profile.sheet_name)
        )
        for field_name, _ in ordered_columns:
            headers.append(
                str(template_profile.matched_headers.get(field_name) or field_name).strip()
            )
        sequence_by_sheet: dict[str, int] = {}
        for record in records:
            sheet_name = template_profile.resolve_sheet_name(record.metadata.contract_number)
            sheet_profile = template_profile.sheet_profile(sheet_name)
            sequence_number = sequence_by_sheet.get(sheet_name, 0) + 1
            sequence_by_sheet[sheet_name] = sequence_number
            localized_values = self._localized_row_values(
                locale_hint=template_profile.locale_hint,
                prepared_record=record,
                sequence_number=sequence_number,
            )
            rows.append(
                tuple(
                    str(localized_values.get(field_name, "")) for field_name, _ in ordered_columns
                )
            )
            row_sheet_names.append(sheet_profile.sheet_name)
        return GS1ExportPreview(
            headers=tuple(headers),
            rows=tuple(rows),
            row_sheet_names=tuple(row_sheet_names),
        )

    def export(
        self,
        template_profile: GS1TemplateProfile,
        records: list[GS1PreparedRecord],
        output_path: str | Path,
    ) -> GS1ExportResult:
        if not records:
            raise ValueError("At least one GS1 record is required for export")

        load_workbook = _load_openpyxl()
        workbook = load_workbook(
            filename=str(template_profile.workbook_path),
            read_only=False,
            data_only=False,
            keep_vba=template_profile.workbook_path.suffix.lower() == ".xlsm",
        )
        start_rows_by_sheet: dict[str, int] = {}
        style_rows_by_sheet: dict[str, int | None] = {}
        sequence_by_sheet: dict[str, int] = {}
        row_numbers: list[int] = []
        sheet_row_numbers: dict[str, list[int]] = {}

        for record in records:
            sheet_name = template_profile.resolve_sheet_name(record.metadata.contract_number)
            sheet_profile = template_profile.sheet_profile(sheet_name)
            worksheet = workbook[sheet_name]
            if sheet_name not in start_rows_by_sheet:
                start_rows_by_sheet[sheet_name] = self._find_start_row(worksheet, sheet_profile)
                style_rows_by_sheet[sheet_name] = (
                    sheet_profile.header_row + 1
                    if worksheet.max_row >= sheet_profile.header_row + 1
                    else None
                )
            sequence_number = sequence_by_sheet.get(sheet_name, 0) + 1
            sequence_by_sheet[sheet_name] = sequence_number
            row_number = start_rows_by_sheet[sheet_name] + sequence_number - 1
            style_template_row = style_rows_by_sheet[sheet_name]
            if style_template_row is not None and row_number > worksheet.max_row:
                self._clone_row_style(worksheet, style_template_row, row_number)
            self._write_record_row(
                worksheet,
                row_number=row_number,
                sheet_profile=sheet_profile,
                locale_hint=template_profile.locale_hint,
                prepared_record=record,
                sequence_number=sequence_number,
            )
            row_numbers.append(row_number)
            sheet_row_numbers.setdefault(sheet_name, []).append(row_number)

        final_output_path = Path(output_path)
        final_output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(final_output_path))
        ordered_sheet_names = list(sheet_row_numbers)
        if len(ordered_sheet_names) == 1:
            result_sheet_name = ordered_sheet_names[0]
        else:
            result_sheet_name = ", ".join(ordered_sheet_names)
        return GS1ExportResult(
            output_path=final_output_path,
            exported_count=len(records),
            sheet_name=result_sheet_name,
            row_numbers=row_numbers,
            sheet_row_numbers={
                sheet_name: tuple(values) for sheet_name, values in sheet_row_numbers.items()
            },
        )

    def _find_start_row(self, worksheet, sheet_profile: GS1TemplateSheetProfile) -> int:
        mapped_columns = list(sheet_profile.column_map.values())
        row_number = sheet_profile.header_row + 1
        while row_number <= max(worksheet.max_row, sheet_profile.header_row + 1):
            if all(
                self._is_blank(worksheet.cell(row=row_number, column=column_index).value)
                for column_index in mapped_columns
            ):
                return row_number
            row_number += 1
        return max(worksheet.max_row + 1, sheet_profile.header_row + 1)

    @staticmethod
    def _is_blank(value) -> bool:
        return value is None or str(value).strip() == ""

    def _clone_row_style(self, worksheet, source_row: int, target_row: int) -> None:
        if source_row <= 0 or source_row > worksheet.max_row or target_row <= worksheet.max_row:
            return
        for column_index in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(row=source_row, column=column_index)
            target_cell = worksheet.cell(row=target_row, column=column_index)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
        source_dimension = worksheet.row_dimensions.get(source_row)
        if source_dimension is not None:
            worksheet.row_dimensions[target_row].height = source_dimension.height
            worksheet.row_dimensions[target_row].hidden = source_dimension.hidden

    def _write_record_row(
        self,
        worksheet,
        *,
        row_number: int,
        sheet_profile: GS1TemplateSheetProfile,
        locale_hint: str,
        prepared_record: GS1PreparedRecord,
        sequence_number: int,
    ) -> None:
        localized_values = self._localized_row_values(
            locale_hint=locale_hint,
            prepared_record=prepared_record,
            sequence_number=sequence_number,
        )
        for field_name, column_index in sheet_profile.column_map.items():
            if field_name not in localized_values:
                continue
            worksheet.cell(
                row=row_number,
                column=column_index,
                value=localized_values[field_name],
            )

    @staticmethod
    def _ordered_export_columns(sheet_profile: GS1TemplateSheetProfile) -> list[tuple[str, int]]:
        return sorted(sheet_profile.column_map.items(), key=lambda item: item[1])

    def _localized_row_values(
        self,
        *,
        locale_hint: str,
        prepared_record: GS1PreparedRecord,
        sequence_number: int,
    ) -> dict[str, str]:
        metadata = prepared_record.metadata
        raw_values = {
            "gtin_request_number": str(sequence_number),
            "status": metadata.status,
            "product_classification": metadata.product_classification,
            "consumer_unit_flag": metadata.consumer_unit_flag,
            "packaging_type": metadata.packaging_type,
            "target_market": metadata.target_market,
            "product_description": metadata.product_description,
            "language": metadata.language,
            "brand": metadata.brand,
            "subbrand": metadata.subbrand,
            "quantity": metadata.quantity,
            "unit": metadata.unit,
            "image_url": metadata.image_url,
        }
        return {
            field_name: str(localize_export_value(field_name, value, locale_hint) or "")
            for field_name, value in raw_values.items()
        }
