"""Workbook verification for user-supplied official GS1 Excel templates."""

from __future__ import annotations

from pathlib import Path
import warnings
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from .gs1_mapping import (
    detect_template_locale,
    missing_core_template_fields,
    normalize_gs1_text,
    optional_template_fields,
    resolve_header_row,
)
from .gs1_models import GS1TemplateCandidate, GS1TemplateProfile, GS1TemplateSheetProfile, GS1TemplateVerificationError


def _load_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:  # pragma: no cover - exercised by UI on systems without the dependency
        from .gs1_models import GS1DependencyError

        raise GS1DependencyError(
            "GS1 workbook support requires the 'openpyxl' package to be installed."
        ) from exc
    return load_workbook, InvalidFileException


class GS1TemplateVerificationService:
    """Verifies a workbook by scanning sheets and scoring likely GS1 export headers."""

    ALLOWED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    MAX_SCAN_ROWS = 25
    MAX_SCAN_COLUMNS = 40
    KEYWORD_MARKERS = (
        "gs1",
        "gtin",
        "gpc",
        "consumer",
        "packaging",
        "productomschrijving",
        "product description",
        "merk",
        "brand",
        "target market",
        "upload",
        "excel uploaden",
    )
    GENERIC_SHEET_NAMES = {
        "instructions",
        "instruction",
        "instructies",
        "reference data",
        "reference",
        "lookup",
        "reference lists",
        "codelijst",
        "code list",
        "template",
        "example",
    }
    SETTINGS_OPTION_FIELDS = (
        "target_market",
        "language",
        "brand",
        "subbrand",
        "packaging_type",
        "product_classification",
    )
    _XML_NS = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
        "xm": "http://schemas.microsoft.com/office/excel/2006/main",
    }

    def verify(self, workbook_path: str | Path) -> GS1TemplateProfile:
        path = Path(workbook_path)
        if not path.exists():
            raise GS1TemplateVerificationError(f"Configured GS1 workbook was not found:\n{path}")
        if path.suffix.lower() not in self.ALLOWED_SUFFIXES:
            raise GS1TemplateVerificationError(
                "The selected file is not a supported Excel workbook. Choose an .xlsx, .xlsm, .xltx, or .xltm file."
            )

        load_workbook, invalid_file_exc = _load_openpyxl()
        try:
            with warnings.catch_warnings():
                # Some official GS1 workbooks contain Excel-only data-validation extensions
                # that openpyxl warns about while reading. Verification only inspects the
                # workbook structure and never saves it, so that warning is just console noise here.
                warnings.filterwarnings(
                    "ignore",
                    message="Data Validation extension is not supported and will be removed",
                )
                workbook = load_workbook(
                    filename=str(path),
                    read_only=True,
                    data_only=False,
                    keep_vba=path.suffix.lower() == ".xlsm",
                )
        except invalid_file_exc as exc:
            raise GS1TemplateVerificationError(
                "The selected file could not be opened as an Excel workbook."
            ) from exc
        except OSError as exc:
            raise GS1TemplateVerificationError(
                f"The selected workbook could not be read:\n{exc}"
            ) from exc

        try:
            workbook_markers = self._collect_workbook_markers(workbook)
            candidates: list[GS1TemplateCandidate] = []
            for sheet in workbook.worksheets:
                candidates.extend(self._scan_sheet_candidates(sheet, workbook_markers))

            if not candidates:
                raise GS1TemplateVerificationError(
                    "This workbook does not look like a recognized GS1 upload template. "
                    "Choose the official workbook from your GS1 portal or environment."
                )

            verified_candidates = [
                candidate
                for candidate in candidates
                if not missing_core_template_fields(candidate.column_map)
            ]
            if not verified_candidates:
                best = max(
                    candidates,
                    key=lambda candidate: (
                        candidate.score,
                        self._sheet_name_priority(candidate.sheet_name),
                        -candidate.header_row,
                    ),
                )
                missing_fields = missing_core_template_fields(best.column_map)
                missing_text = ", ".join(field.replace("_", " ") for field in missing_fields)
                raise GS1TemplateVerificationError(
                    "The workbook looks close to a GS1 template, but required export columns are missing: "
                    f"{missing_text}."
                )

            best_by_sheet: dict[str, GS1TemplateCandidate] = {}
            for candidate in verified_candidates:
                current = best_by_sheet.get(candidate.sheet_name)
                if current is None or (
                    candidate.score,
                    self._sheet_name_priority(candidate.sheet_name),
                    -candidate.header_row,
                ) > (
                    current.score,
                    self._sheet_name_priority(current.sheet_name),
                    -current.header_row,
                ):
                    best_by_sheet[candidate.sheet_name] = candidate

            non_placeholder_sheets = [
                sheet_name
                for sheet_name in best_by_sheet
                if "{" not in sheet_name and "}" not in sheet_name
            ]
            if non_placeholder_sheets:
                best_by_sheet = {
                    sheet_name: candidate
                    for sheet_name, candidate in best_by_sheet.items()
                    if sheet_name in non_placeholder_sheets
                }

            sheet_profiles: dict[str, GS1TemplateSheetProfile] = {}
            merged_field_options: dict[str, list[str]] = {}
            for sheet_name, candidate in best_by_sheet.items():
                try:
                    sheet_field_options = self._extract_field_options(
                        path,
                        workbook,
                        candidate.column_map,
                        candidate.sheet_name,
                        candidate.header_row,
                    )
                except Exception:
                    sheet_field_options = {}
                for field_name, values in sheet_field_options.items():
                    bucket = merged_field_options.setdefault(field_name, [])
                    for value in values:
                        text = str(value or "").strip()
                        if text and text not in bucket:
                            bucket.append(text)
                sheet_profiles[sheet_name] = GS1TemplateSheetProfile(
                    sheet_name=candidate.sheet_name,
                    header_row=candidate.header_row,
                    column_map=dict(candidate.column_map),
                    matched_headers=dict(candidate.matched_headers),
                    score=candidate.score,
                    missing_optional_fields=optional_template_fields(candidate.column_map),
                    field_options=sheet_field_options,
                )

            best = max(
                sheet_profiles.values(),
                key=lambda profile: (
                    profile.score,
                    self._sheet_name_priority(profile.sheet_name),
                    -profile.header_row,
                ),
            )
            locale_hint = detect_template_locale(
                best.matched_headers,
                workbook_markers,
            )
            return GS1TemplateProfile(
                workbook_path=path,
                sheet_name=best.sheet_name,
                header_row=best.header_row,
                column_map=dict(best.column_map),
                matched_headers=dict(best.matched_headers),
                score=best.score,
                workbook_markers=list(workbook_markers),
                locale_hint=locale_hint,
                missing_optional_fields=tuple(best.missing_optional_fields),
                field_options={field_name: tuple(values) for field_name, values in merged_field_options.items() if values},
                sheet_profiles=sheet_profiles,
            )
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()

    def _collect_workbook_markers(self, workbook) -> list[str]:
        markers: list[str] = []
        for sheet in workbook.worksheets:
            sheet_name = str(sheet.title or "").strip()
            if sheet_name:
                markers.append(sheet_name)
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(self.MAX_SCAN_ROWS, max(1, sheet.max_row)),
                max_col=min(self.MAX_SCAN_COLUMNS, max(1, sheet.max_column)),
                values_only=True,
            ):
                for value in row:
                    text = str(value or "").strip()
                    normalized = normalize_gs1_text(text)
                    if normalized and any(keyword in normalized for keyword in self.KEYWORD_MARKERS):
                        markers.append(text)
        return markers

    def _scan_sheet_candidates(self, sheet, workbook_markers: list[str]) -> list[GS1TemplateCandidate]:
        candidates: list[GS1TemplateCandidate] = []
        max_rows = min(self.MAX_SCAN_ROWS, max(1, sheet.max_row))
        max_cols = min(self.MAX_SCAN_COLUMNS, max(1, sheet.max_column))
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True),
            start=1,
        ):
            if not any(str(value or "").strip() for value in row):
                continue
            column_map, matched_headers, score = resolve_header_row(row)
            if len(column_map) < 6:
                continue
            if "gtin_request_number" not in column_map or "product_description" not in column_map:
                continue
            bonus = 0.0
            if "status" in column_map:
                bonus += 1.0
            if "brand" in column_map:
                bonus += 1.0
            if "quantity" in column_map and "unit" in column_map:
                bonus += 1.0
            bonus += self._sheet_name_priority(sheet.title)
            if workbook_markers:
                bonus += 0.5
            candidates.append(
                GS1TemplateCandidate(
                    sheet_name=str(sheet.title or ""),
                    header_row=row_index,
                    column_map=column_map,
                    matched_headers=matched_headers,
                    score=score + bonus,
                    workbook_markers=list(workbook_markers),
                )
            )
        return candidates

    def _sheet_name_priority(self, sheet_name: str) -> float:
        normalized = normalize_gs1_text(sheet_name)
        if not normalized:
            return 0.0
        if normalized in self.GENERIC_SHEET_NAMES:
            return -2.0
        if "{" in sheet_name or "}" in sheet_name:
            return -1.0
        if any(keyword in normalized for keyword in ("contract", "upload", "product", "gtin", "gs1")):
            return 1.25
        if any(ch.isdigit() for ch in sheet_name):
            return 1.0
        return 0.0

    def _extract_field_options(
        self,
        workbook_path: Path,
        workbook,
        column_map: dict[str, int],
        sheet_name: str,
        header_row: int,
    ) -> dict[str, tuple[str, ...]]:
        field_options: dict[str, list[str]] = {}
        for field_name, values in self._validation_options_from_sheet_xml(workbook_path, workbook, column_map, sheet_name).items():
            if values:
                field_options[field_name] = list(values)

        worksheet = workbook[sheet_name]
        for field_name in self.SETTINGS_OPTION_FIELDS:
            if field_options.get(field_name):
                continue
            column_index = column_map.get(field_name)
            if column_index is None:
                continue
            existing_values = self._collect_existing_column_values(worksheet, column_index, start_row=header_row + 1)
            if existing_values:
                field_options[field_name] = existing_values

        return {field_name: tuple(values) for field_name, values in field_options.items() if values}

    def _validation_options_from_sheet_xml(
        self,
        workbook_path: Path,
        workbook,
        column_map: dict[str, int],
        sheet_name: str,
    ) -> dict[str, list[str]]:
        try:
            worksheet_xml_path = self._resolve_sheet_xml_path(workbook_path, sheet_name)
        except Exception:
            return {}
        if not worksheet_xml_path:
            return {}

        validation_entries = self._read_validation_entries(workbook_path, worksheet_xml_path)
        if not validation_entries:
            return {}

        options_by_field: dict[str, list[str]] = {}
        column_to_field = {column_index: field_name for field_name, column_index in column_map.items()}
        for sqref, formula in validation_entries:
            field_name = self._field_name_for_sqref(sqref, column_to_field)
            if not field_name:
                continue
            values = self._resolve_validation_formula_values(workbook, formula)
            if not values:
                continue
            bucket = options_by_field.setdefault(field_name, [])
            for value in values:
                if value not in bucket:
                    bucket.append(value)
        return options_by_field

    def _resolve_sheet_xml_path(self, workbook_path: Path, sheet_name: str) -> str:
        with ZipFile(workbook_path) as archive:
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            relation_id = None
            for sheet in workbook_root.findall("main:sheets/main:sheet", self._XML_NS):
                if str(sheet.attrib.get("name") or "") == sheet_name:
                    relation_id = sheet.attrib.get(f"{{{self._XML_NS['r']}}}id")
                    break
            if not relation_id:
                return ""

            rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            target = ""
            for rel in rels_root.findall("rel:Relationship", self._XML_NS):
                if rel.attrib.get("Id") == relation_id:
                    target = str(rel.attrib.get("Target") or "").strip()
                    break
        if not target:
            return ""
        if target.startswith("/"):
            return target.lstrip("/")
        return f"xl/{target.lstrip('/')}"

    def _read_validation_entries(self, workbook_path: Path, worksheet_xml_path: str) -> list[tuple[str, str]]:
        with ZipFile(workbook_path) as archive:
            worksheet_root = ET.fromstring(archive.read(worksheet_xml_path))

        entries: list[tuple[str, str]] = []
        for validation in worksheet_root.findall(".//x14:dataValidation", self._XML_NS):
            if str(validation.attrib.get("type") or "").strip().lower() != "list":
                continue
            formula_element = validation.find("x14:formula1/xm:f", self._XML_NS)
            sqref_element = validation.find("xm:sqref", self._XML_NS)
            formula = str(formula_element.text or "").strip() if formula_element is not None else ""
            sqref = str(sqref_element.text or "").strip() if sqref_element is not None else str(validation.attrib.get("sqref") or "").strip()
            if formula and sqref:
                entries.append((sqref, formula))

        for validation in worksheet_root.findall(".//main:dataValidation", self._XML_NS):
            if str(validation.attrib.get("type") or "").strip().lower() != "list":
                continue
            formula_element = validation.find("main:formula1", self._XML_NS)
            formula = str(formula_element.text or "").strip() if formula_element is not None else ""
            sqref = str(validation.attrib.get("sqref") or "").strip()
            if formula and sqref:
                entries.append((sqref, formula))
        return entries

    def _field_name_for_sqref(self, sqref: str, column_to_field: dict[int, str]) -> str:
        try:
            from openpyxl.utils import column_index_from_string
        except ImportError:
            return ""

        for token in str(sqref or "").split():
            cell_ref = token.split(":", 1)[0].replace("$", "")
            column_letters = "".join(ch for ch in cell_ref if ch.isalpha())
            if not column_letters:
                continue
            try:
                column_index = column_index_from_string(column_letters)
            except ValueError:
                continue
            field_name = column_to_field.get(column_index)
            if field_name:
                return field_name
        return ""

    def _resolve_validation_formula_values(self, workbook, formula: str) -> list[str]:
        clean_formula = str(formula or "").strip()
        if not clean_formula:
            return []
        if clean_formula.startswith("="):
            clean_formula = clean_formula[1:].strip()
        if clean_formula.startswith('"') and clean_formula.endswith('"'):
            return self._dedupe_preserve_order(part.strip() for part in clean_formula.strip('"').split(","))

        if "!" in clean_formula:
            sheet_ref, range_ref = clean_formula.rsplit("!", 1)
            sheet_name = sheet_ref.strip().strip("'").replace("''", "'")
            return self._read_cell_range_values(workbook, sheet_name, range_ref)

        defined_name = None
        try:
            defined_name = workbook.defined_names.get(clean_formula)
        except Exception:
            defined_name = None
        if defined_name is not None:
            values: list[str] = []
            for sheet_name, range_ref in defined_name.destinations:
                values.extend(self._read_cell_range_values(workbook, sheet_name, range_ref))
            return self._dedupe_preserve_order(values)
        return []

    def _read_cell_range_values(self, workbook, sheet_name: str, range_ref: str) -> list[str]:
        try:
            from openpyxl.utils import range_boundaries
        except ImportError:
            return []

        clean_range = str(range_ref or "").replace("$", "").strip()
        if not clean_range:
            return []
        try:
            worksheet = workbook[sheet_name]
        except KeyError:
            return []
        try:
            min_col, min_row, max_col, max_row = range_boundaries(clean_range)
        except ValueError:
            return []

        values: list[str] = []
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            for value in row:
                text = str(value or "").strip()
                if text:
                    values.append(text)
        return self._dedupe_preserve_order(values)

    @staticmethod
    def _collect_existing_column_values(worksheet, column_index: int, *, start_row: int = 2, max_rows: int = 512) -> list[str]:
        values: list[str] = []
        last_row = min(max(start_row, worksheet.max_row), start_row + max_rows - 1)
        for row in worksheet.iter_rows(
            min_row=start_row,
            max_row=last_row,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            text = str((row[0] if row else "") or "").strip()
            if text and text not in values:
                values.append(text)
        return values

    @staticmethod
    def _dedupe_preserve_order(values) -> list[str]:
        deduped: list[str] = []
        for value in values:
            text = str(value or "").strip()
            if text and text not in deduped:
                deduped.append(text)
        return deduped
