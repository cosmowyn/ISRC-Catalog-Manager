"""Party import/export helpers with mapping, preview, and safe matching."""

from __future__ import annotations

import csv
import json
import sqlite3
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook

from isrc_manager.domain.repertoire import clean_text, normalized_name

from .models import PartyPayload, PartyRecord
from .service import PartyService

if TYPE_CHECKING:
    from isrc_manager.services.settings_mutations import SettingsMutationService

PARTY_JSON_SCHEMA_VERSION = 1
AUTO_CSV_DELIMITERS = ",;\t|"
CSV_SNIFF_SAMPLE_SIZE = 4096
PREVIEW_ROW_LIMIT = 5

PARTY_EXCHANGE_FIELDS = [
    "id",
    "legal_name",
    "display_name",
    "artist_name",
    "artist_aliases",
    "company_name",
    "first_name",
    "middle_name",
    "last_name",
    "party_type",
    "is_owner",
    "contact_person",
    "email",
    "alternative_email",
    "phone",
    "website",
    "street_name",
    "street_number",
    "address_line1",
    "address_line2",
    "city",
    "region",
    "postal_code",
    "country",
    "bank_account_number",
    "chamber_of_commerce_number",
    "tax_id",
    "vat_number",
    "pro_affiliation",
    "pro_number",
    "ipi_cae",
    "notes",
    "profile_name",
]

PARTY_PAYLOAD_FIELDS = [
    "legal_name",
    "display_name",
    "artist_name",
    "company_name",
    "first_name",
    "middle_name",
    "last_name",
    "party_type",
    "contact_person",
    "email",
    "alternative_email",
    "phone",
    "website",
    "street_name",
    "street_number",
    "address_line1",
    "address_line2",
    "city",
    "region",
    "postal_code",
    "country",
    "bank_account_number",
    "chamber_of_commerce_number",
    "tax_id",
    "vat_number",
    "pro_affiliation",
    "pro_number",
    "ipi_cae",
    "notes",
    "profile_name",
    "artist_aliases",
]

_HEADER_ALIASES = {
    "id": ("id", "party id", "party_id"),
    "legal_name": ("legal_name", "legal name", "name", "party name"),
    "display_name": ("display_name", "display name"),
    "artist_name": ("artist_name", "artist name", "stage name"),
    "artist_aliases": ("artist_aliases", "artist aliases", "aliases", "alias names"),
    "company_name": ("company_name", "company name"),
    "first_name": ("first_name", "first name"),
    "middle_name": ("middle_name", "middle name"),
    "last_name": ("last_name", "last name", "surname"),
    "party_type": ("party_type", "party type", "roles", "party roles"),
    "is_owner": ("is_owner", "owner", "owner party", "current owner"),
    "contact_person": ("contact_person", "contact person"),
    "email": ("email", "email address"),
    "alternative_email": (
        "alternative_email",
        "alternative email",
        "alternative email address",
        "secondary email",
    ),
    "phone": ("phone", "phone number", "telephone"),
    "website": ("website", "web site", "url"),
    "street_name": ("street_name", "street name"),
    "street_number": ("street_number", "street number", "house number"),
    "address_line1": ("address_line1", "address line 1"),
    "address_line2": ("address_line2", "address line 2"),
    "city": ("city",),
    "region": ("region", "state", "province"),
    "postal_code": ("postal_code", "postal code", "zip", "zip code"),
    "country": ("country",),
    "bank_account_number": (
        "bank_account_number",
        "bank account",
        "bank account number",
        "iban",
    ),
    "chamber_of_commerce_number": (
        "chamber_of_commerce_number",
        "chamber of commerce number",
        "commerce number",
        "kvk number",
        "coc number",
    ),
    "tax_id": ("tax_id", "tax id", "tax number"),
    "vat_number": ("vat_number", "vat number", "vat / btw number", "btw number"),
    "pro_affiliation": ("pro_affiliation", "pro affiliation", "pro"),
    "pro_number": (
        "pro_number",
        "pro number",
        "relation number",
        "buma/stemra relation number",
    ),
    "ipi_cae": ("ipi_cae", "ipi", "ipi cae", "ipi/cae"),
    "notes": ("notes", "memo"),
    "profile_name": ("profile_name", "profile name"),
}


def _is_blank_like(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return clean_text(value) is None
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) == 0
    return False


def _parse_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "owner", "current owner"}


def _normalize_header_name(value: str) -> str:
    return str(value or "").strip().casefold().replace(" ", "_").replace("-", "_").replace("/", "_")


@dataclass(slots=True)
class PartyExchangeInspection:
    file_path: str
    format_name: str
    headers: list[str]
    preview_rows: list[dict[str, object]]
    suggested_mapping: dict[str, str]
    warnings: list[str] = field(default_factory=list)
    resolved_delimiter: str | None = None


@dataclass(slots=True)
class PartyImportOptions:
    mode: str = "dry_run"
    match_by_internal_id: bool = True
    match_by_legal_name: bool = True
    match_by_identity_keys: bool = True
    match_by_name_fields: bool = True
    preview_apply_mode: str | None = None


@dataclass(slots=True)
class PartyImportReport:
    format_name: str
    mode: str
    passed: int
    failed: int
    skipped: int
    warnings: list[str]
    duplicates: list[str]
    unknown_fields: list[str]
    evaluated_mode: str | None = None
    would_create_parties: int = 0
    would_update_parties: int = 0
    would_set_owner: bool = False
    created_parties: list[int] = field(default_factory=list)
    updated_parties: list[int] = field(default_factory=list)
    owner_party_id: int | None = None


class PartyExchangeService:
    """Import/export helpers for canonical Party records."""

    def __init__(
        self,
        conn: sqlite3.Connection,
        *,
        party_service: PartyService,
        settings_mutations: SettingsMutationService | None = None,
        profile_name: str | None = None,
    ):
        self.conn = conn
        self.party_service = party_service
        self.settings_mutations = settings_mutations
        self.profile_name = clean_text(profile_name)

    def supported_import_targets(self) -> list[str]:
        return list(PARTY_EXCHANGE_FIELDS)

    @staticmethod
    def _report_progress(
        progress_callback,
        value: int,
        message: str,
        *,
        maximum: int = 100,
    ) -> None:
        if progress_callback is None:
            return
        progress_callback(int(value), int(maximum), str(message or ""))

    def export_rows(
        self, party_ids: list[int] | None = None, *, progress_callback=None
    ) -> tuple[list[str], list[dict[str, object]]]:
        self._report_progress(progress_callback, 5, "Collecting Party export rows...")
        exported = self.party_service.export_rows(party_ids)
        owner_party_id = self._current_owner_party_id()
        rows: list[dict[str, object]] = []
        total_rows = max(len(exported), 1)
        for index, source_row in enumerate(exported, start=1):
            self._report_progress(
                progress_callback,
                10 + int(((index - 1) / total_rows) * 30),
                f"Preparing Party export rows ({index} of {total_rows})...",
            )
            row = {field_name: source_row.get(field_name) for field_name in PARTY_EXCHANGE_FIELDS}
            row["artist_aliases"] = list(source_row.get("artist_aliases") or [])
            row["is_owner"] = bool(int(source_row.get("id") or 0) == int(owner_party_id or 0))
            rows.append(row)
        self._report_progress(progress_callback, 40, "Party export rows prepared.")
        return list(PARTY_EXCHANGE_FIELDS), rows

    def export_csv(
        self, path: str | Path, party_ids: list[int] | None = None, progress_callback=None
    ) -> int:
        headers, rows = self.export_rows(party_ids, progress_callback=progress_callback)
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._report_progress(progress_callback, 50, "Writing Party CSV header...")
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            total_rows = max(len(rows), 1)
            for index, row in enumerate(rows, start=1):
                writer.writerow(self._serialized_row(row))
                self._report_progress(
                    progress_callback,
                    50 + int((index / total_rows) * 40),
                    f"Writing Party CSV rows ({index} of {total_rows})...",
                )
        self._report_progress(progress_callback, 90, "Party CSV data written.")
        return len(rows)

    def export_xlsx(
        self, path: str | Path, party_ids: list[int] | None = None, progress_callback=None
    ) -> int:
        headers, rows = self.export_rows(party_ids, progress_callback=progress_callback)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PartyCatalog"
        self._report_progress(progress_callback, 50, "Building Party workbook header...")
        sheet.append(headers)
        total_rows = max(len(rows), 1)
        for index, row in enumerate(rows, start=1):
            sheet.append([self._serialize_cell(row.get(header)) for header in headers])
            self._report_progress(
                progress_callback,
                50 + int((index / total_rows) * 40),
                f"Writing Party workbook rows ({index} of {total_rows})...",
            )
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        self._report_progress(progress_callback, 90, "Party workbook written.")
        return len(rows)

    def export_json(
        self, path: str | Path, party_ids: list[int] | None = None, progress_callback=None
    ) -> int:
        headers, rows = self.export_rows(party_ids, progress_callback=progress_callback)
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._report_progress(progress_callback, 50, "Serializing Party JSON payload...")
        payload = {
            "schema_version": PARTY_JSON_SCHEMA_VERSION,
            "exported_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "columns": headers,
            "rows": rows,
        }
        self._report_progress(progress_callback, 80, "Writing Party JSON file...")
        output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        self._report_progress(progress_callback, 90, "Party JSON data written.")
        return len(rows)

    def inspect_csv(
        self,
        path: str | Path,
        *,
        delimiter: str | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> PartyExchangeInspection:
        self._report_progress(progress_callback, 5, "Reading Party CSV source...")
        if cancel_callback is not None:
            cancel_callback()
        with self._open_csv_dict_reader(path, delimiter=delimiter) as (reader, resolved_delimiter):
            headers = list(reader.fieldnames or [])
            preview_rows = [dict(row) for _, row in zip(range(PREVIEW_ROW_LIMIT), reader)]
        self._report_progress(progress_callback, 60, "Building Party import preview...")
        return PartyExchangeInspection(
            file_path=str(path),
            format_name="csv",
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=self._suggest_mapping(headers),
            resolved_delimiter=resolved_delimiter,
        )

    def inspect_xlsx(
        self,
        path: str | Path,
        *,
        progress_callback=None,
        cancel_callback=None,
    ) -> PartyExchangeInspection:
        self._report_progress(progress_callback, 5, "Reading Party workbook...")
        if cancel_callback is not None:
            cancel_callback()
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in (values[0] if values else ())]
        preview_rows = [
            {header: row[index] for index, header in enumerate(headers)}
            for row in values[1 : 1 + PREVIEW_ROW_LIMIT]
        ]
        self._report_progress(progress_callback, 60, "Building Party import preview...")
        return PartyExchangeInspection(
            file_path=str(path),
            format_name="xlsx",
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=self._suggest_mapping(headers),
        )

    def inspect_json(
        self,
        path: str | Path,
        *,
        progress_callback=None,
        cancel_callback=None,
    ) -> PartyExchangeInspection:
        self._report_progress(progress_callback, 5, "Reading Party JSON source...")
        if cancel_callback is not None:
            cancel_callback()
        _headers, rows = self._load_json_rows(path)
        headers = sorted({str(key) for row in rows for key in row.keys()})
        preview_rows = [dict(row) for row in rows[:PREVIEW_ROW_LIMIT]]
        self._report_progress(progress_callback, 60, "Building Party import preview...")
        return PartyExchangeInspection(
            file_path=str(path),
            format_name="json",
            headers=headers,
            preview_rows=preview_rows,
            suggested_mapping=self._suggest_mapping(headers),
        )

    def import_csv(
        self,
        path: str | Path,
        *,
        mapping: dict[str, str] | None = None,
        options: PartyImportOptions | None = None,
        delimiter: str | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> PartyImportReport:
        self._report_progress(progress_callback, 5, "Reading Party source file...")
        if cancel_callback is not None:
            cancel_callback()
        with self._open_csv_dict_reader(path, delimiter=delimiter) as (reader, _resolved_delimiter):
            rows = [dict(row) for row in reader]
        self._report_progress(progress_callback, 20, "Parsing CSV Party rows...")
        return self._import_rows(
            rows,
            mapping=mapping,
            options=options,
            format_name="csv",
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_xlsx(
        self,
        path: str | Path,
        *,
        mapping: dict[str, str] | None = None,
        options: PartyImportOptions | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> PartyImportReport:
        self._report_progress(progress_callback, 5, "Reading Party workbook...")
        if cancel_callback is not None:
            cancel_callback()
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in (values[0] if values else ())]
        rows = [{header: row[index] for index, header in enumerate(headers)} for row in values[1:]]
        self._report_progress(progress_callback, 20, "Parsing workbook Party rows...")
        return self._import_rows(
            rows,
            mapping=mapping,
            options=options,
            format_name="xlsx",
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def import_json(
        self,
        path: str | Path,
        *,
        mapping: dict[str, str] | None = None,
        options: PartyImportOptions | None = None,
        progress_callback=None,
        cancel_callback=None,
    ) -> PartyImportReport:
        self._report_progress(progress_callback, 5, "Reading Party JSON source...")
        if cancel_callback is not None:
            cancel_callback()
        _headers, rows = self._load_json_rows(path)
        self._report_progress(progress_callback, 20, "Parsing JSON Party rows...")
        return self._import_rows(
            rows,
            mapping=mapping,
            options=options,
            format_name="json",
            progress_callback=progress_callback,
            cancel_callback=cancel_callback,
        )

    def _import_rows(
        self,
        rows: list[dict[str, object]],
        *,
        mapping: dict[str, str] | None,
        options: PartyImportOptions | None,
        format_name: str,
        progress_callback=None,
        cancel_callback=None,
    ) -> PartyImportReport:
        opts = options or PartyImportOptions()
        effective_mode = str(opts.preview_apply_mode or opts.mode or "dry_run").strip().lower()
        self._report_progress(progress_callback, 30, "Mapping and normalizing Party rows...")
        if cancel_callback is not None:
            cancel_callback()
        normalized_rows, unknown_fields = self._normalize_source_rows(rows, mapping)
        owner_row_indexes = [
            index
            for index, row in enumerate(normalized_rows, start=1)
            if _parse_boolean(row.get("is_owner"))
        ]
        conflicting_owner_rows = set(owner_row_indexes) if len(owner_row_indexes) > 1 else set()
        owner_conflict_message = None
        if len(owner_row_indexes) > 1:
            owner_conflict_message = (
                "Exactly one imported row may request current Owner reassignment per run."
            )

        passed = 0
        failed = 0
        skipped = 0
        warnings: list[str] = []
        duplicates: list[str] = []
        would_create_parties = 0
        would_update_parties = 0
        would_set_owner = False
        created_parties: list[int] = []
        updated_parties: list[int] = []
        owner_party_id: int | None = None
        total_rows = max(len(normalized_rows), 1)

        if not normalized_rows:
            return PartyImportReport(
                format_name=format_name,
                mode=opts.mode,
                passed=0,
                failed=0,
                skipped=0,
                warnings=["The selected file did not contain any importable Party rows."],
                duplicates=[],
                unknown_fields=unknown_fields,
                evaluated_mode=effective_mode,
            )

        for index, row in enumerate(normalized_rows, start=1):
            if cancel_callback is not None:
                cancel_callback()
            row_progress = 40 + int(((index - 1) / total_rows) * 50)
            self._report_progress(
                progress_callback,
                row_progress,
                f"Creating and updating Parties for row {index} of {total_rows}...",
            )
            if not any(
                not _is_blank_like(value) for key, value in row.items() if key != "is_owner"
            ):
                skipped += 1
                warnings.append(f"Row {index}: no mapped Party data was supplied.")
                continue
            if index in conflicting_owner_rows:
                failed += 1
                warnings.append(f"Row {index}: {owner_conflict_message}")
                continue
            try:
                matched_party_id = self._resolve_matching_party_id(row, options=opts)
                action, payload = self._resolve_row_action(
                    row,
                    mode=effective_mode,
                    matched_party_id=matched_party_id,
                    row_index=index,
                    warnings=warnings,
                    duplicates=duplicates,
                )
                if opts.mode == "dry_run":
                    if action == "create":
                        would_create_parties += 1
                    else:
                        would_update_parties += 1
                    if _parse_boolean(row.get("is_owner")):
                        would_set_owner = True
                    passed += 1
                    continue
                savepoint_name = f"party_import_row_{index}"
                self.conn.execute(f"SAVEPOINT {savepoint_name}")
                try:
                    if action == "create":
                        party_id = self.party_service.create_party(
                            payload, cursor=self.conn.cursor()
                        )
                        created_parties.append(int(party_id))
                    else:
                        assert matched_party_id is not None
                        self.party_service.update_party(
                            int(matched_party_id),
                            payload,
                            cursor=self.conn.cursor(),
                        )
                        party_id = int(matched_party_id)
                        updated_parties.append(party_id)
                    if _parse_boolean(row.get("is_owner")):
                        owner_party_id = int(party_id)
                    self.conn.execute(f"RELEASE SAVEPOINT {savepoint_name}")
                    passed += 1
                except Exception:
                    self.conn.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
                    self.conn.execute(f"RELEASE SAVEPOINT {savepoint_name}")
                    raise
            except _SkippedRow:
                skipped += 1
            except Exception as exc:
                failed += 1
                warnings.append(f"Row {index}: {exc}")

        if (
            opts.mode != "dry_run"
            and owner_party_id is not None
            and self.settings_mutations is not None
        ):
            self.settings_mutations.set_owner_party_id(owner_party_id)
        self._report_progress(progress_callback, 96, "Finalizing Party import...")
        self._report_progress(progress_callback, 100, "Party import complete.")

        return PartyImportReport(
            format_name=format_name,
            mode=opts.mode,
            passed=passed,
            failed=failed,
            skipped=skipped,
            warnings=warnings,
            duplicates=duplicates,
            unknown_fields=unknown_fields,
            evaluated_mode=effective_mode,
            would_create_parties=would_create_parties,
            would_update_parties=would_update_parties,
            would_set_owner=would_set_owner,
            created_parties=created_parties,
            updated_parties=updated_parties,
            owner_party_id=owner_party_id,
        )

    def _normalize_source_rows(
        self,
        rows: list[dict[str, object]],
        mapping: dict[str, str] | None,
    ) -> tuple[list[dict[str, object]], list[str]]:
        normalized_mapping = {
            str(source): str(target)
            for source, target in (mapping or {}).items()
            if str(source).strip() and str(target).strip()
        }
        if not normalized_mapping and rows:
            headers = [str(key) for key in rows[0].keys()]
            normalized_mapping = self._suggest_mapping(headers)
        normalized_rows: list[dict[str, object]] = []
        unknown_fields = sorted(
            {
                str(source)
                for row in rows
                for source in row.keys()
                if str(source) not in normalized_mapping
            }
        )
        for source_row in rows:
            normalized_row: dict[str, object] = {}
            for source_name, value in source_row.items():
                target_name = normalized_mapping.get(str(source_name))
                if not target_name:
                    continue
                normalized_row[target_name] = self._decode_value(value)
            normalized_rows.append(normalized_row)
        return normalized_rows, unknown_fields

    def _resolve_row_action(
        self,
        row: dict[str, object],
        *,
        mode: str,
        matched_party_id: int | None,
        row_index: int,
        warnings: list[str],
        duplicates: list[str],
    ) -> tuple[str, PartyPayload]:
        mode = str(mode or "dry_run").strip().lower()
        existing_record = (
            self.party_service.fetch_party(int(matched_party_id))
            if matched_party_id is not None
            else None
        )
        if mode == "create" and matched_party_id is not None:
            duplicates.append(f"Row {row_index}: matched existing party {matched_party_id}")
            raise _SkippedRow()
        if mode == "update" and matched_party_id is None:
            warnings.append(f"Row {row_index}: no safe existing Party match was found.")
            raise _SkippedRow()

        if existing_record is None:
            payload = self._build_create_payload(row, row_index=row_index, warnings=warnings)
            self._validate_payload(payload)
            return "create", payload

        if mode == "merge":
            payload = self._build_merged_payload(existing_record, row)
        else:
            payload = self._build_updated_payload(existing_record, row)
        self._validate_payload(payload, party_id=existing_record.id)
        return "update", payload

    def _build_create_payload(
        self,
        row: dict[str, object],
        *,
        row_index: int,
        warnings: list[str],
    ) -> PartyPayload:
        payload_values = self._incoming_payload_values(row)
        legal_name = clean_text(payload_values.get("legal_name"))
        if not legal_name:
            legal_name = self._derived_legal_name(payload_values)
            if legal_name:
                payload_values["legal_name"] = legal_name
                warnings.append(
                    f"Row {row_index}: legal_name was blank and was seeded from the imported identity fields."
                )
        if not legal_name:
            raise ValueError("Legal name is required.")
        if not clean_text(payload_values.get("profile_name")) and self.profile_name:
            payload_values["profile_name"] = self.profile_name
        return PartyPayload(**payload_values)

    def _build_updated_payload(
        self,
        existing_record: PartyRecord,
        row: dict[str, object],
    ) -> PartyPayload:
        data = self._payload_data_from_record(existing_record)
        incoming = self._incoming_payload_values(row)
        for field_name, value in incoming.items():
            if field_name == "artist_aliases":
                if value:
                    data[field_name] = list(value)
                continue
            if not _is_blank_like(value):
                data[field_name] = value
        legal_name = clean_text(data.get("legal_name"))
        if not legal_name:
            raise ValueError("Legal name is required.")
        return PartyPayload(**data)

    def _build_merged_payload(
        self,
        existing_record: PartyRecord,
        row: dict[str, object],
    ) -> PartyPayload:
        data = self._payload_data_from_record(existing_record)
        incoming = self._incoming_payload_values(row)
        for field_name, value in incoming.items():
            if field_name == "artist_aliases":
                merged_aliases = list(data.get(field_name) or [])
                for alias in list(value or []):
                    if alias not in merged_aliases:
                        merged_aliases.append(alias)
                data[field_name] = merged_aliases
                continue
            if _is_blank_like(data.get(field_name)) and not _is_blank_like(value):
                data[field_name] = value
        legal_name = clean_text(data.get("legal_name"))
        if not legal_name:
            raise ValueError("Legal name is required.")
        return PartyPayload(**data)

    def _incoming_payload_values(self, row: dict[str, object]) -> dict[str, object]:
        values: dict[str, object] = {}
        for field_name in PARTY_PAYLOAD_FIELDS:
            if field_name not in row:
                continue
            if field_name == "artist_aliases":
                values[field_name] = self._parse_artist_aliases(row.get(field_name))
            elif field_name == "party_type":
                values[field_name] = self.party_service._clean_party_type(
                    str(row.get(field_name) or "")
                )
            else:
                values[field_name] = clean_text(row.get(field_name))
        values.setdefault("party_type", "organization")
        values.setdefault("artist_aliases", [])
        return values

    @staticmethod
    def _payload_data_from_record(record: PartyRecord) -> dict[str, object]:
        return {
            "legal_name": record.legal_name,
            "display_name": record.display_name,
            "artist_name": record.artist_name,
            "company_name": record.company_name,
            "first_name": record.first_name,
            "middle_name": record.middle_name,
            "last_name": record.last_name,
            "party_type": record.party_type,
            "contact_person": record.contact_person,
            "email": record.email,
            "alternative_email": record.alternative_email,
            "phone": record.phone,
            "website": record.website,
            "street_name": record.street_name,
            "street_number": record.street_number,
            "address_line1": record.address_line1,
            "address_line2": record.address_line2,
            "city": record.city,
            "region": record.region,
            "postal_code": record.postal_code,
            "country": record.country,
            "bank_account_number": record.bank_account_number,
            "chamber_of_commerce_number": record.chamber_of_commerce_number,
            "tax_id": record.tax_id,
            "vat_number": record.vat_number,
            "pro_affiliation": record.pro_affiliation,
            "pro_number": record.pro_number,
            "ipi_cae": record.ipi_cae,
            "notes": record.notes,
            "profile_name": record.profile_name,
            "artist_aliases": list(record.artist_aliases),
        }

    def _validate_payload(self, payload: PartyPayload, party_id: int | None = None) -> None:
        errors = self.party_service.validate_party(
            payload, party_id=party_id, cursor=self.conn.cursor()
        )
        if errors:
            raise ValueError("\n".join(errors))

    def _resolve_matching_party_id(
        self,
        row: dict[str, object],
        *,
        options: PartyImportOptions,
    ) -> int | None:
        candidate_sets: list[set[int]] = []
        if options.match_by_internal_id:
            internal_id = self._coerce_int(row.get("id"))
            if internal_id is not None:
                ids = self._select_ids("SELECT id FROM Parties WHERE id=?", (int(internal_id),))
                if ids:
                    candidate_sets.append(ids)
        if options.match_by_legal_name:
            legal_name = clean_text(row.get("legal_name"))
            if legal_name:
                ids = self._select_ids(
                    "SELECT id FROM Parties WHERE lower(legal_name)=lower(?)",
                    (legal_name,),
                )
                if ids:
                    candidate_sets.append(ids)
        if options.match_by_identity_keys:
            identity_ids: set[int] = set()
            for column_name in (
                "email",
                "alternative_email",
                "chamber_of_commerce_number",
                "pro_number",
                "ipi_cae",
            ):
                clean_value = clean_text(row.get(column_name))
                if not clean_value:
                    continue
                if column_name in {"email", "alternative_email"}:
                    matched = self._select_ids(
                        """
                        SELECT id
                        FROM Parties
                        WHERE lower(coalesce(email, ''))=lower(?)
                           OR lower(coalesce(alternative_email, ''))=lower(?)
                        """,
                        (clean_value, clean_value),
                    )
                else:
                    matched = self._select_ids(
                        f"SELECT id FROM Parties WHERE lower(coalesce({column_name}, ''))=lower(?)",
                        (clean_value,),
                    )
                identity_ids.update(matched)
            if identity_ids:
                candidate_sets.append(identity_ids)
        if options.match_by_name_fields:
            name_ids: set[int] = set()
            for column_name in ("display_name", "artist_name", "company_name"):
                clean_value = clean_text(row.get(column_name))
                if not clean_value:
                    continue
                name_ids.update(
                    self._select_ids(
                        f"SELECT id FROM Parties WHERE lower(coalesce({column_name}, ''))=lower(?)",
                        (clean_value,),
                    )
                )
            full_name = self._joined_name(
                clean_text(row.get("first_name")),
                clean_text(row.get("middle_name")),
                clean_text(row.get("last_name")),
            )
            if full_name:
                name_ids.update(
                    self._select_ids(
                        """
                        SELECT id
                        FROM Parties
                        WHERE lower(trim(
                            coalesce(first_name, '')
                            || CASE
                                WHEN coalesce(middle_name, '') <> '' THEN ' ' || middle_name
                                ELSE ''
                            END
                            || CASE
                                WHEN coalesce(last_name, '') <> '' THEN ' ' || last_name
                                ELSE ''
                            END
                        )) = lower(?)
                        """,
                        (full_name,),
                    )
                )
            artist_aliases = list(self._parse_artist_aliases(row.get("artist_aliases")))
            if clean_text(row.get("artist_name")):
                artist_aliases.append(str(clean_text(row.get("artist_name"))))
            for alias in artist_aliases:
                normalized_alias = normalized_name(alias)
                if not normalized_alias:
                    continue
                name_ids.update(
                    self._select_ids(
                        "SELECT party_id FROM PartyArtistAliases WHERE normalized_alias=?",
                        (normalized_alias,),
                    )
                )
            if name_ids:
                candidate_sets.append(name_ids)

        if not candidate_sets:
            return None
        intersection = set(candidate_sets[0])
        for ids in candidate_sets[1:]:
            intersection &= ids
        if len(intersection) == 1:
            return int(next(iter(intersection)))

        all_candidates = sorted({party_id for ids in candidate_sets for party_id in ids})
        if not all_candidates:
            raise ValueError("The enabled Party match rules resolved to conflicting records.")
        if len(all_candidates) == 1:
            return int(all_candidates[0])
        raise ValueError(
            "The enabled Party match rules resolved to multiple existing Parties. Review the row mapping or import mode."
        )

    def _suggest_mapping(self, headers: list[str]) -> dict[str, str]:
        reverse_lookup: dict[str, str] = {}
        for target_name, aliases in _HEADER_ALIASES.items():
            for alias in aliases:
                reverse_lookup[_normalize_header_name(alias)] = target_name
        mapping: dict[str, str] = {}
        for header in headers:
            normalized = _normalize_header_name(header)
            target = reverse_lookup.get(normalized)
            if target:
                mapping[str(header)] = target
        return mapping

    def _load_json_rows(self, path: str | Path) -> tuple[list[str], list[dict[str, object]]]:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
        if isinstance(payload, list):
            rows = [dict(row) for row in payload]
            headers = sorted({str(key) for row in rows for key in row.keys()})
            return headers, rows
        if not isinstance(payload, dict):
            raise ValueError("Party JSON import expects either a rows array or a schema wrapper.")
        schema_version = int(payload.get("schema_version") or PARTY_JSON_SCHEMA_VERSION)
        if schema_version != PARTY_JSON_SCHEMA_VERSION:
            raise ValueError(
                f"Unsupported Party JSON schema version {schema_version}. Expected {PARTY_JSON_SCHEMA_VERSION}."
            )
        rows = payload.get("rows")
        if not isinstance(rows, list):
            rows = payload.get("parties")
        rows = [dict(row) for row in list(rows or [])]
        headers = list(payload.get("columns") or []) or sorted(
            {str(key) for row in rows for key in row.keys()}
        )
        return headers, rows

    def _current_owner_party_id(self) -> int | None:
        try:
            row = self.conn.execute(
                "SELECT party_id FROM ApplicationOwnerBinding WHERE id=1"
            ).fetchone()
        except sqlite3.OperationalError:
            return None
        if row is None or row[0] in (None, ""):
            return None
        return int(row[0])

    @staticmethod
    def _serialize_cell(value: object) -> object:
        if isinstance(value, (list, tuple, dict)):
            return json.dumps(value, ensure_ascii=False)
        return value

    def _serialized_row(self, row: dict[str, object]) -> dict[str, object]:
        return {key: self._serialize_cell(value) for key, value in row.items()}

    @staticmethod
    def _decode_value(value: object) -> object:
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return ""
            if text.startswith("[") or text.startswith("{"):
                try:
                    return json.loads(text)
                except Exception:
                    return value
        return value

    @staticmethod
    def _parse_artist_aliases(value: object) -> list[str]:
        if value is None:
            return []
        if isinstance(value, (list, tuple)):
            raw_values = list(value)
        else:
            text = str(value).strip()
            if not text:
                return []
            if text.startswith("["):
                try:
                    decoded = json.loads(text)
                except Exception:
                    decoded = None
                if isinstance(decoded, list):
                    raw_values = list(decoded)
                else:
                    raw_values = [part.strip() for part in text.split("|")]
            elif "|" in text:
                raw_values = [part.strip() for part in text.split("|")]
            elif ";" in text:
                raw_values = [part.strip() for part in text.split(";")]
            else:
                raw_values = [text]
        cleaned: list[str] = []
        seen: set[str] = set()
        for raw_value in raw_values:
            clean_alias = clean_text(raw_value)
            normalized_alias = normalized_name(clean_alias)
            if not clean_alias or not normalized_alias or normalized_alias in seen:
                continue
            seen.add(normalized_alias)
            cleaned.append(str(clean_alias))
        return cleaned

    @staticmethod
    def _joined_name(
        first_name: str | None,
        middle_name: str | None,
        last_name: str | None,
    ) -> str | None:
        parts = [part for part in (first_name, middle_name, last_name) if part]
        if not parts:
            return None
        return " ".join(parts)

    def _derived_legal_name(self, values: dict[str, object]) -> str | None:
        for field_name in ("display_name", "artist_name", "company_name"):
            derived = clean_text(values.get(field_name))
            if derived:
                return derived
        person_name = self._joined_name(
            clean_text(values.get("first_name")),
            clean_text(values.get("middle_name")),
            clean_text(values.get("last_name")),
        )
        if person_name:
            return person_name
        email = clean_text(values.get("email"))
        if email:
            return email
        return None

    @staticmethod
    def _coerce_int(value: object) -> int | None:
        text = str(value or "").strip()
        if not text or not text.lstrip("-").isdigit():
            return None
        return int(text)

    def _select_ids(self, sql: str, params: tuple[object, ...]) -> set[int]:
        return {
            int(row[0])
            for row in self.conn.execute(sql, params).fetchall()
            if row and row[0] not in (None, "")
        }

    @contextmanager
    def _open_csv_dict_reader(self, path: str | Path, *, delimiter: str | None = None):
        handle = Path(path).open("r", encoding="utf-8-sig", newline="")
        try:
            sample = handle.read(CSV_SNIFF_SAMPLE_SIZE)
            handle.seek(0)
            dialect, resolved_delimiter = self._csv_dialect_for_sample(sample, delimiter=delimiter)
            if dialect is None:
                reader = csv.DictReader(handle, delimiter=resolved_delimiter)
            else:
                reader = csv.DictReader(handle, dialect=dialect)
            yield reader, resolved_delimiter
        finally:
            handle.close()

    @staticmethod
    def _validate_csv_delimiter(delimiter: str | None) -> str | None:
        if delimiter is None:
            return None
        clean = str(delimiter)
        if len(clean) != 1 or clean in {"\r", "\n"}:
            raise ValueError("CSV delimiter must be a single non-newline character.")
        return clean

    def _csv_dialect_for_sample(
        self,
        sample: str,
        *,
        delimiter: str | None = None,
    ) -> tuple[type[csv.Dialect] | csv.Dialect | None, str]:
        explicit_delimiter = self._validate_csv_delimiter(delimiter)
        if explicit_delimiter is not None:
            return None, explicit_delimiter
        try:
            dialect = (
                csv.Sniffer().sniff(sample, delimiters=AUTO_CSV_DELIMITERS)
                if sample.strip()
                else csv.excel
            )
        except csv.Error:
            dialect = csv.excel
        return dialect, str(getattr(dialect, "delimiter", ",") or ",")


class _SkippedRow(ValueError):
    pass
