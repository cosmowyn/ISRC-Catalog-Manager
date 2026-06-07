import sqlite3
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from openpyxl import Workbook
from PySide6.QtCore import QSettings

from isrc_manager.services import (
    DatabaseSchemaService,
    GS1BatchValidationError,
    GS1ContractEntry,
    GS1IntegrationService,
    GS1MetadataRecord,
    GS1MetadataRepository,
    GS1PreparedRecord,
    GS1ProfileDefaults,
    GS1RecordContext,
    GS1SettingsService,
    GS1TemplateAsset,
    GS1TemplateProfile,
    GS1TemplateVerificationError,
    GS1ValidationError,
    GS1ValidationIssue,
    GS1ValidationResult,
    TrackService,
)

HEADERS = [
    "GS1 Artikelcode (GTIN)",
    "Status",
    "Productclassificatie",
    "Gaat naar de consument",
    "Verpakkings type",
    "Landen of Regio's",
    "Productomschrijving (max 300 tekens)",
    "Taal",
    "Merk",
    "Submerk",
    "Aantal",
    "Eenheid",
    "Afbeelding (max 500 tekens)",
]


def build_template(path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "GS1 article code (GTIN)"
    instructions["A2"] = "Use 1, 2, 3 in the first column to request new GTINs."
    placeholder = workbook.create_sheet("{ContractNr}")
    placeholder.append(HEADERS)
    target = workbook.create_sheet("10070050")
    target.append(HEADERS)
    workbook.save(path)


def build_multi_contract_template(path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("Instructions")["A1"] = "GS1 article code (GTIN)"
    workbook.create_sheet("10064976").append(HEADERS)
    workbook.create_sheet("10070050").append(HEADERS)
    workbook.save(path)


def make_conn():
    conn = sqlite3.connect(":memory:")
    schema = DatabaseSchemaService(conn)
    schema.init_db()
    schema.migrate_schema()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS app_kv (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """)
    conn.execute("""
        INSERT INTO Parties(id, legal_name, display_name, artist_name, party_type)
        VALUES (1, 'Main Artist', 'Main Artist', 'Main Artist', 'artist')
        """)
    conn.execute("INSERT INTO Albums(id, title) VALUES (1, 'Orbit Release')")
    conn.execute("INSERT INTO Albums(id, title) VALUES (2, 'Solar Release')")
    conn.execute("INSERT INTO Albums(id, title) VALUES (3, 'Single')")
    conn.execute("INSERT INTO Albums(id, title) VALUES (4, 'single')")
    conn.execute("""
        INSERT INTO Tracks(
            id, isrc, isrc_compact, track_title, main_artist_party_id, album_id, release_date, track_length_sec, iswc, upc, genre
        )
        VALUES(1, 'NL-ABC-26-00001', 'NLABC2600001', 'Orbit Release', 1, 1, '2026-03-14', 180, NULL, '123456789012', 'Pop')
        """)
    conn.execute("""
        INSERT INTO Tracks(
            id, isrc, isrc_compact, track_title, main_artist_party_id, album_id, release_date, track_length_sec, iswc, upc, genre
        )
        VALUES(2, 'NL-ABC-26-00002', 'NLABC2600002', 'Orbit Reprise', 1, 1, '2026-03-14', 210, NULL, '999999999999', 'Pop')
        """)
    conn.execute("""
        INSERT INTO Tracks(
            id, isrc, isrc_compact, track_title, main_artist_party_id, album_id, release_date, track_length_sec, iswc, upc, genre
        )
        VALUES(3, 'NL-ABC-26-00003', 'NLABC2600003', 'Solar Flare', 1, 2, '2026-03-15', 195, NULL, '', 'Electronic')
        """)
    conn.execute("""
        INSERT INTO Tracks(
            id, isrc, isrc_compact, track_title, main_artist_party_id, album_id, release_date, track_length_sec, iswc, upc, genre
        )
        VALUES(4, 'NL-ABC-26-00004', 'NLABC2600004', 'Standalone Echo', 1, NULL, '2026-03-16', 205, NULL, '', 'Ambient')
        """)
    conn.execute("""
        INSERT INTO Tracks(
            id, isrc, isrc_compact, track_title, main_artist_party_id, album_id, release_date, track_length_sec, iswc, upc, genre
        )
        VALUES(5, 'NL-ABC-26-00005', 'NLABC2600005', 'Night Current', 1, 3, '2026-03-17', 215, NULL, '', 'Ambient')
        """)
    conn.execute("""
        INSERT INTO Tracks(
            id, isrc, isrc_compact, track_title, main_artist_party_id, album_id, release_date, track_length_sec, iswc, upc, genre
        )
        VALUES(6, 'NL-ABC-26-00006', 'NLABC2600006', 'Silent Orbit', 1, 4, '2026-03-18', 225, NULL, '', 'Ambient')
        """)
    conn.commit()
    return conn


class GS1IntegrationServiceTests(unittest.TestCase):
    def setUp(self):
        self.conn = make_conn()
        self.tmpdir = tempfile.TemporaryDirectory()
        self.settings_path = Path(self.tmpdir.name) / "settings.ini"
        self.settings = QSettings(str(self.settings_path), QSettings.IniFormat)
        self.settings.setFallbacksEnabled(False)
        self.service = GS1IntegrationService(
            GS1MetadataRepository(self.conn),
            GS1SettingsService(self.conn, self.settings),
            TrackService(self.conn, self.tmpdir.name),
        )
        self.service.settings_service.set_profile_defaults(
            GS1ProfileDefaults(contract_number="10070050")
        )

    def tearDown(self):
        self.settings.clear()
        self.conn.close()
        self.tmpdir.cleanup()

    def test_build_context_rejects_zero_or_negative_track_ids(self):
        with self.assertRaises(ValueError):
            self.service.build_context(0)

        with self.assertRaises(ValueError):
            self.service.build_context(-1)

    def test_build_context_loads_existing_track(self):
        context = self.service.build_context(1, current_profile_path="/tmp/Orbit_Label.db")

        self.assertEqual(context.track_id, 1)
        self.assertEqual(context.track_title, "Orbit Release")
        self.assertEqual(context.album_title, "Orbit Release")
        self.assertEqual(context.artist_name, "Main Artist")
        self.assertEqual(context.upc, "123456789012")
        self.assertEqual(context.profile_label, "Orbit Label")

    def test_build_context_rejects_missing_track(self):
        with self.assertRaisesRegex(ValueError, "Track 999"):
            self.service.build_context(999)

    def test_build_default_metadata_uses_active_contract_from_settings(self):
        self.service.settings_service.set_profile_defaults(
            GS1ProfileDefaults(
                contract_number="10070050",
                target_market="Worldwide",
                language="English",
                brand="Orbit Label",
                subbrand="",
                packaging_type="Digital file",
                product_classification="Audio",
            )
        )

        record = self.service.build_default_metadata(
            1, current_profile_path="/tmp/Orbit_Label.db", window_title="Orbit Window"
        )

        self.assertEqual(record.contract_number, "10070050")

    def test_build_default_metadata_uses_window_title_unbranded_and_profile_subbrand_fallbacks(
        self,
    ):
        self.service.track_service = mock.Mock(
            fetch_track_snapshot=mock.Mock(
                return_value=SimpleNamespace(
                    track_title="",
                    album_title="",
                    artist_name="",
                    upc="",
                    release_date="",
                    catalog_number="",
                )
            )
        )
        self.service.settings_service.set_profile_defaults(GS1ProfileDefaults())

        window_record = self.service.build_default_metadata(
            7,
            current_profile_path="",
            window_title="Catalog Studio",
        )
        default_title_record = self.service.build_default_metadata(
            7,
            current_profile_path="",
            window_title="Music Catalog Manager",
        )

        self.assertEqual(window_record.brand, "Catalog Studio")
        self.assertEqual(default_title_record.brand, "UNBRANDED")
        self.assertEqual(default_title_record.product_description, "Track 7")

        self.service.settings_service.set_profile_defaults(
            GS1ProfileDefaults(brand="Parent Brand", subbrand="")
        )
        profile_record = self.service.build_default_metadata(
            1,
            current_profile_path="/tmp/Orbit_Subbrand.db",
            window_title="Ignored",
        )

        self.assertEqual(profile_record.brand, "Parent Brand")
        self.assertEqual(profile_record.subbrand, "Orbit Subbrand")

    def test_contract_and_template_wrappers_delegate_to_settings_services(self):
        contracts = (GS1ContractEntry(contract_number="1001"),)
        self.service.settings_service.set_contracts(contracts, source_path="/tmp/contracts.csv")

        self.assertEqual(self.service.load_imported_contracts(), contracts)
        self.assertEqual(self.service.load_contracts_csv_path(), "/tmp/contracts.csv")

        importer = mock.Mock(load_contracts=mock.Mock(return_value=contracts))
        self.service.contract_import_service = importer
        self.assertEqual(self.service.import_contracts_from_csv("/tmp/new.csv"), contracts)
        importer.load_contracts.assert_called_once_with("/tmp/new.csv")

        destination = Path(self.tmpdir.name) / "contracts.csv"
        exported = self.service.export_contracts_csv(destination, contracts=contracts)
        self.assertEqual(exported, destination)
        self.assertIn("1001", destination.read_text(encoding="utf-8"))

        self.assertEqual(
            self.service.save_template_path("/tmp/template.xlsx"),
            "/tmp/template.xlsx",
        )
        self.assertEqual(self.service.load_template_asset().source_path, "/tmp/template.xlsx")
        with self.assertRaisesRegex(GS1TemplateVerificationError, "No official GS1 workbook"):
            self.service.export_template_workbook(Path(self.tmpdir.name) / "template.xlsx")

    def test_load_or_create_metadata_repairs_legacy_brand_subbrand_defaults(self):
        self.service.settings_service.set_profile_defaults(
            GS1ProfileDefaults(
                target_market="Global Market",
                language="English",
                brand="Orbit Label Group",
                subbrand="Orbit Series",
                packaging_type="Digital file",
                product_classification="Audio",
            )
        )
        self.service.repository.save(
            GS1MetadataRecord(
                track_id=1,
                status="Concept",
                product_classification="Audio",
                consumer_unit_flag=True,
                packaging_type="Digital file",
                target_market="Worldwide",
                language="English",
                product_description="Orbit Release",
                brand="Orbit Label",
                subbrand="",
                quantity="1",
                unit="Each",
                image_url="",
                notes="",
                export_enabled=True,
            )
        )

        record, context, exists = self.service.load_or_create_metadata(
            1,
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertTrue(exists)
        self.assertEqual(context.profile_label, "Orbit Label")
        self.assertEqual(record.brand, "Orbit Label Group")
        self.assertEqual(record.subbrand, "Orbit Series")

    def test_legacy_default_repairs_preserve_record_when_no_defaults_are_configured(self):
        self.service.settings_service.set_profile_defaults(GS1ProfileDefaults())
        record = GS1MetadataRecord(track_id=1, brand="Existing", subbrand="Existing Subbrand")
        context = self.service.build_context(1, current_profile_path="/tmp/profile.db")

        repaired = self.service._apply_legacy_default_repairs(
            record,
            context,
            window_title="ISRC Catalog Manager",
        )

        self.assertIs(repaired, record)

    def test_legacy_default_repairs_apply_missing_contract_only(self):
        self.service.settings_service.set_profile_defaults(
            GS1ProfileDefaults(contract_number="10070050")
        )
        record = GS1MetadataRecord(track_id=1, brand="Current Brand", subbrand="Current Subbrand")
        context = self.service.build_context(1, current_profile_path="/tmp/profile.db")

        repaired = self.service._apply_legacy_default_repairs(
            record,
            context,
            window_title="ISRC Catalog Manager",
        )

        self.assertIsNot(repaired, record)
        self.assertEqual(repaired.contract_number, "10070050")
        self.assertEqual(repaired.brand, "Current Brand")
        self.assertEqual(repaired.subbrand, "Current Subbrand")

    def test_save_metadata_and_group_raise_validation_errors(self):
        invalid = GS1ValidationResult([GS1ValidationIssue("brand", "Brand is required.")])
        self.service.validation_service = mock.Mock(validate=mock.Mock(return_value=invalid))

        record = GS1MetadataRecord(track_id=1)
        group = self.service.build_metadata_groups([1])[0]

        with self.assertRaises(GS1ValidationError):
            self.service.save_metadata(record)

        with self.assertRaises(GS1ValidationError):
            self.service.save_metadata_group(group, record)

    def test_validation_and_normalization_wrappers_use_group_record(self):
        validation = GS1ValidationResult([])
        self.service.validation_service = mock.Mock(validate=mock.Mock(return_value=validation))
        group = self.service.build_metadata_groups([1, 2])[0]
        record = group.record.copy()
        record.product_description = "Custom"

        normalized = self.service.normalize_group_record(group, record)
        result = self.service.validate_group_metadata(group, record, for_export=True)
        direct_result = self.service.validate_metadata(record, for_export=False)

        self.assertEqual(normalized.track_id, group.representative_context.track_id)
        self.assertEqual(normalized.product_description, group.display_title)
        self.assertIs(result, validation)
        self.assertIs(direct_result, validation)
        self.service.validation_service.validate.assert_any_call(
            normalized,
            for_export=True,
        )

    def test_load_template_profile_reports_missing_stored_bytes_and_missing_configuration(self):
        stored_asset = GS1TemplateAsset(filename="stored-template.xlsx")
        self.service.settings_service.load_stored_template_info = mock.Mock(
            return_value=stored_asset
        )
        self.service.settings_service.load_stored_template_bytes = mock.Mock(return_value=None)

        with self.assertRaisesRegex(GS1TemplateVerificationError, "could not be loaded"):
            self.service.load_template_profile()

        self.service.settings_service.load_stored_template_info = mock.Mock(return_value=None)
        self.service.settings_service.load_template_path = mock.Mock(return_value="")

        with self.assertRaisesRegex(GS1TemplateVerificationError, "No GS1 workbook"):
            self.service.load_template_profile()

    def test_replace_template_workbook_delegates_to_import_template_workbook(self):
        asset = GS1TemplateAsset(filename="template.xlsx")
        with mock.patch.object(
            self.service,
            "import_template_workbook",
            return_value=asset,
        ) as import_template:
            self.assertIs(
                self.service.replace_template_workbook(
                    "/tmp/template.xlsx",
                    storage_mode="managed_file",
                ),
                asset,
            )

        import_template.assert_called_once_with(
            "/tmp/template.xlsx",
            storage_mode="managed_file",
        )

    def test_prepare_records_for_export_collapses_same_album_selection_to_one_product(self):
        prepared = self.service.prepare_records_for_export(
            [1, 2],
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertEqual(len(prepared), 1)
        self.assertEqual(prepared[0].metadata.product_description, "Orbit Release")
        self.assertEqual(prepared[0].source_track_ids, (1, 2))
        self.assertEqual(
            prepared[0].source_track_labels, ("Orbit Release", "Orbit Reprise (Orbit Release)")
        )

    def test_prepare_records_for_export_rejects_empty_and_invalid_groups(self):
        with self.assertRaisesRegex(ValueError, "At least one track"):
            self.service.prepare_records_for_export([])

        invalid = GS1ValidationResult([GS1ValidationIssue("brand", "Brand is required.")])
        self.service.validation_service = mock.Mock(validate=mock.Mock(return_value=invalid))

        with self.assertRaises(GS1BatchValidationError) as cm:
            self.service.prepare_records_for_export([1])

        self.assertEqual(cm.exception.issues[0].track_id, 1)
        self.assertIn("Brand is required.", cm.exception.issues[0].messages)

    def test_build_metadata_groups_split_selection_into_album_groups_and_singles(self):
        groups = self.service.build_metadata_groups(
            [1, 2, 3, 4],
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertEqual(
            [group.display_title for group in groups],
            ["Orbit Release", "Solar Release", "Standalone Echo - Single"],
        )
        self.assertEqual([group.track_ids for group in groups], [(1, 2), (3,), (4,)])
        self.assertEqual([group.mode for group in groups], ["album", "album", "single"])

    def test_prepare_records_for_export_groups_by_album_title_and_keeps_singles_separate(self):
        prepared = self.service.prepare_records_for_export(
            [1, 2, 3, 4],
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertEqual(len(prepared), 3)
        self.assertEqual(prepared[0].metadata.product_description, "Orbit Release")
        self.assertEqual(prepared[1].metadata.product_description, "Solar Release")
        self.assertEqual(prepared[2].metadata.product_description, "Standalone Echo - Single")
        self.assertEqual(prepared[0].source_track_ids, (1, 2))
        self.assertEqual(prepared[1].source_track_ids, (3,))
        self.assertEqual(prepared[2].source_track_ids, (4,))

    def test_album_title_single_is_treated_as_a_single_not_an_album_group(self):
        groups = self.service.build_metadata_groups(
            [5, 6],
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertEqual(
            [group.display_title for group in groups],
            ["Night Current - Single", "Silent Orbit - Single"],
        )
        self.assertEqual([group.mode for group in groups], ["single", "single"])
        self.assertEqual([group.track_ids for group in groups], [(5,), (6,)])

        prepared = self.service.prepare_records_for_export(
            [5, 6],
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertEqual(
            [record.metadata.product_description for record in prepared],
            ["Night Current - Single", "Silent Orbit - Single"],
        )
        self.assertEqual(prepared[0].source_track_labels, ("Night Current",))
        self.assertEqual(prepared[1].source_track_labels, ("Silent Orbit",))

    def test_save_metadata_group_applies_album_values_to_every_track_in_group(self):
        group = self.service.build_metadata_groups(
            [1, 2],
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )[0]
        record = group.record.copy()
        record.brand = "Unified Label"
        record.subbrand = "Album Series"
        record.target_market = "Worldwide"

        saved_records = self.service.save_metadata_group(group, record)

        self.assertEqual(len(saved_records), 2)
        self.assertEqual(self.service.repository.fetch_by_track_id(1).brand, "Unified Label")
        self.assertEqual(self.service.repository.fetch_by_track_id(2).brand, "Unified Label")
        self.assertEqual(
            self.service.repository.fetch_by_track_id(1).product_description, "Orbit Release"
        )
        self.assertEqual(
            self.service.repository.fetch_by_track_id(2).product_description, "Orbit Release"
        )

    def test_prepare_export_plan_includes_preview_and_upc_warning_details(self):
        template_path = Path(self.tmpdir.name) / "gs1-template.xlsx"
        build_template(template_path)

        plan = self.service.prepare_export_plan(
            [1, 2, 3, 4],
            template_path=str(template_path),
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertEqual(plan.mode, "mixed_groups")
        self.assertEqual(plan.preview.headers[0], "GS1 Artikelcode (GTIN)")
        self.assertEqual(plan.preview.rows[0][0], "1")
        self.assertIn("Orbit Release", plan.preview.rows[0])
        self.assertIn("Solar Release", plan.preview.rows[1])
        self.assertIn("Standalone Echo - Single", plan.preview.rows[2])
        self.assertTrue(plan.warnings)
        self.assertIn("UPC/EAN", plan.warnings[0])
        self.assertIn("Orbit Release: 123456789012", plan.warnings[1])
        self.assertIn("Orbit Reprise (Orbit Release): 999999999999", plan.warnings[1])

    def test_prepare_export_plan_routes_preview_rows_by_contract_sheet(self):
        template_path = Path(self.tmpdir.name) / "gs1-multi-template.xlsx"
        build_multi_contract_template(template_path)
        self.service.repository.save(
            GS1MetadataRecord(
                track_id=1,
                contract_number="10064976",
                status="Concept",
                product_classification="Audio",
                consumer_unit_flag=True,
                packaging_type="Digital file",
                target_market="Worldwide",
                language="English",
                product_description="Orbit Release",
                brand="Orbit Label",
                subbrand="Series A",
                quantity="1",
                unit="Each",
                image_url="",
                notes="",
                export_enabled=True,
            )
        )
        self.service.repository.save(
            GS1MetadataRecord(
                track_id=3,
                contract_number="10070050",
                status="Concept",
                product_classification="Audio",
                consumer_unit_flag=True,
                packaging_type="Digital file",
                target_market="Worldwide",
                language="English",
                product_description="Solar Release",
                brand="Orbit Label",
                subbrand="Series A",
                quantity="1",
                unit="Each",
                image_url="",
                notes="",
                export_enabled=True,
            )
        )

        plan = self.service.prepare_export_plan(
            [1, 3],
            template_path=str(template_path),
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertEqual(plan.preview.row_sheet_names, ("10064976", "10070050"))
        self.assertIn("contract sheets", plan.summary_lines[0])
        self.assertIn("Sheet '10064976': 1 product row(s).", plan.summary_lines)
        self.assertIn("Sheet '10070050': 1 product row(s).", plan.summary_lines)

    def test_prepare_export_plan_uses_stored_template_when_no_path_is_provided(self):
        template_path = Path(self.tmpdir.name) / "stored-gs1-template.xlsx"
        build_template(template_path)
        self.service.import_template_workbook(template_path)

        plan = self.service.prepare_export_plan(
            [1, 2],
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertTrue(plan.template_profile.stored_in_database)
        self.assertEqual(plan.template_profile.template_filename, "stored-gs1-template.xlsx")
        self.assertIsNotNone(plan.template_profile.source_bytes)
        self.assertEqual(plan.preview.rows[0][0], "1")
        self.assertIn("Orbit Release", plan.preview.rows[0])

    def test_export_plan_and_export_records_delegate_to_excel_export_service(self):
        plan = SimpleNamespace(template_profile="template", prepared_records=("record",))
        self.service.excel_export_service.export = mock.Mock(return_value="exported")

        self.assertEqual(self.service.export_plan(plan, output_path="/tmp/out.xlsx"), "exported")
        self.service.excel_export_service.export.assert_called_once_with(
            "template",
            ["record"],
            "/tmp/out.xlsx",
        )

        with (
            mock.patch.object(self.service, "prepare_export_plan", return_value=plan) as prepare,
            mock.patch.object(self.service, "export_plan", return_value="exported again") as export,
        ):
            self.assertEqual(
                self.service.export_records(
                    [1],
                    output_path="/tmp/again.xlsx",
                    template_path="/tmp/template.xlsx",
                    current_profile_path="/tmp/profile.db",
                    window_title="Window",
                ),
                "exported again",
            )

        prepare.assert_called_once_with(
            [1],
            template_path="/tmp/template.xlsx",
            current_profile_path="/tmp/profile.db",
            window_title="Window",
        )
        export.assert_called_once_with(plan, output_path="/tmp/again.xlsx")

    def test_static_summary_mode_and_warning_helpers_cover_edge_shapes(self):
        template_profile = GS1TemplateProfile(
            workbook_path=Path("template.xlsx"),
            sheet_name="Input",
            header_row=1,
            column_map={},
            matched_headers={},
            score=1.0,
            workbook_markers=[],
        )
        context = GS1RecordContext(track_id=9, track_title="Already - Single", upc="123")
        metadata = GS1MetadataRecord(track_id=9, product_description="Single Product")
        single_prepared = GS1PreparedRecord(
            metadata=metadata,
            context=context,
            source_track_ids=(9,),
            source_track_labels=("Already - Single", "Duplicate UPC"),
            source_upc_values=("123", "123"),
        )
        album_prepared = GS1PreparedRecord(
            metadata=GS1MetadataRecord(track_id=10),
            context=GS1RecordContext(track_id=10, track_title="Album Track", album_title="Album"),
            source_track_ids=(10,),
        )

        self.assertEqual(
            self.service._single_product_name("Already - Single", 9), "Already - Single"
        )
        self.assertEqual(self.service._single_product_name("", 9), "Track 9 - Single")
        self.assertEqual(self.service._export_mode([]), "single")
        self.assertEqual(self.service._export_mode([album_prepared]), "album")
        self.assertEqual(
            self.service._export_mode([album_prepared, album_prepared]),
            "album_groups",
        )
        self.assertEqual(
            self.service._export_mode([single_prepared, single_prepared]),
            "single_groups",
        )
        self.assertEqual(self.service._profile_label("/tmp/default.db"), "")
        self.assertEqual(self.service._profile_label(""), "")
        self.assertEqual(self.service._group_album_title([]), "")
        self.assertEqual(
            self.service._source_track_label(
                GS1RecordContext(track_id=11, track_title="", album_title="")
            ),
            "Track 11",
        )
        self.assertEqual(self.service._build_upc_warnings([]), [])
        self.assertEqual(
            self.service._build_upc_warnings(
                [
                    GS1PreparedRecord(
                        metadata=GS1MetadataRecord(track_id=12),
                        context=GS1RecordContext(track_id=12, track_title="No UPC"),
                    )
                ]
            ),
            [],
        )
        duplicate_warning = self.service._build_upc_warnings([single_prepared])
        self.assertIn("Already - Single: 123", duplicate_warning[1])
        self.assertIn("Duplicate UPC: 123", duplicate_warning[1])

        summary = self.service._build_export_summary(
            track_ids=[1, 2],
            prepared_records=[],
            template_profile=template_profile,
            records_by_sheet={"Input": [], "Other": []},
        )

        self.assertIn("across 2 contract sheets", summary[0])
        self.assertIn("Sheet 'Input': 0 product row(s).", summary)

    def test_load_export_entries_ignores_invalid_duplicate_and_non_positive_ids(self):
        entries = self.service._load_export_entries(
            [None, "bad", 0, -1, 1, 1, 2],
            current_profile_path="/tmp/Orbit_Label.db",
            window_title="Orbit Window",
        )

        self.assertEqual([context.track_id for _record, context in entries], [1, 2])

    def test_load_stored_template_profile_ignores_temp_cleanup_failures(self):
        profile = GS1TemplateProfile(
            workbook_path=Path("temporary.xlsx"),
            sheet_name="Input",
            header_row=1,
            column_map={},
            matched_headers={},
            score=1.0,
            workbook_markers=[],
        )
        self.service.template_verification_service = mock.Mock(
            verify=mock.Mock(return_value=profile)
        )
        asset = GS1TemplateAsset(filename="stored-template.xlsx")

        with mock.patch(
            "isrc_manager.services.gs1_integration.Path.unlink",
            side_effect=OSError("locked"),
        ):
            finalized = self.service._load_stored_template_profile(asset, b"template bytes")

        self.assertTrue(finalized.stored_in_database)
        self.assertEqual(finalized.source_bytes, b"template bytes")
        self.assertEqual(finalized.workbook_path, Path("stored-template.xlsx"))


if __name__ == "__main__":
    unittest.main()
