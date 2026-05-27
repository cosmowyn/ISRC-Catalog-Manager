import csv
import json
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook
from PySide6.QtCore import QSettings

from isrc_manager.parties import (
    PartyExchangeService,
    PartyImportOptions,
    PartyPayload,
    PartyService,
)
from isrc_manager.parties import exchange_service as party_exchange_module
from isrc_manager.services import (
    DatabaseSchemaService,
    SettingsMutationService,
    SettingsReadService,
)


class PartyExchangeServiceTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.data_root = Path(self.tmpdir.name)
        self.conn = sqlite3.connect(":memory:")
        self.conn.execute("PRAGMA foreign_keys = ON")
        schema = DatabaseSchemaService(self.conn, data_root=self.data_root)
        schema.init_db()
        schema.migrate_schema()
        settings_path = self.data_root / "party-exchange.ini"
        self.settings = QSettings(str(settings_path), QSettings.IniFormat)
        self.settings.setFallbacksEnabled(False)
        self.party_service = PartyService(self.conn)
        self.settings_reads = SettingsReadService(self.conn)
        self.settings_mutations = SettingsMutationService(self.conn, self.settings)
        self.exchange_service = PartyExchangeService(
            self.conn,
            party_service=self.party_service,
            settings_mutations=self.settings_mutations,
            profile_name="TestProfile.db",
        )

    def tearDown(self):
        self.settings.sync()
        self.conn.close()
        self.tmpdir.cleanup()

    def test_export_selected_and_full_catalog_across_supported_formats(self):
        selected_party_id = self.party_service.create_party(
            PartyPayload(
                legal_name="Aeonium Holdings B.V.",
                display_name="Aeonium",
                artist_name="Aeonium Official",
                email="hello@aeonium.test",
                artist_aliases=["Aeonium", "Lyra Cosmos"],
            )
        )
        other_party_id = self.party_service.create_party(
            PartyPayload(
                legal_name="North Star Music B.V.",
                display_name="North Star",
                chamber_of_commerce_number="COC-778899",
            )
        )
        self.settings_mutations.set_owner_party_id(other_party_id)

        json_path = self.data_root / "party-catalog.json"
        csv_path = self.data_root / "selected-parties.csv"
        xlsx_path = self.data_root / "selected-parties.xlsx"

        self.assertEqual(self.exchange_service.export_json(json_path), 2)
        self.assertEqual(self.exchange_service.export_csv(csv_path, [selected_party_id]), 1)
        self.assertEqual(self.exchange_service.export_xlsx(xlsx_path, [selected_party_id]), 1)

        payload = json.loads(json_path.read_text(encoding="utf-8"))
        self.assertEqual(payload["schema_version"], 1)
        self.assertEqual(len(payload["rows"]), 2)
        owner_rows = [row for row in payload["rows"] if row.get("is_owner")]
        self.assertEqual(len(owner_rows), 1)
        self.assertEqual(owner_rows[0]["id"], other_party_id)

        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["id"], str(selected_party_id))
        self.assertEqual(json.loads(rows[0]["artist_aliases"]), ["Aeonium", "Lyra Cosmos"])
        self.assertEqual(rows[0]["is_owner"], "False")

        workbook = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        headers = list(values[0])
        body = values[1]
        self.assertEqual(sheet.title, "PartyCatalog")
        self.assertEqual(body[headers.index("id")], selected_party_id)
        self.assertEqual(body[headers.index("legal_name")], "Aeonium Holdings B.V.")

    def test_import_csv_creates_new_party_with_mapping_and_owner_binding(self):
        csv_path = self.data_root / "parties.csv"
        csv_path.write_text(
            "\n".join(
                [
                    "Legal Name,Display Name,Artist Aliases,Email Address,Party Type,Owner",
                    'Aeonium Holdings B.V.,Aeonium,"Aeonium | Lyra Cosmos",hello@aeonium.test,artist,true',
                ]
            ),
            encoding="utf-8",
        )

        inspection = self.exchange_service.inspect_csv(csv_path)
        report = self.exchange_service.import_csv(
            csv_path,
            mapping=inspection.suggested_mapping,
            options=PartyImportOptions(mode="upsert"),
        )

        self.assertEqual(report.passed, 1)
        self.assertEqual(report.failed, 0)
        self.assertEqual(len(report.created_parties), 1)
        party = self.party_service.fetch_party(report.created_parties[0])
        assert party is not None
        self.assertEqual(party.display_name, "Aeonium")
        self.assertEqual(list(party.artist_aliases), ["Aeonium", "Lyra Cosmos"])
        self.assertEqual(party.profile_name, "TestProfile.db")
        self.assertEqual(self.settings_reads.load_owner_party_id(), party.id)

    def test_dry_run_reports_planned_party_changes_without_writing_or_binding_owner(self):
        json_path = self.data_root / "party-dry-run.json"
        json_path.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "rows": [
                        {
                            "legal_name": "Dry Run Party B.V.",
                            "display_name": "Dry Run Party",
                            "email": "hello@dry-run.test",
                            "is_owner": True,
                        }
                    ],
                },
                indent=2,
            ),
            encoding="utf-8",
        )

        inspection = self.exchange_service.inspect_json(json_path)
        report = self.exchange_service.import_json(
            json_path,
            mapping=inspection.suggested_mapping,
            options=PartyImportOptions(mode="dry_run"),
        )

        self.assertEqual(report.mode, "dry_run")
        self.assertEqual(report.failed, 0)
        self.assertEqual(report.created_parties, [])
        self.assertEqual(report.updated_parties, [])
        self.assertEqual(report.would_create_parties, 1)
        self.assertTrue(report.would_set_owner)
        self.assertEqual(self.party_service.list_parties(), [])
        self.assertIsNone(self.settings_reads.load_owner_party_id())

    def test_import_xlsx_round_trip_preserves_aliases_and_business_fields(self):
        source_party_id = self.party_service.create_party(
            PartyPayload(
                legal_name="North Star Music B.V.",
                display_name="North Star",
                alternative_email="legal@northstar.test",
                chamber_of_commerce_number="COC-998877",
                pro_number="PRO-998877",
                artist_aliases=["North Star", "NSM"],
            )
        )
        xlsx_path = self.data_root / "party-catalog.xlsx"
        self.exchange_service.export_xlsx(xlsx_path, [source_party_id])

        other_conn = sqlite3.connect(":memory:")
        other_conn.execute("PRAGMA foreign_keys = ON")
        try:
            schema = DatabaseSchemaService(other_conn, data_root=self.data_root / "other")
            schema.init_db()
            schema.migrate_schema()
            other_settings = QSettings(str(self.data_root / "other.ini"), QSettings.IniFormat)
            other_party_service = PartyService(other_conn)
            other_service = PartyExchangeService(
                other_conn,
                party_service=other_party_service,
                settings_mutations=SettingsMutationService(other_conn, other_settings),
                profile_name="Imported.db",
            )
            inspection = other_service.inspect_xlsx(xlsx_path)
            report = other_service.import_xlsx(
                xlsx_path,
                mapping=inspection.suggested_mapping,
                options=PartyImportOptions(mode="upsert"),
            )
            self.assertEqual(report.passed, 1)
            imported = other_party_service.fetch_party(report.created_parties[0])
            assert imported is not None
            self.assertEqual(imported.alternative_email, "legal@northstar.test")
            self.assertEqual(imported.chamber_of_commerce_number, "COC-998877")
            self.assertEqual(imported.pro_number, "PRO-998877")
            self.assertEqual(list(imported.artist_aliases), ["North Star", "NSM"])
        finally:
            other_conn.close()

    def test_import_json_can_update_existing_party_by_safe_match(self):
        existing_party_id = self.party_service.create_party(
            PartyPayload(
                legal_name="Aeonium Holdings B.V.",
                display_name="Old Aeonium",
                artist_name="Aeonium Official",
                email="hello@aeonium.test",
                artist_aliases=["Aeonium"],
            )
        )
        json_path = self.data_root / "party-update.json"
        json_path.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "columns": [
                        "legal_name",
                        "display_name",
                        "artist_aliases",
                        "email",
                    ],
                    "rows": [
                        {
                            "legal_name": "Aeonium Holdings B.V.",
                            "display_name": "Aeonium",
                            "artist_aliases": ["Aeonium", "Lyra Cosmos"],
                            "email": "hello@aeonium.test",
                        }
                    ],
                },
                indent=2,
            ),
            encoding="utf-8",
        )

        inspection = self.exchange_service.inspect_json(json_path)
        report = self.exchange_service.import_json(
            json_path,
            mapping=inspection.suggested_mapping,
            options=PartyImportOptions(mode="upsert"),
        )

        self.assertEqual(report.passed, 1)
        self.assertEqual(report.created_parties, [])
        self.assertEqual(report.updated_parties, [existing_party_id])
        party = self.party_service.fetch_party(existing_party_id)
        assert party is not None
        self.assertEqual(party.display_name, "Aeonium")
        self.assertEqual(list(party.artist_aliases), ["Aeonium", "Lyra Cosmos"])

    def test_import_rejects_ambiguous_matches_instead_of_guessing(self):
        self.party_service.create_party(PartyPayload(legal_name="Duplicate Name"))
        self.party_service.create_party(PartyPayload(legal_name="Duplicate Name"))
        json_path = self.data_root / "party-ambiguous.json"
        json_path.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "rows": [
                        {
                            "legal_name": "Duplicate Name",
                            "display_name": "Duplicate Name",
                        }
                    ],
                },
                indent=2,
            ),
            encoding="utf-8",
        )

        inspection = self.exchange_service.inspect_json(json_path)
        report = self.exchange_service.import_json(
            json_path,
            mapping=inspection.suggested_mapping,
            options=PartyImportOptions(mode="upsert"),
        )

        self.assertEqual(report.passed, 0)
        self.assertEqual(report.failed, 1)
        self.assertIn("multiple existing Parties", "\n".join(report.warnings))

    def test_import_rejects_multiple_owner_rows_to_preserve_single_owner_integrity(self):
        json_path = self.data_root / "party-owner-conflict.json"
        json_path.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "rows": [
                        {"legal_name": "Owner One", "is_owner": True},
                        {"legal_name": "Owner Two", "is_owner": True},
                    ],
                },
                indent=2,
            ),
            encoding="utf-8",
        )

        inspection = self.exchange_service.inspect_json(json_path)
        report = self.exchange_service.import_json(
            json_path,
            mapping=inspection.suggested_mapping,
            options=PartyImportOptions(mode="upsert"),
        )

        self.assertEqual(report.passed, 0)
        self.assertEqual(report.failed, 2)
        self.assertIsNone(self.settings_reads.load_owner_party_id())
        self.assertEqual(self.party_service.list_parties(), [])

    def test_import_reports_staged_progress_to_completion(self):
        json_path = self.data_root / "party-progress.json"
        json_path.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "rows": [
                        {
                            "legal_name": "Progress Party B.V.",
                            "display_name": "Progress Party",
                            "email": "hello@progress.test",
                        }
                    ],
                },
                indent=2,
            ),
            encoding="utf-8",
        )
        inspection = self.exchange_service.inspect_json(json_path)
        progress_events: list[tuple[int, int, str]] = []

        report = self.exchange_service.import_json(
            json_path,
            mapping=inspection.suggested_mapping,
            options=PartyImportOptions(mode="upsert"),
            progress_callback=lambda value, maximum, message: progress_events.append(
                (value, maximum, message)
            ),
        )

        self.assertEqual(report.failed, 0)
        self.assertGreaterEqual(len(progress_events), 4)
        self.assertEqual(progress_events[0][0], 5)
        self.assertEqual(progress_events[-1], (100, 100, "Party import complete."))
        self.assertTrue(
            any("Creating and updating Parties" in message for *_rest, message in progress_events)
        )

    def test_export_reports_staged_progress(self):
        party_id = self.party_service.create_party(
            PartyPayload(
                legal_name="Progress Party Export B.V.",
                display_name="Progress Party Export",
                email="hello@progress-export.test",
            )
        )
        export_path = self.data_root / "party-progress-export.json"
        progress_events: list[tuple[int, int, str]] = []

        exported = self.exchange_service.export_json(
            export_path,
            [party_id],
            progress_callback=lambda value, maximum, message: progress_events.append(
                (value, maximum, message)
            ),
        )

        self.assertEqual(exported, 1)
        self.assertTrue(export_path.exists())
        self.assertGreaterEqual(len(progress_events), 4)
        self.assertEqual(progress_events[0], (5, 100, "Collecting Party export rows..."))
        self.assertEqual(progress_events[-1], (90, 100, "Party JSON data written."))
        self.assertEqual(
            [value for value, _maximum, _message in progress_events],
            sorted(value for value, _maximum, _message in progress_events),
        )
        self.assertTrue(
            any("Preparing Party export rows" in message for *_rest, message in progress_events)
        )
        self.assertTrue(
            any("Serializing Party JSON payload" in message for *_rest, message in progress_events)
        )

    def test_party_exchange_mapping_json_alias_and_csv_edge_helpers(self):
        normalized_rows, unknown_fields = self.exchange_service._normalize_source_rows(
            [
                {
                    "Legal Name": "Mapped Party",
                    "Artist Aliases": '["Mapped", "Mapped", ""]',
                    "Mystery": "ignored",
                }
            ],
            mapping=None,
        )

        self.assertEqual(unknown_fields, ["Mystery"])
        self.assertEqual(normalized_rows[0]["legal_name"], "Mapped Party")
        self.assertEqual(normalized_rows[0]["artist_aliases"], ["Mapped", "Mapped", ""])
        self.assertEqual(self.exchange_service._decode_value(""), "")
        self.assertEqual(self.exchange_service._decode_value("{bad json"), "{bad json")
        self.assertEqual(
            self.exchange_service._parse_artist_aliases("One | one | Two"),
            ["One", "Two"],
        )
        self.assertEqual(
            self.exchange_service._parse_artist_aliases("Three; Four; three"),
            ["Three", "Four"],
        )
        self.assertEqual(self.exchange_service._parse_artist_aliases(None), [])
        self.assertEqual(self.exchange_service._joined_name(None, None, None), None)
        self.assertEqual(self.exchange_service._coerce_int("bad"), None)
        self.assertTrue(party_exchange_module._parse_boolean("current owner"))
        self.assertTrue(party_exchange_module._is_blank_like({}))

        list_json = self.data_root / "party-list.json"
        list_json.write_text('[{"legal_name": "List Party"}]', encoding="utf-8")
        self.assertEqual(self.exchange_service._load_json_rows(list_json)[0], ["legal_name"])

        parties_json = self.data_root / "party-wrapper.json"
        parties_json.write_text(
            json.dumps({"schema_version": 1, "parties": [{"legal_name": "Wrapped"}]}),
            encoding="utf-8",
        )
        self.assertEqual(
            self.exchange_service._load_json_rows(parties_json)[1][0]["legal_name"], "Wrapped"
        )

        bad_json = self.data_root / "party-bad.json"
        bad_json.write_text('"not rows"', encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "rows array"):
            self.exchange_service._load_json_rows(bad_json)

        bad_schema = self.data_root / "party-bad-schema.json"
        bad_schema.write_text(
            json.dumps({"schema_version": 999, "rows": []}),
            encoding="utf-8",
        )
        with self.assertRaisesRegex(ValueError, "Unsupported Party JSON schema"):
            self.exchange_service._load_json_rows(bad_schema)

        with self.assertRaisesRegex(ValueError, "single non-newline"):
            self.exchange_service._validate_csv_delimiter("\n")
        self.assertEqual(self.exchange_service._csv_dialect_for_sample("", delimiter=";")[1], ";")
        self.assertEqual(self.exchange_service._csv_dialect_for_sample("not,csv?")[1], ",")

    def test_party_import_row_actions_cover_empty_create_update_merge_and_savepoint_failure(self):
        empty_report = self.exchange_service._import_rows(
            [],
            mapping=None,
            options=PartyImportOptions(mode="dry_run"),
            format_name="json",
        )
        self.assertEqual(
            empty_report.warnings, ["The selected file did not contain any importable Party rows."]
        )

        existing_id = self.party_service.create_party(
            PartyPayload(
                legal_name="Merge Target B.V.",
                display_name="Existing Display",
                company_name=None,
                email="merge@test.example",
                artist_aliases=["Existing Alias"],
            )
        )

        create_duplicate = self.exchange_service._import_rows(
            [{"legal_name": "Merge Target B.V."}],
            mapping={"legal_name": "legal_name"},
            options=PartyImportOptions(mode="create"),
            format_name="json",
        )
        self.assertEqual(create_duplicate.skipped, 1)
        self.assertIn("matched existing party", create_duplicate.duplicates[0])

        update_missing = self.exchange_service._import_rows(
            [{"legal_name": "Missing Target B.V."}],
            mapping={"legal_name": "legal_name"},
            options=PartyImportOptions(mode="update"),
            format_name="json",
        )
        self.assertEqual(update_missing.skipped, 1)
        self.assertIn("no safe existing Party match", update_missing.warnings[0])

        merge_report = self.exchange_service._import_rows(
            [
                {
                    "legal_name": "Merge Target B.V.",
                    "display_name": "",
                    "company_name": "Merged Company",
                    "artist_aliases": "Existing Alias|New Alias",
                }
            ],
            mapping={
                "legal_name": "legal_name",
                "display_name": "display_name",
                "company_name": "company_name",
                "artist_aliases": "artist_aliases",
            },
            options=PartyImportOptions(mode="merge"),
            format_name="json",
        )
        self.assertEqual(merge_report.updated_parties, [existing_id])
        merged = self.party_service.fetch_party(existing_id)
        assert merged is not None
        self.assertEqual(merged.display_name, "Existing Display")
        self.assertEqual(merged.company_name, "Merged Company")
        self.assertEqual(list(merged.artist_aliases), ["Existing Alias", "New Alias"])

        with mock.patch.object(
            self.party_service,
            "create_party",
            side_effect=RuntimeError("write failed"),
        ):
            failed = self.exchange_service._import_rows(
                [{"legal_name": "Exploding Party"}],
                mapping={"legal_name": "legal_name"},
                options=PartyImportOptions(mode="upsert"),
                format_name="json",
            )
        self.assertEqual(failed.failed, 1)
        self.assertIn("write failed", failed.warnings[0])

    def test_party_matching_uses_identity_name_alias_and_conflict_guards(self):
        alias_party_id = self.party_service.create_party(
            PartyPayload(
                legal_name="Alias Holder",
                artist_name="Alias Stage",
                artist_aliases=["Alias One"],
            )
        )
        person_party_id = self.party_service.create_party(
            PartyPayload(
                legal_name="Person Holder",
                first_name="Ada",
                middle_name="L.",
                last_name="Writer",
            )
        )

        self.assertEqual(
            self.exchange_service._resolve_matching_party_id(
                {"artist_aliases": "Alias One"},
                options=PartyImportOptions(match_by_legal_name=False),
            ),
            alias_party_id,
        )
        self.assertEqual(
            self.exchange_service._resolve_matching_party_id(
                {"first_name": "Ada", "middle_name": "L.", "last_name": "Writer"},
                options=PartyImportOptions(match_by_legal_name=False),
            ),
            person_party_id,
        )

        with self.assertRaisesRegex(ValueError, "multiple existing Parties"):
            self.exchange_service._resolve_matching_party_id(
                {
                    "id": alias_party_id,
                    "legal_name": "Person Holder",
                },
                options=PartyImportOptions(
                    match_by_internal_id=True,
                    match_by_legal_name=True,
                    match_by_identity_keys=False,
                    match_by_name_fields=False,
                ),
            )
