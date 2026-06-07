import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from isrc_manager.services import (
    GS1ExcelExportService,
    GS1ExportResult,
    GS1MetadataRecord,
    GS1PreparedRecord,
    GS1RecordContext,
    GS1TemplateVerificationService,
)
from isrc_manager.services.gs1_models import GS1TemplateSheetProfile

HEADERS = [
    "GS1 Artikelcode (GTIN)",
    "Status",
    "Productclassificatie",
    "Gaat naar de consument",
    "Verpakkings type",
    "Landen of Regio's",
    "Productomschrijving (max 300 tekens)",
    "Taal",
    "Merk",
    "Submerk",
    "Aantal",
    "Eenheid",
    "Afbeelding (max 500 tekens)",
]


def build_template(path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "GS1 article code (GTIN)"
    instructions["A2"] = "Use 1, 2, 3 in the first column to request new GTINs."
    placeholder = workbook.create_sheet("{ContractNr}")
    placeholder.append(HEADERS)
    target = workbook.create_sheet("10070050")
    target.append(HEADERS)
    workbook.save(path)


def build_multi_contract_template(path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("Instructions")["A1"] = "GS1 article code (GTIN)"
    workbook.create_sheet("10064976").append(HEADERS)
    workbook.create_sheet("10070050").append(HEADERS)
    workbook.save(path)


def prepared_record(
    track_id: int, title: str, *, contract_number: str = "10070050"
) -> GS1PreparedRecord:
    return GS1PreparedRecord(
        metadata=GS1MetadataRecord(
            track_id=track_id,
            contract_number=contract_number,
            status="Active",
            product_classification="Audio",
            consumer_unit_flag=True,
            packaging_type="Digital file",
            target_market="Worldwide",
            language="English",
            product_description=title,
            brand="Orbit Label",
            subbrand="Series A",
            quantity="1",
            unit="Each",
            image_url="https://example.com/cover.png",
            notes="",
            export_enabled=True,
        ),
        context=GS1RecordContext(
            track_id=track_id,
            track_title=title,
            album_title=title,
            artist_name="Main Artist",
            upc="123456789012",
            release_date="2026-03-14",
            profile_label="Orbit Label",
        ),
    )


class GS1ExcelExportServiceTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tmpdir.name)
        self.template_path = self.root / "gs1-template.xlsx"
        build_template(self.template_path)
        self.template_profile = GS1TemplateVerificationService().verify(self.template_path)
        self.service = GS1ExcelExportService()

    def tearDown(self):
        self.tmpdir.cleanup()

    def test_single_export_writes_into_output_copy_without_mutating_template(self):
        output_path = self.root / "single-export.xlsx"

        result = self.service.export(
            self.template_profile, [prepared_record(1, "Orbit Release")], output_path
        )

        self.assertIsInstance(result, GS1ExportResult)
        self.assertEqual(result.exported_count, 1)
        self.assertEqual(result.row_numbers, [2])

        template_book = load_workbook(self.template_path, data_only=False)
        self.assertIsNone(template_book["10070050"]["A2"].value)

        output_book = load_workbook(output_path, data_only=False)
        target_sheet = output_book["10070050"]
        self.assertEqual(target_sheet["A2"].value, "1")
        self.assertEqual(target_sheet["B2"].value, "Actief")
        self.assertEqual(target_sheet["D2"].value, "Ja")
        self.assertEqual(target_sheet["G2"].value, "Orbit Release")
        self.assertEqual(target_sheet["H2"].value, "Engels")
        self.assertEqual(target_sheet["I2"].value, "Orbit Label")
        self.assertEqual(target_sheet["J2"].value, "Series A")
        self.assertEqual(target_sheet["L2"].value, "Aantal")

    def test_batch_export_smoke_and_sequence_regression(self):
        output_path = self.root / "batch-export.xlsx"
        records = [
            prepared_record(1, "Orbit Release"),
            prepared_record(2, "Solar Release"),
            prepared_record(3, "Lunar Release"),
        ]

        result = self.service.export(self.template_profile, records, output_path)

        self.assertEqual(result.exported_count, 3)
        self.assertEqual(result.row_numbers, [2, 3, 4])
        workbook = load_workbook(output_path, data_only=False)
        target_sheet = workbook["10070050"]
        self.assertEqual(target_sheet["A2"].value, "1")
        self.assertEqual(target_sheet["A3"].value, "2")
        self.assertEqual(target_sheet["A4"].value, "3")
        self.assertEqual(target_sheet["G3"].value, "Solar Release")
        self.assertEqual(target_sheet["G4"].value, "Lunar Release")

    def test_build_preview_matches_final_sheet_values(self):
        preview = self.service.build_preview(
            self.template_profile,
            [prepared_record(1, "Orbit Release"), prepared_record(2, "Solar Release")],
        )

        self.assertEqual(preview.headers[0], "GS1 Artikelcode (GTIN)")
        self.assertEqual(preview.headers[8], "Merk")
        self.assertEqual(preview.rows[0][0], "1")
        self.assertEqual(preview.rows[1][0], "2")
        self.assertEqual(preview.rows[0][1], "Actief")
        self.assertEqual(preview.rows[0][8], "Orbit Label")
        self.assertEqual(preview.rows[0][9], "Series A")
        self.assertEqual(preview.rows[1][6], "Solar Release")

    def test_export_routes_rows_by_contract_sheet_and_restarts_sequence_per_sheet(self):
        self.template_path = self.root / "multi-contract-template.xlsx"
        build_multi_contract_template(self.template_path)
        self.template_profile = GS1TemplateVerificationService().verify(self.template_path)
        output_path = self.root / "multi-contract-export.xlsx"
        records = [
            prepared_record(1, "Orbit Release", contract_number="10064976"),
            prepared_record(2, "Solar Release", contract_number="10070050"),
            prepared_record(3, "Lunar Release", contract_number="10064976"),
        ]

        result = self.service.export(self.template_profile, records, output_path)

        self.assertEqual(result.sheet_row_numbers["10064976"], (2, 3))
        self.assertEqual(result.sheet_row_numbers["10070050"], (2,))
        workbook = load_workbook(output_path, data_only=False)
        sheet_a = workbook["10064976"]
        sheet_b = workbook["10070050"]
        self.assertEqual(sheet_a["A2"].value, "1")
        self.assertEqual(sheet_a["A3"].value, "2")
        self.assertEqual(sheet_b["A2"].value, "1")
        self.assertEqual(sheet_a["G2"].value, "Orbit Release")
        self.assertEqual(sheet_a["G3"].value, "Lunar Release")
        self.assertEqual(sheet_b["G2"].value, "Solar Release")

    def test_export_supports_embedded_template_bytes_without_a_live_source_path(self):
        embedded_profile = GS1TemplateVerificationService().verify(self.template_path)
        embedded_profile.source_name = "embedded-template.xlsx"
        embedded_profile.source_label = "embedded-template.xlsx"
        embedded_profile.stored_in_database = True
        embedded_profile.source_bytes = self.template_path.read_bytes()
        embedded_profile.workbook_path = Path("embedded-template.xlsx")
        output_path = self.root / "embedded-export.xlsx"

        result = self.service.export(
            embedded_profile,
            [prepared_record(1, "Orbit Release"), prepared_record(2, "Solar Release")],
            output_path,
        )

        self.assertEqual(result.exported_count, 2)
        workbook = load_workbook(output_path, data_only=False)
        target_sheet = workbook["10070050"]
        self.assertEqual(target_sheet["A2"].value, "1")
        self.assertEqual(target_sheet["A3"].value, "2")
        self.assertEqual(target_sheet["G3"].value, "Solar Release")

    def test_export_rejects_empty_records(self):
        with self.assertRaisesRegex(ValueError, "At least one GS1 record"):
            self.service.export(self.template_profile, [], self.root / "empty.xlsx")

    def test_materialized_template_cleanup_tolerates_unlink_failure(self):
        embedded_profile = GS1TemplateVerificationService().verify(self.template_path)
        embedded_profile.source_name = "embedded-template.xlsx"
        embedded_profile.source_bytes = self.template_path.read_bytes()
        temp_path = None
        original_unlink = Path.unlink

        def _raise_for_materialized_template(path, *args, **kwargs):
            if path == temp_path:
                raise OSError("cleanup blocked")
            return original_unlink(path, *args, **kwargs)

        with mock.patch.object(Path, "unlink", _raise_for_materialized_template):
            with self.service._materialized_template_path(embedded_profile) as workbook_path:
                temp_path = workbook_path
                self.assertTrue(workbook_path.is_file())

        self.assertTrue(temp_path.is_file())
        original_unlink(temp_path)

    def test_row_discovery_style_cloning_and_unknown_field_skip(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "10070050"
        worksheet.append(HEADERS)
        worksheet.append(["existing", "data"])
        sheet_profile = self.template_profile.sheet_profile("10070050")

        self.assertEqual(self.service._find_start_row(worksheet, sheet_profile), 3)
        self.service._clone_row_style(worksheet, 0, 3)

        worksheet["A2"].font = Font(bold=True)
        worksheet["A2"].fill = PatternFill("solid", fgColor="FF0000")
        worksheet["A2"].border = Border(left=Side(style="thin"))
        worksheet["A2"].alignment = Alignment(horizontal="center")
        worksheet["A2"].protection = Protection(locked=False)
        worksheet["A2"].number_format = "0"
        worksheet.row_dimensions[2].height = 42
        worksheet.row_dimensions[2].hidden = True

        self.service._clone_row_style(worksheet, 2, 4)

        self.assertTrue(worksheet["A4"].font.bold)
        self.assertEqual(worksheet["A4"].number_format, "0")
        self.assertEqual(worksheet.row_dimensions[4].height, 42)
        self.assertTrue(worksheet.row_dimensions[4].hidden)

        custom_profile = GS1TemplateSheetProfile(
            sheet_name="10070050",
            header_row=1,
            column_map={"gtin_request_number": 1, "unknown_field": 2},
            matched_headers={},
            score=1.0,
        )
        self.service._write_record_row(
            worksheet,
            row_number=5,
            sheet_profile=custom_profile,
            locale_hint=self.template_profile.locale_hint,
            prepared_record=prepared_record(99, "Unknown field skip"),
            sequence_number=7,
        )

        self.assertEqual(worksheet["A5"].value, "7")
        self.assertIsNone(worksheet["B5"].value)


if __name__ == "__main__":
    unittest.main()
