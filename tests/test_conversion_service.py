import json
import sqlite3
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace
from unittest import mock
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from isrc_manager.conversion import ConversionService, ConversionTemplateStoreService
from isrc_manager.conversion.adapters import csv as csv_adapter_module
from isrc_manager.conversion.adapters import xlsx as xlsx_adapter_module
from isrc_manager.conversion.adapters.csv import (
    CsvSourceAdapter,
    CsvTemplateAdapter,
    _first_non_empty_row,
    _load_rows_from_bytes,
    _rows_from_text,
)
from isrc_manager.conversion.adapters.xlsx import XlsxSourceAdapter, XlsxTemplateAdapter
from isrc_manager.conversion.adapters.xml import (
    XmlSourceAdapter,
    XmlTemplateAdapter,
    _local_name,
    _lookup_by_path,
    _row_from_node,
)
from isrc_manager.conversion.models import (
    MAPPING_KIND_CONSTANT,
    MAPPING_KIND_SKIP,
    MAPPING_KIND_SOURCE,
    MAPPING_KIND_UNMAPPED,
    REQUIRED_STATUS_OPTIONAL,
    REQUIRED_STATUS_REQUIRED,
    SOURCE_MODE_FILE,
    ConversionMappingEntry,
    ConversionPreview,
    ConversionSourceProfile,
    ConversionTargetField,
    ConversionTemplateProfile,
)


class _FakeExchangeService:
    def __init__(self):
        self.calls = []

    def export_rows(self, track_ids):
        normalized = list(track_ids or [])
        self.calls.append(normalized)
        rows = [
            {
                "track_id": track_id,
                "track_title": f"Track {track_id}",
                "catalog_number": f"CAT-{track_id}",
            }
            for track_id in normalized
        ]
        return ["track_id", "track_title", "catalog_number"], rows


class _FakeSettingsReadService:
    def __init__(self, *, sena_number="", owner_values=None):
        self.sena_number = sena_number
        self.owner_values = dict(owner_values or {})

    def load_sena_number(self):
        return self.sena_number

    def load_owner_party_settings(self):
        return SimpleNamespace(**self.owner_values)


class ConversionServiceTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tmpdir.name)

    def tearDown(self):
        self.tmpdir.cleanup()

    def _write_csv(self, name: str, text: str) -> Path:
        path = self.root / name
        path.write_text(text, encoding="utf-8")
        return path

    def test_csv_template_export_preserves_dialect_and_trailing_rows(self):
        template_path = self._write_csv(
            "template.csv",
            "\nCatalog Number*;Title;Year\nTEMPLATE;Sample;2024\n\nFooter;Keep;Tail\n",
        )
        source_path = self._write_csv(
            "source.csv",
            "catalog_number,title,release_date\nCAT-001,Orbit,2025-04-06\nCAT-002,Signal,2024-11-09\n",
        )
        service = ConversionService()

        template_profile = service.inspect_template(template_path)
        source_profile = service.inspect_source_file(source_path)
        session = service.build_session(template_profile, source_profile)
        suggestions = service.suggest_mapping(session)
        session.mapping_entries = tuple(
            suggestions.get(field.field_key)
            or (
                ConversionMappingEntry(
                    target_field_key=field.field_key,
                    target_display_name=field.display_name,
                    mapping_kind=MAPPING_KIND_SOURCE,
                    source_field="release_date",
                    transform_name="date_to_year",
                )
                if field.field_key == "year"
                else ConversionMappingEntry(
                    target_field_key=field.field_key,
                    target_display_name=field.display_name,
                )
            )
            for field in template_profile.target_fields
        )

        preview = service.build_preview(session)
        self.assertEqual(preview.rendered_rows[0], ("CAT-001", "Orbit", "2025"))
        self.assertFalse(preview.blocking_issues)

        output_path = self.root / "converted.csv"
        result = service.export_preview(preview, output_path)

        self.assertEqual(result.exported_row_count, 2)
        exported_text = output_path.read_text(encoding="utf-8")
        self.assertIn("Catalog Number*;Title;Year", exported_text)
        self.assertIn("CAT-001;Orbit;2025", exported_text)
        self.assertIn("CAT-002;Signal;2024", exported_text)
        self.assertTrue(exported_text.rstrip().endswith("Footer;Keep;Tail"))

    def test_csv_adapter_helpers_cover_delimiters_bytes_and_empty_header_errors(self):
        rows, dialect = _rows_from_text("Catalog;Title\nCAT-1;Orbit\n", preferred_delimiter=";")
        self.assertEqual(rows, [["Catalog", "Title"], ["CAT-1", "Orbit"]])
        self.assertEqual(getattr(dialect, "delimiter", ""), ";")

        decoded_rows, decoded_dialect = _load_rows_from_bytes(
            b"\xef\xbb\xbfCatalog,Title\nCAT-2,Signal\n",
        )
        self.assertEqual(decoded_rows[0], ["Catalog", "Title"])
        self.assertEqual(getattr(decoded_dialect, "delimiter", ","), ",")

        with mock.patch.object(
            csv_adapter_module.csv.Sniffer,
            "sniff",
            side_effect=csv_adapter_module.csv.Error("ambiguous"),
        ):
            _rows, fallback_dialect = _rows_from_text("No obvious delimiter")
        self.assertIs(fallback_dialect, csv_adapter_module.csv.excel)

        with self.assertRaisesRegex(ValueError, "non-empty header"):
            _first_non_empty_row([[""], ["  "]])

    def test_csv_template_adapter_handles_empty_headers_no_samples_and_select_scope(self):
        template_path = self._write_csv("header-only.csv", "\nCatalog*, ,Title\n")
        adapter = CsvTemplateAdapter()

        profile = adapter.inspect_template(template_path)

        self.assertEqual(profile.chosen_scope, "csv")
        self.assertEqual(profile.adapter_state["sample_start"], 2)
        self.assertEqual(profile.adapter_state["sample_end"], 1)
        self.assertEqual(
            [(field.field_key, field.location) for field in profile.target_fields],
            [("catalog", "CSV!A2"), ("title", "CSV!C2")],
        )
        self.assertIs(adapter.select_scope(profile, "ignored"), profile)

    def test_csv_source_adapter_skips_blank_rows_and_select_scope(self):
        source_path = self._write_csv(
            "source.csv",
            "Catalog,Title\n\nCAT-1,Orbit\n,\nCAT-2,Signal\n",
        )
        adapter = CsvSourceAdapter()

        profile = adapter.inspect_source(source_path, preferred_csv_delimiter=",")

        self.assertEqual(profile.headers, ("Catalog", "Title"))
        self.assertEqual(
            profile.rows,
            (
                {"Catalog": "CAT-1", "Title": "Orbit"},
                {"Catalog": "CAT-2", "Title": "Signal"},
            ),
        )
        self.assertEqual(profile.resolved_delimiter, ",")
        self.assertIs(adapter.select_scope(profile, "anything"), profile)

    def test_csv_export_preview_uses_stored_bytes_expands_rows_and_reports_progress(self):
        template_profile = ConversionTemplateProfile(
            template_path=self.root / "stored.csv",
            format_name="csv",
            output_suffix=".csv",
            structure_label="Stored CSV",
            target_fields=(
                ConversionTargetField(
                    field_key="catalog_number",
                    display_name="Catalog Number",
                    location="CSV!D1",
                    required_status=REQUIRED_STATUS_REQUIRED,
                    metadata={"column_index": 3},
                ),
            ),
            template_signature="csv|stored|catalog_number",
            template_bytes=b"Catalog Number\nTEMPLATE\nFooter\n",
            adapter_state={
                "header_row_index": 0,
                "sample_start": 1,
                "sample_end": 1,
                "header_row": ["Catalog Number"],
            },
        )
        preview = ConversionPreview(
            template_profile=template_profile,
            source_profile=ConversionSourceProfile(
                source_mode=SOURCE_MODE_FILE,
                format_name="csv",
                source_label="memory",
                headers=("catalog_number",),
                rows=(),
                preview_rows=(),
            ),
            mapping_entries=(),
            included_row_indices=(0,),
            rendered_rows=(("CAT-9", "unused", "overflow"),),
        )
        progress: list[tuple[int, int, str]] = []

        result = CsvTemplateAdapter().export_preview(
            preview,
            self.root / "stored-output.csv",
            progress_callback=lambda value, maximum, message: progress.append(
                (value, maximum, message)
            ),
        )

        self.assertEqual(result.exported_row_count, 1)
        self.assertEqual(
            progress,
            [
                (20, 100, "Preparing CSV conversion export..."),
                (90, 100, "CSV conversion export written."),
            ],
        )
        exported = (self.root / "stored-output.csv").read_text(encoding="utf-8")
        self.assertIn(",,,CAT-9", exported)
        self.assertTrue(exported.rstrip().endswith("Footer"))

    def test_xlsx_template_export_clones_sample_row_style_for_appended_rows(self):
        template_path = self.root / "template.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Producer"
        worksheet["A1"] = "Catalog Number*"
        worksheet["B1"] = "Title"
        worksheet["C1"] = "Length"
        worksheet["A2"] = "TEMPLATE"
        worksheet["B2"] = "Sample"
        worksheet["C2"] = "00:00:00"
        worksheet["A2"].font = Font(bold=True)
        worksheet["B2"].fill = PatternFill("solid", fgColor="FFF2CC")
        workbook.create_sheet("Notes")
        workbook.save(template_path)

        source_path = self.root / "source.json"
        source_path.write_text(
            json.dumps(
                [
                    {"catalog_number": "CAT-100", "title": "First", "length_seconds": 245},
                    {"catalog_number": "CAT-200", "title": "Second", "length_seconds": 302},
                ]
            ),
            encoding="utf-8",
        )
        service = ConversionService()

        template_profile = service.inspect_template(template_path)
        source_profile = service.inspect_source_file(source_path)
        session = service.build_session(template_profile, source_profile)
        session.mapping_entries = (
            ConversionMappingEntry(
                target_field_key="catalog_number",
                target_display_name="Catalog Number*",
                mapping_kind=MAPPING_KIND_SOURCE,
                source_field="catalog_number",
            ),
            ConversionMappingEntry(
                target_field_key="title",
                target_display_name="Title",
                mapping_kind=MAPPING_KIND_SOURCE,
                source_field="title",
            ),
            ConversionMappingEntry(
                target_field_key="length",
                target_display_name="Length",
                mapping_kind=MAPPING_KIND_SOURCE,
                source_field="length_seconds",
                transform_name="duration_seconds_to_hms",
            ),
        )

        preview = service.build_preview(session)
        output_path = self.root / "converted.xlsx"
        service.export_preview(preview, output_path)

        exported = load_workbook(output_path)
        try:
            sheet = exported["Producer"]
            self.assertEqual(sheet["A2"].value, "CAT-100")
            self.assertEqual(sheet["B2"].value, "First")
            self.assertEqual(sheet["C2"].value, "00:04:05")
            self.assertEqual(sheet["A3"].value, "CAT-200")
            self.assertEqual(sheet["B3"].value, "Second")
            self.assertEqual(sheet["C3"].value, "00:05:02")
            self.assertTrue(sheet["A3"].font.bold)
            self.assertEqual(sheet["B3"].fill.fill_type, sheet["B2"].fill.fill_type)
            self.assertEqual(sheet["B3"].fill.fgColor.rgb, sheet["B2"].fill.fgColor.rgb)
            self.assertIn("Notes", exported.sheetnames)
        finally:
            exported.close()

    def test_xlsx_template_inspects_multi_sheet_scopes_and_preview_rows(self):
        template_path = self.root / "multi-template.xlsx"
        workbook = Workbook()
        blank = workbook.active
        blank.title = "Blank"
        producer = workbook.create_sheet("Producer")
        producer["A3"] = "Catalog Number*"
        producer["C3"] = "Title"
        producer["A4"] = "TEMPLATE"
        producer["C4"] = "Sample"
        producer["A6"] = "Ignored after blank"
        alternate = workbook.create_sheet("Alternate")
        alternate["B1"] = "ISRC"
        workbook.save(template_path)

        adapter = XlsxTemplateAdapter()
        profile = adapter.inspect_template(template_path)

        self.assertEqual(profile.chosen_scope, "Producer")
        self.assertEqual(
            profile.available_scopes, (("Producer", "Producer"), ("Alternate", "Alternate"))
        )
        self.assertIn("Multiple workbook sheets look writable", profile.warnings[0])
        self.assertEqual(
            [
                (field.field_key, field.display_name, field.location, field.required_status)
                for field in profile.target_fields
            ],
            [
                ("catalog_number", "Catalog Number*", "Producer!A3", REQUIRED_STATUS_REQUIRED),
                ("title", "Title", "Producer!C3", REQUIRED_STATUS_OPTIONAL),
            ],
        )
        producer_state = profile.adapter_state["sheet_profiles"]["Producer"]
        self.assertEqual(producer_state["sample_start"], 4)
        self.assertEqual(producer_state["sample_end"], 4)
        self.assertEqual(producer_state["style_row"], 4)

        alternate_profile = adapter.select_scope(profile, "Alternate")
        self.assertEqual(alternate_profile.chosen_scope, "Alternate")
        self.assertEqual(alternate_profile.target_fields[0].field_key, "isrc")
        self.assertIn("sheet:Alternate", alternate_profile.template_signature)
        self.assertIs(adapter.select_scope(profile, "Missing"), profile)

        headers, rows, xml_text, warnings, state = adapter.build_preview(
            profile,
            [{"catalog_number": "CAT-1"}],
        )
        self.assertEqual(headers, ("Catalog Number*", "Title"))
        self.assertEqual(rows, (("CAT-1", ""),))
        self.assertEqual(xml_text, "")
        self.assertEqual(warnings, ())
        self.assertEqual(state, {})

        empty_path = self.root / "empty-template.xlsx"
        empty_workbook = Workbook()
        empty_workbook.save(empty_path)
        with self.assertRaisesRegex(ValueError, "usable data sheet"):
            adapter.inspect_template(empty_path)

    def test_xlsx_template_export_from_stored_bytes_clears_stale_rows_and_reports_progress(self):
        template_path = self.root / "stored-template.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["Catalog Number*", "Title"])
        worksheet.append(["TEMPLATE", "Sample"])
        worksheet.append(["OLD-1", "Stale 1"])
        worksheet.append(["OLD-2", "Stale 2"])
        workbook.save(template_path)

        service = ConversionService()
        template_profile = service.inspect_template(template_path)
        stored_profile = replace(template_profile, template_bytes=template_path.read_bytes())
        preview = ConversionPreview(
            template_profile=stored_profile,
            source_profile=ConversionSourceProfile(
                source_mode=SOURCE_MODE_FILE,
                format_name="json",
                source_label="memory",
                headers=("catalog_number", "title"),
                rows=(),
                preview_rows=(),
            ),
            mapping_entries=(),
            included_row_indices=(0,),
            rendered_rows=(("CAT-9", "Fresh"),),
        )
        progress: list[tuple[int, int, str]] = []

        output_path = self.root / "stored-output.xlsx"
        result = XlsxTemplateAdapter().export_preview(
            preview,
            output_path,
            progress_callback=lambda value, maximum, message: progress.append(
                (value, maximum, message)
            ),
        )

        self.assertEqual(result.exported_row_count, 1)
        self.assertEqual(
            progress,
            [
                (15, 100, "Opening workbook template..."),
                (80, 100, "Writing converted workbook..."),
                (90, 100, "Workbook conversion export written."),
            ],
        )
        exported = load_workbook(output_path)
        try:
            sheet = exported["Data"]
            self.assertEqual(sheet["A2"].value, "CAT-9")
            self.assertEqual(sheet["B2"].value, "Fresh")
            self.assertIsNone(sheet["A3"].value)
            self.assertIsNone(sheet["B3"].value)
            self.assertIsNone(sheet["A4"].value)
            self.assertIsNone(sheet["B4"].value)
        finally:
            exported.close()

    def test_xlsx_template_export_supports_workbooks_without_callable_close(self):
        class _FakeCell:
            def __init__(self):
                self.value = None

        class _FakeWorksheet:
            max_row = 1

            def __init__(self):
                self.cells = {}

            def cell(self, *, row, column, value=None):
                key = (row, column)
                cell = self.cells.setdefault(key, _FakeCell())
                if value is not None:
                    cell.value = value
                return cell

        class _FakeWorkbook:
            close = None

            def __init__(self):
                self.worksheet = _FakeWorksheet()
                self.saved_path = None

            def __getitem__(self, sheet_name):
                if sheet_name != "Data":
                    raise KeyError(sheet_name)
                return self.worksheet

            def save(self, path):
                self.saved_path = path

        fake_workbook = _FakeWorkbook()
        template_path = self.root / "fake-template.xlsx"
        preview = ConversionPreview(
            template_profile=ConversionTemplateProfile(
                template_path=template_path,
                format_name="xlsx",
                output_suffix=".xlsx",
                structure_label="Fake workbook",
                target_fields=(
                    ConversionTargetField(
                        field_key="catalog_number",
                        display_name="Catalog Number",
                        location="Data!A1",
                        required_status=REQUIRED_STATUS_REQUIRED,
                        metadata={"column_index": 1},
                    ),
                ),
                template_signature="xlsx|sheet:Data|catalog_number",
                chosen_scope="Data",
                adapter_state={
                    "sheet_profiles": {
                        "Data": {
                            "header_row_index": 1,
                            "sample_start": 2,
                            "sample_end": 1,
                        }
                    }
                },
            ),
            source_profile=ConversionSourceProfile(
                source_mode=SOURCE_MODE_FILE,
                format_name="json",
                source_label="memory",
                headers=(),
                rows=(),
                preview_rows=(),
            ),
            mapping_entries=(),
            included_row_indices=(0,),
            rendered_rows=(("CAT-77",),),
        )

        with mock.patch.object(
            xlsx_adapter_module,
            "_open_workbook",
            return_value=fake_workbook,
        ):
            result = XlsxTemplateAdapter().export_preview(preview, self.root / "fake-output.xlsx")

        self.assertEqual(result.exported_row_count, 1)
        self.assertEqual(fake_workbook.worksheet.cells[(2, 1)].value, "CAT-77")
        self.assertEqual(fake_workbook.saved_path, str(self.root / "fake-output.xlsx"))

    def test_xlsx_template_row_style_clone_copies_full_style_and_respects_noop_guards(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "Styled"
        worksheet["A1"].font = Font(bold=True, color="FF0000")
        worksheet["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
        worksheet["A1"].border = Border(left=Side(style="thin", color="000000"))
        worksheet["A1"].alignment = Alignment(horizontal="center")
        worksheet["A1"].protection = Protection(locked=False)
        worksheet["A1"].number_format = "0.00"
        worksheet.row_dimensions[1].height = 33
        worksheet.row_dimensions[1].hidden = True

        adapter = XlsxTemplateAdapter()
        adapter._clone_row_style(worksheet, 0, 4)
        adapter._clone_row_style(worksheet, 1, 1)
        adapter._clone_row_style(worksheet, 1, 3)

        target = worksheet["A3"]
        self.assertTrue(target.font.bold)
        self.assertEqual(target.fill.fgColor.rgb, "00FFF2CC")
        self.assertEqual(target.border.left.style, "thin")
        self.assertEqual(target.alignment.horizontal, "center")
        self.assertFalse(target.protection.locked)
        self.assertEqual(target.number_format, "0.00")
        self.assertEqual(worksheet.row_dimensions[3].height, 33)
        self.assertTrue(worksheet.row_dimensions[3].hidden)

    def test_xlsx_source_adapter_reads_scopes_rows_and_rejects_empty_workbooks(self):
        source_path = self.root / "source.xlsx"
        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        data.append(["Catalog Number", "", "Title"])
        data.append(["CAT-1", "ignored", "First"])
        data.append([None, None, None])
        data.append(["CAT-2", "ignored", "Second"])
        alternate = workbook.create_sheet("Alternate")
        alternate.append(["ISRC"])
        alternate.append(["NLA1"])
        workbook.save(source_path)

        adapter = XlsxSourceAdapter()
        profile = adapter.inspect_source(source_path, preferred_csv_delimiter=";")

        self.assertEqual(profile.chosen_scope, "Data")
        self.assertEqual(profile.available_scopes, (("Data", "Data"), ("Alternate", "Alternate")))
        self.assertIn("Multiple source sheets look usable", profile.warnings[0])
        self.assertEqual(profile.headers, ("Catalog Number", "Title"))
        self.assertEqual(
            profile.rows,
            (
                {"Catalog Number": "CAT-1", "Title": "First"},
                {"Catalog Number": "CAT-2", "Title": "Second"},
            ),
        )
        self.assertEqual(profile.preview_rows, profile.rows[:10])
        self.assertIs(adapter.select_scope(profile, "Missing"), profile)

        alternate_profile = adapter.select_scope(profile, "Alternate")
        self.assertEqual(alternate_profile.chosen_scope, "Alternate")
        self.assertEqual(alternate_profile.headers, ("ISRC",))
        self.assertEqual(alternate_profile.rows, ({"ISRC": "NLA1"},))

        empty_path = self.root / "empty-source.xlsx"
        empty_workbook = Workbook()
        empty_workbook.save(empty_path)
        with self.assertRaisesRegex(ValueError, "usable source sheet"):
            adapter.inspect_source(empty_path)

    def test_xlsx_source_adapter_select_scope_supports_workbooks_without_callable_close(self):
        class _FakeCell:
            def __init__(self, value=None):
                self.value = value

        class _FakeWorksheet:
            max_column = 2
            max_row = 3

            def __init__(self):
                self.values = {
                    (1, 1): "Catalog Number",
                    (1, 2): "Title",
                    (2, 1): "CAT-1",
                    (2, 2): "First",
                    (3, 1): "",
                    (3, 2): "",
                }

            def cell(self, *, row, column):
                return _FakeCell(self.values.get((row, column)))

        class _FakeWorkbook:
            close = None

            def __getitem__(self, sheet_name):
                if sheet_name != "Data":
                    raise KeyError(sheet_name)
                return _FakeWorksheet()

        profile = ConversionSourceProfile(
            source_mode=SOURCE_MODE_FILE,
            format_name="xlsx",
            source_label="fake.xlsx",
            source_path=str(self.root / "fake.xlsx"),
            headers=(),
            rows=(),
            preview_rows=(),
            available_scopes=(("Data", "Data"),),
            chosen_scope="Data",
            adapter_state={
                "sheet_profiles": {
                    "Data": {
                        "header_row_index": 1,
                    }
                }
            },
        )

        with mock.patch.object(
            xlsx_adapter_module,
            "load_workbook",
            return_value=_FakeWorkbook(),
        ):
            selected = XlsxSourceAdapter().select_scope(profile, "Data")

        self.assertEqual(selected.headers, ("Catalog Number", "Title"))
        self.assertEqual(selected.rows, ({"Catalog Number": "CAT-1", "Title": "First"},))
        self.assertEqual(selected.preview_rows, selected.rows)

    def test_xml_template_export_preserves_tree_and_xml_declaration(self):
        template_path = self.root / "template.xml"
        template_path.write_text(
            """<?xml version="1.0" encoding="utf-8"?>
<Envelope>
  <Meta>
    <Sender>Catalog Manager</Sender>
  </Meta>
  <Records>
    <Record code="TEMPLATE">
      <Title>Sample</Title>
      <Artist>Template Artist</Artist>
    </Record>
    <Record code="SECOND">
      <Title>Second</Title>
      <Artist>Second Artist</Artist>
    </Record>
  </Records>
</Envelope>
""",
            encoding="utf-8",
        )
        source_path = self.root / "source.xml"
        source_path.write_text(
            """<Rows>
  <Row code="CAT-1">
    <Title>Alpha</Title>
    <Artist>One</Artist>
  </Row>
  <Row code="CAT-2">
    <Title>Beta</Title>
    <Artist>Two</Artist>
  </Row>
</Rows>
""",
            encoding="utf-8",
        )
        service = ConversionService()

        template_profile = service.inspect_template(template_path)
        source_profile = service.inspect_source_file(source_path)
        session = service.build_session(template_profile, source_profile)
        suggestions = service.suggest_mapping(session)
        session.mapping_entries = tuple(
            suggestions.get(field.field_key)
            or ConversionMappingEntry(
                target_field_key=field.field_key,
                target_display_name=field.display_name,
                mapping_kind=MAPPING_KIND_SOURCE,
                source_field=field.display_name,
            )
            for field in template_profile.target_fields
        )

        preview = service.build_preview(session)
        output_path = self.root / "converted.xml"
        service.export_preview(preview, output_path)

        xml_text = output_path.read_text(encoding="utf-8")
        self.assertTrue(xml_text.lstrip().startswith("<?xml"))
        root = ET.fromstring(xml_text)
        self.assertEqual(root.findtext("./Meta/Sender"), "Catalog Manager")
        records = root.findall("./Records/Record")
        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].attrib["code"], "CAT-1")
        self.assertEqual(records[0].findtext("Title"), "Alpha")
        self.assertEqual(records[1].attrib["code"], "CAT-2")
        self.assertEqual(records[1].findtext("Artist"), "Two")

    def test_xml_template_adapter_handles_scopes_bytes_progress_and_defensive_paths(self):
        template_path = self.root / "multi-template.xml"
        template_path.write_text(
            """<?xml version="1.0" encoding="utf-8"?>
<Envelope>
  <Records>
    <Record code="TEMPLATE">
      <Title lang="en">Sample</Title>
      <Aliases>
        <Alias>Nested</Alias>
      </Aliases>
      <Tag>First</Tag>
      <Tag>Second</Tag>
    </Record>
    <Record code="OLD">
      <Title>Old</Title>
      <Tag>Old</Tag>
    </Record>
  </Records>
  <Artists>
    <Artist id="A"><Name>Ada</Name></Artist>
    <Artist id="B"><Name>Ben</Name></Artist>
  </Artists>
</Envelope>
""",
            encoding="utf-8",
        )
        adapter = XmlTemplateAdapter()

        profile = adapter.inspect_template(template_path)

        self.assertEqual(_local_name("{urn:test}Record"), "Record")
        self.assertTrue(profile.warnings)
        self.assertIs(adapter.select_scope(profile, "missing"), profile)
        tag_scope = next(
            scope for scope, _label in profile.available_scopes if scope.endswith("Tag[1]")
        )
        self.assertEqual(adapter.select_scope(profile, tag_scope).chosen_scope, tag_scope)
        self.assertEqual(
            [field.display_name for field in profile.target_fields],
            ["@code", "Title", "Title/@lang", "Aliases/Alias", "Tag"],
        )

        root = ET.fromstring(template_path.read_text(encoding="utf-8"))
        self.assertIs(_lookup_by_path(root, ""), root)
        self.assertIsNone(_lookup_by_path(root, "/Other/Records"))
        self.assertIs(_lookup_by_path(root, "/Envelope/Records/Record[x]"), root[0][0])
        self.assertIsNone(_lookup_by_path(root, "/Envelope/Records/Record[0]"))
        self.assertIsNone(_lookup_by_path(root, "/Envelope/Records/Record[99]"))

        missing_field = ConversionTargetField(
            field_key="missing",
            display_name="Missing",
            location="/Envelope/Records/Record/Missing",
            required_status=REQUIRED_STATUS_OPTIONAL,
            metadata={"path": "Missing"},
        )
        bytes_profile = replace(
            profile,
            template_bytes=template_path.read_bytes(),
            target_fields=profile.target_fields + (missing_field,),
        )
        headers, rows, xml_text, warnings, state = adapter.build_preview(
            bytes_profile,
            [
                {
                    "code": "CAT-9",
                    "title": "Orbit",
                    "title_lang": "nl",
                    "tag": "Focus",
                    "aliases_alias": "Alias One",
                    "missing": "ignored",
                }
            ],
        )

        self.assertEqual(headers[-1], "Missing")
        self.assertEqual(rows[0][headers.index("Title")], "Orbit")
        self.assertEqual(warnings, ())
        self.assertEqual(state, {})
        self.assertTrue(xml_text.lstrip().startswith("<?xml"))
        rendered_root = ET.fromstring(xml_text)
        rendered_records = rendered_root.findall("./Records/Record")
        self.assertEqual(len(rendered_records), 1)
        self.assertEqual(rendered_records[0].attrib["code"], "CAT-9")
        self.assertEqual(rendered_records[0].findtext("Title"), "Orbit")
        self.assertEqual(rendered_records[0].find("Title").attrib["lang"], "nl")
        self.assertEqual(rendered_records[0].findtext("./Aliases/Alias"), "Alias One")
        self.assertIsNone(rendered_records[0].find("Missing"))

        preview = ConversionPreview(
            template_profile=bytes_profile,
            source_profile=ConversionSourceProfile(
                source_mode=SOURCE_MODE_FILE,
                format_name="xml",
                source_label="memory",
                headers=(),
                rows=(),
                preview_rows=(),
            ),
            mapping_entries=(),
            included_row_indices=(0,),
            rendered_field_rows=({"title": "Orbit"},),
            rendered_xml_text=xml_text,
        )
        progress: list[tuple[int, int, str]] = []
        result = adapter.export_preview(
            preview,
            self.root / "rendered.xml",
            progress_callback=lambda value, maximum, message: progress.append(
                (value, maximum, message)
            ),
        )

        self.assertEqual(result.exported_row_count, 1)
        self.assertEqual(
            progress,
            [
                (20, 100, "Preparing XML conversion export..."),
                (90, 100, "XML conversion export written."),
            ],
        )
        self.assertEqual((self.root / "rendered.xml").read_text(encoding="utf-8"), xml_text)

        bad_profile = replace(profile, chosen_scope="/Envelope/Records/Missing")
        with self.assertRaisesRegex(ValueError, "repeat-node path"):
            adapter._render_xml_text(bad_profile, [])

        empty_template = self.root / "empty-template.xml"
        empty_template.write_text("<Envelope />", encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "repeat-node candidate"):
            adapter.inspect_template(empty_template)

    def test_xml_template_adapter_uses_single_nested_node_when_no_repeats_exist(self):
        template_path = self.root / "single-template.xml"
        template_path.write_text(
            '<Envelope><Record code="ONE"><Title>Single</Title></Record></Envelope>',
            encoding="utf-8",
        )

        profile = XmlTemplateAdapter().inspect_template(template_path)

        self.assertEqual(profile.chosen_scope, "/Envelope/Record")
        self.assertEqual(profile.warnings, ())
        self.assertEqual(
            [field.display_name for field in profile.target_fields], ["@code", "Title"]
        )

    def test_xml_source_adapter_rows_scopes_and_defensive_paths(self):
        source_path = self.root / "source-adapter.xml"
        source_path.write_text(
            """<Envelope>
  <Records>
    <Record code="CAT-1"><Title lang="en">Orbit</Title></Record>
    <Record code="CAT-2"><Title>Signal</Title></Record>
  </Records>
  <Artists>
    <Artist id="A"><Name>Ada</Name></Artist>
    <Artist id="B"><Name>Ben</Name></Artist>
  </Artists>
</Envelope>
""",
            encoding="utf-8",
        )
        adapter = XmlSourceAdapter()

        profile = adapter.inspect_source(source_path, preferred_csv_delimiter=";")

        self.assertTrue(profile.warnings)
        self.assertEqual(profile.headers, ("@code", "Title", "Title/@lang"))
        self.assertEqual(
            profile.rows,
            (
                {"@code": "CAT-1", "Title": "Orbit", "Title/@lang": "en"},
                {"@code": "CAT-2", "Title": "Signal", "Title/@lang": ""},
            ),
        )
        self.assertIs(adapter.select_scope(profile, "missing"), profile)

        artist_scope = next(
            scope for scope, _label in profile.available_scopes if scope.endswith("Artist[1]")
        )
        artist_profile = adapter.select_scope(profile, artist_scope)
        self.assertEqual(artist_profile.headers, ("@id", "Name"))
        self.assertEqual(
            artist_profile.rows,
            ({"@id": "A", "Name": "Ada"}, {"@id": "B", "Name": "Ben"}),
        )

        broken_profile = replace(
            profile,
            adapter_state={
                "candidates": {
                    "broken": {
                        "sample_path": "/Envelope/Records/Missing",
                        "parent_path": "/Envelope/Records",
                        "sample_fields": profile.adapter_state["candidates"][profile.chosen_scope][
                            "sample_fields"
                        ],
                    }
                }
            },
        )
        self.assertIs(adapter.select_scope(broken_profile, "broken"), broken_profile)

        root = ET.fromstring(source_path.read_text(encoding="utf-8"))
        missing_value = _row_from_node(
            root.find("./Records/Record"),
            (
                ConversionTargetField(
                    field_key="missing",
                    display_name="Missing",
                    location="/Envelope/Records/Record/Missing",
                    required_status=REQUIRED_STATUS_OPTIONAL,
                    metadata={"path": "Missing"},
                ),
            ),
        )
        self.assertEqual(missing_value, {"Missing": ""})

        empty_source = self.root / "empty-source.xml"
        empty_source.write_text("<Envelope />", encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "repeat-node candidate"):
            adapter.inspect_source(empty_source)

    def test_database_source_uses_exchange_export_rows(self):
        exchange_service = _FakeExchangeService()
        settings_read_service = _FakeSettingsReadService(
            sena_number="SENA-778899",
            owner_values={
                "party_id": 12,
                "legal_name": "Aeon Cosmowyn Records B.V.",
                "company_name": "Cosmowyn Records",
                "display_name": "Cosmowyn Records",
                "email": "hello@cosmowyn.test",
            },
        )
        service = ConversionService(
            exchange_service=exchange_service,
            settings_read_service=settings_read_service,
        )

        profile = service.inspect_database_tracks([7, 11])

        self.assertEqual(exchange_service.calls, [[7, 11]])
        self.assertEqual(profile.source_mode, "database_tracks")
        self.assertEqual(
            profile.headers,
            (
                "track_id",
                "track_title",
                "catalog_number",
                "pro_number",
                "owner_party_id",
                "owner_legal_name",
                "owner_display_name",
                "owner_artist_name",
                "owner_company_name",
                "owner_first_name",
                "owner_middle_name",
                "owner_last_name",
                "owner_contact_person",
                "owner_email",
                "owner_alternative_email",
                "owner_phone",
                "owner_website",
                "owner_street_name",
                "owner_street_number",
                "owner_address_line1",
                "owner_address_line2",
                "owner_city",
                "owner_region",
                "owner_postal_code",
                "owner_country",
                "owner_bank_account_number",
                "owner_chamber_of_commerce_number",
                "owner_tax_id",
                "owner_vat_number",
                "owner_pro_affiliation",
                "owner_pro_number",
                "owner_ipi_cae",
            ),
        )
        self.assertEqual(profile.rows[0]["track_title"], "Track 7")
        self.assertEqual(profile.rows[0]["pro_number"], "SENA-778899")
        self.assertEqual(profile.rows[0]["owner_party_id"], "12")
        self.assertEqual(profile.rows[0]["owner_legal_name"], "Aeon Cosmowyn Records B.V.")
        self.assertEqual(profile.rows[0]["owner_company_name"], "Cosmowyn Records")
        self.assertEqual(profile.rows[0]["owner_email"], "hello@cosmowyn.test")
        self.assertTrue(any("release-aware" in warning.lower() for warning in profile.warnings))

    def test_json_source_wrapper_variants_and_invalid_payloads(self):
        service = ConversionService()
        rows_path = self.root / "rows.json"
        rows_path.write_text(json.dumps({"rows": [{"catalog": "CAT-1"}]}), encoding="utf-8")
        items_path = self.root / "items.json"
        items_path.write_text(json.dumps({"items": [{"title": "Orbit"}]}), encoding="utf-8")
        nested_path = self.root / "nested.json"
        nested_path.write_text(json.dumps({"payload": [{"artist": "Nova"}]}), encoding="utf-8")
        single_path = self.root / "single.json"
        single_path.write_text(
            json.dumps({"catalog": "CAT-2", "title": "Signal"}),
            encoding="utf-8",
        )
        invalid_path = self.root / "invalid.json"
        invalid_path.write_text(json.dumps("bad payload"), encoding="utf-8")

        self.assertEqual(service.inspect_source_file(rows_path).headers, ("catalog",))
        self.assertEqual(service.inspect_source_file(items_path).headers, ("title",))
        self.assertEqual(service.inspect_source_file(nested_path).headers, ("artist",))
        single_profile = service.inspect_source_file(single_path)
        self.assertEqual(single_profile.rows, ({"catalog": "CAT-2", "title": "Signal"},))

        with self.assertRaisesRegex(ValueError, "JSON source"):
            service.inspect_source_file(invalid_path)
        with self.assertRaisesRegex(ValueError, "Unsupported source file format"):
            service.inspect_source_file(self.root / "source.txt")

    def test_preview_blocks_unmapped_required_targets_and_supports_constants(self):
        service = ConversionService()
        template_profile = ConversionTemplateProfile(
            template_path=self.root / "virtual.csv",
            format_name="csv",
            output_suffix=".csv",
            structure_label="Virtual CSV",
            target_fields=(
                ConversionTargetField(
                    field_key="catalog_number",
                    display_name="Catalog Number*",
                    location="CSV!A1",
                    required_status=REQUIRED_STATUS_REQUIRED,
                ),
                ConversionTargetField(
                    field_key="territory",
                    display_name="Territory",
                    location="CSV!B1",
                    required_status=REQUIRED_STATUS_OPTIONAL,
                ),
            ),
            template_signature="csv|virtual|catalog_number,territory",
        )
        source_profile = ConversionSourceProfile(
            source_mode=SOURCE_MODE_FILE,
            format_name="json",
            source_label="memory",
            headers=("catalog number",),
            rows=({"catalog number": "CAT-55"},),
            preview_rows=({"catalog number": "CAT-55"},),
        )
        session = service.build_session(template_profile, source_profile)

        preview = service.build_preview(session)
        self.assertTrue(preview.blocking_issues)

        suggestions = service.suggest_mapping(session)
        session.mapping_entries = (
            suggestions["catalog_number"],
            ConversionMappingEntry(
                target_field_key="territory",
                target_display_name="Territory",
                mapping_kind=MAPPING_KIND_CONSTANT,
                constant_value="Worldwide",
            ),
        )
        preview = service.build_preview(session)

        self.assertFalse(preview.blocking_issues)
        self.assertEqual(preview.rendered_rows[0], ("CAT-55", "Worldwide"))
        self.assertEqual(preview.mapping_entries[0].mapping_kind, MAPPING_KIND_SOURCE)
        self.assertNotEqual(preview.mapping_entries[0].mapping_kind, MAPPING_KIND_UNMAPPED)

    def test_preview_allows_optional_skip_without_empty_value_warning(self):
        service = ConversionService()
        template_profile = ConversionTemplateProfile(
            template_path=self.root / "virtual.csv",
            format_name="csv",
            output_suffix=".csv",
            structure_label="Virtual CSV",
            target_fields=(
                ConversionTargetField(
                    field_key="catalog_number",
                    display_name="Catalog Number*",
                    location="CSV!A1",
                    required_status=REQUIRED_STATUS_REQUIRED,
                ),
                ConversionTargetField(
                    field_key="territory",
                    display_name="Territory",
                    location="CSV!B1",
                    required_status=REQUIRED_STATUS_OPTIONAL,
                ),
            ),
            template_signature="csv|virtual|catalog_number,territory",
        )
        source_profile = ConversionSourceProfile(
            source_mode=SOURCE_MODE_FILE,
            format_name="json",
            source_label="memory",
            headers=("catalog number",),
            rows=({"catalog number": "CAT-55"},),
            preview_rows=({"catalog number": "CAT-55"},),
        )
        session = service.build_session(template_profile, source_profile)
        suggestions = service.suggest_mapping(session)
        session.mapping_entries = (
            suggestions["catalog_number"],
            ConversionMappingEntry(
                target_field_key="territory",
                target_display_name="Territory",
                mapping_kind=MAPPING_KIND_SKIP,
            ),
        )

        preview = service.build_preview(session)

        self.assertFalse(preview.blocking_issues)
        self.assertEqual(preview.rendered_rows[0], ("CAT-55", ""))
        self.assertNotIn("Territory", " ".join(preview.warnings))
        self.assertEqual(preview.mapping_entries[1].status, "skipped")

    def test_preview_blocks_required_skip(self):
        service = ConversionService()
        template_profile = ConversionTemplateProfile(
            template_path=self.root / "virtual.csv",
            format_name="csv",
            output_suffix=".csv",
            structure_label="Virtual CSV",
            target_fields=(
                ConversionTargetField(
                    field_key="catalog_number",
                    display_name="Catalog Number*",
                    location="CSV!A1",
                    required_status=REQUIRED_STATUS_REQUIRED,
                ),
            ),
            template_signature="csv|virtual|catalog_number",
        )
        source_profile = ConversionSourceProfile(
            source_mode=SOURCE_MODE_FILE,
            format_name="json",
            source_label="memory",
            headers=("catalog number",),
            rows=({"catalog number": "CAT-55"},),
            preview_rows=({"catalog number": "CAT-55"},),
        )
        session = service.build_session(template_profile, source_profile)
        session.mapping_entries = (
            ConversionMappingEntry(
                target_field_key="catalog_number",
                target_display_name="Catalog Number*",
                mapping_kind=MAPPING_KIND_SKIP,
            ),
        )

        preview = service.build_preview(session)

        self.assertIn("Required target 'Catalog Number*' is skipped.", preview.blocking_issues)

    def test_conversion_service_guardrails_for_scopes_preview_export_and_mapping_payloads(self):
        service = ConversionService()
        required_field = ConversionTargetField(
            field_key="catalog_number",
            display_name="Catalog Number*",
            location="CSV!A1",
            required_status=REQUIRED_STATUS_REQUIRED,
        )
        optional_field = ConversionTargetField(
            field_key="title",
            display_name="Title",
            location="CSV!B1",
            required_status=REQUIRED_STATUS_OPTIONAL,
        )
        template_profile = ConversionTemplateProfile(
            template_path=Path("template.csv"),
            format_name="csv",
            output_suffix=".csv",
            structure_label="CSV",
            target_fields=(required_field, optional_field),
            template_signature="csv|test",
        )
        unsupported_profile = replace(template_profile, format_name="unknown")
        source_profile = ConversionSourceProfile(
            source_mode=SOURCE_MODE_FILE,
            format_name="unknown",
            source_label="memory",
            headers=("catalog", "title"),
            rows=({"catalog": "", "title": ""},),
            preview_rows=({"catalog": "", "title": ""},),
        )

        self.assertIs(
            service.select_template_scope(unsupported_profile, "anything"),
            unsupported_profile,
        )
        self.assertIs(service.select_source_scope(source_profile, "anything"), source_profile)

        empty_session = service.build_session(template_profile, source_profile)
        empty_session.included_row_indices = ()
        empty_session.mapping_entries = (
            ConversionMappingEntry(
                target_field_key="catalog_number",
                target_display_name="Catalog Number*",
                mapping_kind=MAPPING_KIND_SOURCE,
                source_field="catalog",
            ),
            ConversionMappingEntry(
                target_field_key="title",
                target_display_name="Title",
                mapping_kind=MAPPING_KIND_SOURCE,
                source_field="title",
            ),
        )
        empty_preview = service.build_preview(empty_session)
        self.assertIn("Select at least one source row to convert.", empty_preview.blocking_issues)

        session = service.build_session(template_profile, source_profile)
        session.mapping_entries = empty_session.mapping_entries
        preview = service.build_preview(session)
        self.assertIn(
            "Required target 'Catalog Number*' resolves empty in one or more rendered rows.",
            preview.blocking_issues,
        )
        self.assertIn(
            "Target 'Title' resolves empty in one or more rendered rows.",
            preview.warnings,
        )

        unsupported_session = service.build_session(unsupported_profile, source_profile)
        with self.assertRaisesRegex(ValueError, "Unsupported template format"):
            service.build_preview(unsupported_session)
        with self.assertRaisesRegex(ValueError, "Unsupported template format"):
            service.export_preview(
                ConversionPreview(
                    template_profile=unsupported_profile,
                    source_profile=source_profile,
                    mapping_entries=(),
                    included_row_indices=(),
                ),
                self.root / "output.dat",
            )

        self.assertEqual(service.deserialize_mapping_entries("{not json", template_profile), ())
        restored_entries = service.deserialize_mapping_entries(
            json.dumps(
                [
                    {"target_field_key": "missing", "mapping_kind": MAPPING_KIND_SOURCE},
                    {
                        "target_field_key": "title",
                        "mapping_kind": MAPPING_KIND_SOURCE,
                        "source_field": "title",
                    },
                ]
            ),
            template_profile,
        )
        self.assertEqual(len(restored_entries), 1)
        self.assertEqual(restored_entries[0].target_field_key, "title")

        with self.assertRaisesRegex(ValueError, "Unsupported template format"):
            service.inspect_template(self.root / "template.txt")

    def test_inspect_template_bytes_ignores_temporary_cleanup_failures(self):
        service = ConversionService()
        template_bytes = b"Catalog Number*,Title\nTEMPLATE,Sample\n"

        with mock.patch(
            "isrc_manager.conversion.service.Path.unlink",
            side_effect=RuntimeError("temp file locked"),
        ):
            profile = service.inspect_template_bytes(
                "saved-template.csv",
                template_bytes,
                source_label="Saved Template",
                source_path="/original/template.csv",
            )

        self.assertEqual(profile.template_path, Path("saved-template.csv"))
        self.assertEqual(profile.template_bytes, template_bytes)
        self.assertEqual(profile.adapter_state["source_label"], "Saved Template")
        self.assertEqual(profile.adapter_state["source_path"], "/original/template.csv")

    def test_saved_template_store_round_trips_template_bytes_and_mapping(self):
        conn = sqlite3.connect(":memory:")
        try:
            service = ConversionService()
            store = ConversionTemplateStoreService(conn)
            template_path = self._write_csv(
                "template.csv",
                "Catalog Number*;Title;Year\nTEMPLATE;Sample;2024\n",
            )
            source_path = self._write_csv(
                "source.csv",
                "catalog_number,title,release_date\nCAT-001,Orbit,2025-04-06\n",
            )

            template_profile = service.inspect_template(template_path)
            source_profile = service.inspect_source_file(source_path)
            session = service.build_session(template_profile, source_profile)
            suggestions = service.suggest_mapping(session)
            session.mapping_entries = (
                suggestions["catalog_number"],
                suggestions["title"],
                ConversionMappingEntry(
                    target_field_key="year",
                    target_display_name="Year",
                    mapping_kind=MAPPING_KIND_CONSTANT,
                    constant_value="2030",
                ),
            )
            mapping_payload = service.serialize_mapping_entries(session.mapping_entries)

            stored = store.save_template(
                name="Producer Export",
                template_profile=template_profile,
                mapping_payload=mapping_payload,
                source_mode=SOURCE_MODE_FILE,
            )

            listed = store.list_saved_templates()
            self.assertEqual(len(listed), 1)
            self.assertEqual(listed[0].name, "Producer Export")
            self.assertEqual(listed[0].source_mode, SOURCE_MODE_FILE)
            self.assertIsNone(listed[0].template_bytes)

            loaded = store.load_saved_template(stored.id)
            self.assertIsNotNone(loaded.template_bytes)
            self.assertEqual(loaded.mapping_payload, mapping_payload)

            loaded_profile = service.inspect_template_bytes(
                loaded.filename,
                loaded.template_bytes or b"",
                source_label=f"Saved in profile: {loaded.name}",
                source_path=loaded.source_path,
            )
            loaded_entries = service.deserialize_mapping_entries(
                loaded.mapping_payload,
                loaded_profile,
            )

            self.assertEqual(len(loaded_entries), 3)
            self.assertEqual(loaded_entries[2].mapping_kind, MAPPING_KIND_CONSTANT)
            self.assertEqual(loaded_entries[2].constant_value, "2030")
            self.assertEqual(loaded_profile.template_bytes, loaded.template_bytes)
        finally:
            conn.close()

    def test_saved_template_store_validation_memory_templates_and_update_paths(self):
        conn = sqlite3.connect(":memory:")
        try:
            store = ConversionTemplateStoreService(conn)
            memory_profile = ConversionTemplateProfile(
                template_path=Path(""),
                format_name="",
                output_suffix=".csv",
                structure_label="Memory template",
                target_fields=(),
                template_signature="memory",
                template_bytes=b"catalog,title\nCAT-1,Orbit\n",
                chosen_scope="tracks",
                adapter_state={"source_path": " /imports/source.csv "},
            )

            with self.assertRaisesRegex(ValueError, "profile template name"):
                store.save_template(name="  ", template_profile=memory_profile)
            with self.assertRaisesRegex(ValueError, "no longer exists"):
                store.load_saved_template(999)

            stored = store.save_template(
                name="Memory Template",
                template_profile=memory_profile,
                mapping_payload='{"field": "catalog"}',
                source_mode=" file ",
            )

            self.assertEqual(stored.filename, "conversion-template")
            self.assertEqual(stored.format_name, "unknown")
            self.assertEqual(stored.source_path, "/imports/source.csv")
            self.assertEqual(stored.chosen_scope, "tracks")
            self.assertEqual(stored.source_mode, "file")
            self.assertEqual(stored.size_bytes, len(memory_profile.template_bytes or b""))

            updated_profile = replace(
                memory_profile,
                template_path=Path("updated-template.csv"),
                template_bytes=b"updated\n",
                adapter_state={},
            )
            updated = store.save_template(
                name="memory template",
                template_profile=updated_profile,
            )

            self.assertEqual(updated.id, stored.id)
            self.assertEqual(updated.filename, "updated-template.csv")
            self.assertEqual(updated.source_path, "")
            loaded = store.load_saved_template(updated.id)
            self.assertEqual(loaded.template_bytes, b"updated\n")
        finally:
            conn.close()

    def test_saved_template_store_reads_file_templates_and_reports_missing_files(self):
        conn = sqlite3.connect(":memory:")
        try:
            store = ConversionTemplateStoreService(conn)
            template_path = self._write_csv("disk-template.csv", "catalog,title\nCAT-1,Orbit\n")
            disk_profile = ConversionTemplateProfile(
                template_path=template_path,
                format_name="csv",
                output_suffix=".csv",
                structure_label="CSV template",
                target_fields=(),
                template_signature="disk",
            )

            stored = store.save_template(name="Disk Template", template_profile=disk_profile)
            self.assertEqual(stored.source_path, str(template_path))
            self.assertEqual(stored.size_bytes, template_path.stat().st_size)

            missing_profile = replace(disk_profile, template_path=self.root / "missing.csv")
            with self.assertRaisesRegex(ValueError, "no longer available"):
                store.save_template(name="Missing Template", template_profile=missing_profile)
        finally:
            conn.close()

    def test_saved_template_store_reports_failed_reload_after_insert(self):
        class _Cursor:
            def __init__(self, row=None):
                self._row = row

            def fetchone(self):
                return self._row

        class _FakeConnection:
            def __enter__(self):
                return self

            def __exit__(self, _exc_type, _exc, _tb):
                return False

            def execute(self, sql, _params=()):
                if "WHERE name = ?" in sql:
                    return _Cursor(row=None)
                return _Cursor()

        store = ConversionTemplateStoreService(_FakeConnection())
        profile = ConversionTemplateProfile(
            template_path=Path("template.csv"),
            format_name="csv",
            output_suffix=".csv",
            structure_label="CSV template",
            target_fields=(),
            template_signature="memory",
            template_bytes=b"catalog,title\n",
        )

        with self.assertRaisesRegex(RuntimeError, "Failed to store"):
            store.save_template(name="Defensive Reload", template_profile=profile)


if __name__ == "__main__":
    unittest.main()
