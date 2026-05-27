import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from isrc_manager.services import GS1TemplateVerificationError, GS1TemplateVerificationService
from isrc_manager.services.gs1_mapping import (
    COMMON_MARKET_CHOICES,
    localize_export_value,
    resolve_header_row,
)

HEADERS = [
    "GS1 Artikelcode (GTIN)",
    "Status",
    "Productclassificatie",
    "Gaat naar de consument",
    "Verpakkings type",
    "Landen of Regio's",
    "Productomschrijving (max 300 tekens)",
    "Taal",
    "Merk",
    "Submerk",
    "Aantal",
    "Eenheid",
    "Afbeelding (max 500 tekens)",
]


def build_verified_workbook(path: Path, *, include_placeholder_sheet: bool = True):
    workbook = Workbook()
    reference_sheet = workbook.active
    reference_sheet.title = "Reference Data"
    reference_sheet["A1"] = "Ampul"
    reference_sheet["A2"] = "Bag-In-Box"
    reference_sheet["B1"] = "Netherlands"
    reference_sheet["B2"] = "United States"
    reference_sheet["C1"] = "Dutch"
    reference_sheet["C2"] = "English"
    reference_sheet["D1"] = "Each"
    reference_sheet["D2"] = "Box"
    reference_sheet["E1"] = "Audio"
    reference_sheet["E2"] = "Music"
    reference_sheet["F1"] = "Concept"
    reference_sheet["F2"] = "Active"
    reference_sheet["G1"] = "Yes"
    reference_sheet["G2"] = "No"
    instruction_sheet = workbook.create_sheet("Instructions")
    instruction_sheet["A1"] = "GS1 article code (GTIN)"
    instruction_sheet["A2"] = "Upload the workbook after filling the fields."
    if include_placeholder_sheet:
        placeholder_sheet = workbook.create_sheet("{ContractNr}")
        placeholder_sheet.append(HEADERS)
    target_sheet = workbook.create_sheet("10070050")
    target_sheet.append(HEADERS)
    target_sheet["I2"] = "Orbit Label"
    target_sheet["J2"] = "Digital Series"
    for cell_range, formula in (
        ("B2:B200", "'Reference Data'!$F$1:$F$2"),
        ("C2:C200", "'Reference Data'!$E$1:$E$2"),
        ("D2:D200", "'Reference Data'!$G$1:$G$2"),
        ("E2:E200", "'Reference Data'!$A$1:$A$2"),
        ("F2:F200", "'Reference Data'!$B$1:$B$2"),
        ("H2:H200", "'Reference Data'!$C$1:$C$2"),
        ("L2:L200", "'Reference Data'!$D$1:$D$2"),
    ):
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        target_sheet.add_data_validation(validation)
        validation.add(cell_range)
    workbook.save(path)


class GS1HeaderMappingTests(unittest.TestCase):
    def test_resolve_header_row_maps_canonical_fields(self):
        column_map, matched_headers, total_score = resolve_header_row(HEADERS)

        self.assertEqual(column_map["gtin_request_number"], 1)
        self.assertEqual(column_map["consumer_unit_flag"], 4)
        self.assertEqual(column_map["product_description"], 7)
        self.assertEqual(column_map["image_url"], 13)
        self.assertEqual(matched_headers["brand"], "Merk")
        self.assertGreater(total_score, 20.0)

    def test_localize_export_value_handles_known_dutch_variants(self):
        self.assertEqual(localize_export_value("consumer_unit_flag", True, "nl"), "Ja")
        self.assertEqual(localize_export_value("status", "Active", "nl"), "Actief")
        self.assertEqual(localize_export_value("language", "English", "nl"), "Engels")

    def test_common_market_choices_include_non_country_specific_gs1_markets(self):
        self.assertIn("Global Market", COMMON_MARKET_CHOICES)
        self.assertIn("European Union", COMMON_MARKET_CHOICES)
        self.assertIn("Non-EU", COMMON_MARKET_CHOICES)
        self.assertIn("Developing Countries Support", COMMON_MARKET_CHOICES)
        self.assertIn("Europese Unie", COMMON_MARKET_CHOICES)
        self.assertIn("Niet EU", COMMON_MARKET_CHOICES)

    def test_localize_export_value_handles_non_country_specific_target_markets(self):
        self.assertEqual(
            localize_export_value("target_market", "Global Market", "nl"), "Global Market"
        )
        self.assertEqual(
            localize_export_value("target_market", "European Union", "nl"), "Europese Unie"
        )
        self.assertEqual(localize_export_value("target_market", "Non-EU", "nl"), "Niet EU")
        self.assertEqual(
            localize_export_value("target_market", "Developing Countries Support", "nl"),
            "Ontwikkelingslanden ondersteuning",
        )


class GS1TemplateVerificationServiceTests(unittest.TestCase):
    def setUp(self):
        self.service = GS1TemplateVerificationService()
        self.tmpdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tmpdir.name)

    def tearDown(self):
        self.tmpdir.cleanup()

    def test_verify_selects_actual_contract_sheet(self):
        workbook_path = self.root / "official_template.xlsx"
        build_verified_workbook(workbook_path)

        result = self.service.verify(workbook_path)

        self.assertEqual(result.sheet_name, "10070050")
        self.assertEqual(result.available_sheet_names, ("10070050",))
        self.assertEqual(result.header_row, 1)
        self.assertEqual(result.column_map["gtin_request_number"], 1)
        self.assertEqual(result.column_map["quantity"], 11)
        self.assertEqual(result.locale_hint, "nl")
        self.assertEqual(result.field_options["packaging_type"], ("Ampul", "Bag-In-Box"))
        self.assertEqual(result.field_options["target_market"], ("Netherlands", "United States"))
        self.assertEqual(result.field_options["language"], ("Dutch", "English"))
        self.assertEqual(result.field_options["product_classification"], ("Audio", "Music"))
        self.assertEqual(result.field_options["brand"], ("Orbit Label",))
        self.assertEqual(result.field_options["subbrand"], ("Digital Series",))

    def test_verify_rejects_arbitrary_workbook(self):
        workbook_path = self.root / "random.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["First Name", "Last Name", "Email"])
        workbook.save(workbook_path)

        with self.assertRaises(GS1TemplateVerificationError):
            self.service.verify(workbook_path)

    def test_verify_rejects_missing_and_unsupported_paths_before_loading_workbook(self):
        with self.assertRaisesRegex(GS1TemplateVerificationError, "was not found"):
            self.service.verify(self.root / "missing.xlsx")

        unsupported = self.root / "template.csv"
        unsupported.write_text("not an excel workbook", encoding="utf-8")
        with self.assertRaisesRegex(GS1TemplateVerificationError, "supported Excel workbook"):
            self.service.verify(unsupported)

    def test_verify_rejects_workbook_with_missing_required_headers(self):
        workbook_path = self.root / "missing_headers.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(
            [
                "GS1 Artikelcode (GTIN)",
                "Status",
                "Productclassificatie",
                "Gaat naar de consument",
                "Verpakkings type",
                "Landen of Regio's",
                "Productomschrijving (max 300 tekens)",
                "Taal",
                "Merk",
            ]
        )
        workbook.save(workbook_path)

        with self.assertRaises(GS1TemplateVerificationError):
            self.service.verify(workbook_path)

    def test_helper_scoring_formula_resolution_and_column_value_fallbacks(self):
        loaded = Workbook()
        loaded.remove(loaded.active)
        ref = loaded.create_sheet("Reference Data")
        ref["A1"] = "One"
        ref["A2"] = "Two"
        sheet = loaded.create_sheet("Product Upload")
        sheet.append(HEADERS)
        sheet["I2"] = "Brand A"
        sheet["I3"] = "Brand B"

        self.assertEqual(self.service._sheet_name_priority("Instructions"), -2.0)
        self.assertEqual(self.service._sheet_name_priority("{ContractNr}"), -1.0)
        self.assertGreater(self.service._sheet_name_priority("Product Upload"), 0)
        self.assertEqual(self.service._field_name_for_sqref("$I$2:$I$200", {9: "brand"}), "brand")
        self.assertEqual(self.service._field_name_for_sqref("not-a-cell", {9: "brand"}), "")
        self.assertEqual(
            self.service._resolve_validation_formula_values(loaded, '"A, B, A"'),
            ["A", "B"],
        )
        self.assertEqual(
            self.service._resolve_validation_formula_values(
                loaded,
                "'Reference Data'!$A$1:$A$2",
            ),
            ["One", "Two"],
        )
        self.assertEqual(
            self.service._resolve_validation_formula_values(loaded, "MissingName"),
            [],
        )
        self.assertEqual(
            self.service._read_cell_range_values(loaded, "Missing", "$A$1:$A$2"),
            [],
        )
        self.assertEqual(
            self.service._read_cell_range_values(loaded, "Reference Data", "not-a-range"),
            [],
        )
        self.assertEqual(
            self.service._collect_existing_column_values(sheet, 9, start_row=2),
            ["Brand A", "Brand B"],
        )
        self.assertEqual(
            self.service._dedupe_preserve_order([" One ", "", "Two", "One"]),
            ["One", "Two"],
        )

    def test_xml_validation_helpers_extract_sheet_options(self):
        workbook_path = self.root / "validation-options.xlsx"
        build_verified_workbook(workbook_path)
        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=False,
        )
        self.addCleanup(workbook.close)

        worksheet_xml_path = self.service._resolve_sheet_xml_path(workbook_path, "10070050")
        entries = self.service._read_validation_entries(workbook_path, worksheet_xml_path)
        options = self.service._validation_options_from_sheet_xml(
            workbook_path,
            workbook,
            {
                "status": 2,
                "product_classification": 3,
                "consumer_unit_flag": 4,
                "packaging_type": 5,
                "target_market": 6,
                "language": 8,
                "unit": 12,
            },
            "10070050",
        )

        self.assertTrue(worksheet_xml_path.endswith(".xml"))
        self.assertTrue(any("B2:B200" in sqref for sqref, _formula in entries))
        self.assertEqual(options["status"], ["Concept", "Active"])
        self.assertEqual(options["packaging_type"], ["Ampul", "Bag-In-Box"])
        self.assertEqual(
            self.service._validation_options_from_sheet_xml(
                workbook_path / "missing",
                workbook,
                {"status": 2},
                "10070050",
            ),
            {},
        )
        self.assertEqual(self.service._resolve_sheet_xml_path(workbook_path, "Missing"), "")


if __name__ == "__main__":
    unittest.main()
